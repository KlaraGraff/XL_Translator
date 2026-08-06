"""补丁式 xlsx 写入器的契约测试。

重点验证三件事：
  1. 未被修改的部件在输出包里逐字节保持原样（尤其是 WPS cellimages.xml、
     Excel 富数据部件、media 图片）；
  2. 悬浮图片在行高调整后不被拉伸——twoCellAnchor 变成带原渲染尺寸的 oneCellAnchor；
  3. 译文、复核标色、行高、原文分表克隆的语义与旧路径一致。
"""
from __future__ import annotations

import hashlib
import shutil
import tempfile
import unittest
import zipfile
from pathlib import Path

from lxml import etree
from openpyxl import Workbook, load_workbook

from core.mixed_language import MIXED_MARK_FOREIGN_NOISE, MIXED_MARK_UNRESOLVED
from core.bilingual_writer import write_bilingual_file
from core.xlsx_patcher import (
    EMU_PER_POINT,
    NS_MAIN,
    NS_XDR,
    _column_letter,
    _SheetGeometry,
    fix_drawing_anchors,
)

CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

WPS_CELLIMAGES = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    b'<etc:cellImages xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData">'
    b"<etc:cellImage>ID_1</etc:cellImage></etc:cellImages>"
)
RICH_VALUE_PART = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    b'<rvData xmlns="http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"'
    b' count="1"><rv s="0"><v>0</v></rv></rvData>'
)

PNG_BYTES = bytes.fromhex(
    "89504e470d0a1a0a0000000d494844520000000100000001080600000"
    "01f15c4890000000a49444154789c6360000002000100ffff03000006"
    "000557bfabd40000000049454e44ae426082"
)


def _sha(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _part_hashes(path: Path) -> dict[str, str]:
    with zipfile.ZipFile(path) as archive:
        return {name: _sha(archive.read(name)) for name in archive.namelist()}


def _read_part(path: Path, name: str) -> bytes:
    with zipfile.ZipFile(path) as archive:
        return archive.read(name)


class XlsxPatcherFixture:
    """手工往 openpyxl 生成的工作簿里塞进 openpyxl 自己不认识的部件。"""

    @staticmethod
    def build(root: Path) -> Path:
        base = root / "base.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "报价"
        sheet["A1"] = "施工内容"
        sheet["A2"] = "很长很长的施工说明文字需要换行才放得下所以行高会被自动调整"
        sheet["A3"] = "配电箱"
        sheet["B1"] = "=SUM(C1:C2)"
        sheet["C1"] = 123
        second = workbook.create_sheet("附表")
        second["A1"] = "配电箱"
        workbook.save(base)
        workbook.close()

        target = root / "fixture.xlsx"
        XlsxPatcherFixture._inject(base, target)
        return target

    @staticmethod
    def _inject(source: Path, target: Path) -> None:
        with zipfile.ZipFile(source) as archive:
            parts = {name: archive.read(name) for name in archive.namelist()}

        # ── 分表 1：DISPIMG 公式 + 富数据单元格 + 共享字符串引用 ──────────
        sheet_xml = etree.fromstring(parts["xl/worksheets/sheet1.xml"])
        sheet_data = sheet_xml.find(f"{{{NS_MAIN}}}sheetData")
        row = etree.SubElement(sheet_data, f"{{{NS_MAIN}}}row")
        row.set("r", "5")
        dispimg = etree.SubElement(row, f"{{{NS_MAIN}}}c")
        dispimg.set("r", "A5")
        dispimg.set("t", "str")
        etree.SubElement(dispimg, f"{{{NS_MAIN}}}f").text = '_xlfn.DISPIMG("ID_1",1)'
        etree.SubElement(dispimg, f"{{{NS_MAIN}}}v").text = "配电箱"

        rich = etree.SubElement(row, f"{{{NS_MAIN}}}c")
        rich.set("r", "B5")
        rich.set("vm", "1")
        rich.set("t", "s")
        etree.SubElement(rich, f"{{{NS_MAIN}}}v").text = "0"

        # 共享字符串引用（openpyxl 保存时只会写 inlineStr，这里手工补一个）
        shared = etree.SubElement(row, f"{{{NS_MAIN}}}c")
        shared.set("r", "C5")
        shared.set("t", "s")
        etree.SubElement(shared, f"{{{NS_MAIN}}}v").text = "0"

        # ── 悬浮图片：twoCellAnchor 跨 1-4 行 ───────────────────────────
        drawing_el = etree.SubElement(sheet_xml, f"{{{NS_MAIN}}}drawing")
        drawing_el.set(f"{{{REL_NS}}}id", "rIdDraw")
        parts["xl/worksheets/sheet1.xml"] = etree.tostring(
            sheet_xml, xml_declaration=True, encoding="UTF-8", standalone=True
        )

        parts["xl/worksheets/_rels/sheet1.xml.rels"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{PKG_REL_NS}">'
            f'<Relationship Id="rIdDraw" Type="{REL_NS}/drawing"'
            f' Target="../drawings/drawing1.xml"/></Relationships>'
        ).encode()

        parts["xl/drawings/drawing1.xml"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<xdr:wsDr xmlns:xdr="{NS_XDR}"'
            f' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
            f' xmlns:r="{REL_NS}">'
            f'<xdr:twoCellAnchor editAs="oneCell">'
            f"<xdr:from><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>0</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f"<xdr:to><xdr:col>5</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>3</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
            f"<xdr:pic><xdr:nvPicPr>"
            f'<xdr:cNvPr id="1" name="Picture 1"/><xdr:cNvPicPr/></xdr:nvPicPr>'
            f'<xdr:blipFill><a:blip r:embed="rIdImg"/>'
            f"<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>"
            f'<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>'
            f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr></xdr:pic>'
            f"<xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>"
        ).encode()

        parts["xl/drawings/_rels/drawing1.xml.rels"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{PKG_REL_NS}">'
            f'<Relationship Id="rIdImg" Type="{REL_NS}/image"'
            f' Target="../media/image1.png"/></Relationships>'
        ).encode()
        parts["xl/media/image1.png"] = PNG_BYTES

        # ── 共享字符串表 ───────────────────────────────────────────────
        parts["xl/sharedStrings.xml"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<sst xmlns="{NS_MAIN}" count="2" uniqueCount="1">'
            f"<si><t>配电箱</t></si></sst>"
        ).encode()

        # ── WPS cellimages + Excel 富数据部件 ─────────────────────────
        parts["xl/cellimages.xml"] = WPS_CELLIMAGES
        parts["xl/richData/rdrichvalue.xml"] = RICH_VALUE_PART

        workbook_rels = etree.fromstring(parts["xl/_rels/workbook.xml.rels"])
        for rel_id, rel_type, rel_target in (
            ("rIdShared", f"{REL_NS}/sharedStrings", "sharedStrings.xml"),
            (
                "rIdCellImages",
                "http://www.wps.cn/officeDocument/2020/cellImage",
                "cellimages.xml",
            ),
            (
                "rIdRichValue",
                "http://schemas.microsoft.com/office/2017/06/relationships/rdrichvalue",
                "richData/rdrichvalue.xml",
            ),
        ):
            rel = etree.SubElement(workbook_rels, f"{{{PKG_REL_NS}}}Relationship")
            rel.set("Id", rel_id)
            rel.set("Type", rel_type)
            rel.set("Target", rel_target)
        parts["xl/_rels/workbook.xml.rels"] = etree.tostring(
            workbook_rels, xml_declaration=True, encoding="UTF-8", standalone=True
        )

        content_types = etree.fromstring(parts["[Content_Types].xml"])
        for part_name, content_type in (
            (
                "/xl/sharedStrings.xml",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml",
            ),
            ("/xl/drawings/drawing1.xml", "application/vnd.openxmlformats-officedocument.drawing+xml"),
            ("/xl/cellimages.xml", "application/vnd.wps-officedocument.cellimage+xml"),
            ("/xl/richData/rdrichvalue.xml", "application/vnd.ms-excel.rdrichvalue+xml"),
        ):
            override = etree.SubElement(content_types, f"{{{CT_NS}}}Override")
            override.set("PartName", part_name)
            override.set("ContentType", content_type)
        default = etree.SubElement(content_types, f"{{{CT_NS}}}Default")
        default.set("Extension", "png")
        default.set("ContentType", "image/png")
        parts["[Content_Types].xml"] = etree.tostring(
            content_types, xml_declaration=True, encoding="UTF-8", standalone=True
        )

        with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as archive:
            for name, data in parts.items():
                archive.writestr(name, data)


class XlsxPatcherTests(unittest.TestCase):
    TRANSLATIONS = {
        "施工内容": "Construction scope",
        "很长很长的施工说明文字需要换行才放得下所以行高会被自动调整": (
            "A very long construction note that needs wrapping"
        ),
        "配电箱": "Distribution box",
    }

    def setUp(self) -> None:
        self._temp = tempfile.TemporaryDirectory()
        self.root = Path(self._temp.name)
        self.fixture = XlsxPatcherFixture.build(self.root)
        self.before = _part_hashes(self.fixture)

    def tearDown(self) -> None:
        self._temp.cleanup()

    def _write(self, **overrides) -> Path:
        kwargs = dict(
            source_path=self.fixture,
            output_dir=self.root / "out",
            translations=dict(self.TRANSLATIONS),
            target_lang="en",
            source_lang="zh",
            keep_original_sheets=False,
            formula_display_value_backfill=False,
            enable_print_guard=False,
        )
        kwargs.update(overrides)
        return write_bilingual_file(**kwargs)

    # ── 部件保真 ──────────────────────────────────────────────────────────
    def test_untouched_parts_are_byte_identical(self) -> None:
        output = self._write()
        after = _part_hashes(output)

        self.assertEqual(set(self.before) - set(after), set(), "输出包丢失了原有部件")

        expected_modified = {
            "xl/worksheets/sheet1.xml",
            "xl/worksheets/sheet2.xml",
            "xl/drawings/drawing1.xml",
            "xl/styles.xml",
        }
        actually_modified = {
            name for name, digest in self.before.items() if after[name] != digest
        }
        self.assertTrue(
            actually_modified <= expected_modified,
            f"意外重写了部件：{sorted(actually_modified - expected_modified)}",
        )
        # 未修改部件必须逐字节一致
        for name in set(self.before) - expected_modified:
            self.assertEqual(after[name], self.before[name], f"部件被改动：{name}")

    def test_embedded_image_parts_survive(self) -> None:
        output = self._write()
        self.assertEqual(_read_part(output, "xl/cellimages.xml"), WPS_CELLIMAGES)
        self.assertEqual(_read_part(output, "xl/richData/rdrichvalue.xml"), RICH_VALUE_PART)
        self.assertEqual(_read_part(output, "xl/media/image1.png"), PNG_BYTES)
        self.assertEqual(
            _read_part(output, "xl/sharedStrings.xml"),
            _read_part(self.fixture, "xl/sharedStrings.xml"),
            "回填走 inlineStr，共享字符串表不应被改写",
        )

    def test_dispimg_and_rich_data_cells_are_untouched(self) -> None:
        output = self._write(formula_display_value_backfill=True)
        root = etree.fromstring(_read_part(output, "xl/worksheets/sheet1.xml"))
        cells = {
            cell.get("r"): cell
            for cell in root.iter(f"{{{NS_MAIN}}}c")
        }

        dispimg = cells["A5"]
        self.assertEqual(
            dispimg.findtext(f"{{{NS_MAIN}}}f"), '_xlfn.DISPIMG("ID_1",1)'
        )
        self.assertEqual(dispimg.findtext(f"{{{NS_MAIN}}}v"), "配电箱")

        rich = cells["B5"]
        self.assertEqual(rich.get("vm"), "1")
        self.assertEqual(rich.get("t"), "s")
        self.assertEqual(rich.findtext(f"{{{NS_MAIN}}}v"), "0")

        # 同样的原文在普通共享字符串单元格里必须照常翻译
        self.assertEqual(cells["C5"].get("t"), "inlineStr")
        self.assertEqual(
            "".join(cells["C5"].itertext()).strip(), "配电箱\nDistribution box"
        )

    # ── 悬浮图片防变形 ────────────────────────────────────────────────────
    def test_floating_image_keeps_original_rendered_size(self) -> None:
        source_sheet = etree.fromstring(_read_part(self.fixture, "xl/worksheets/sheet1.xml"))
        original_heights = {
            int(row.get("r")): float(row.get("ht") or 15.0)
            for row in source_sheet.iter(f"{{{NS_MAIN}}}row")
        }
        expected_cy = int(
            round(sum(original_heights.get(index, 15.0) for index in (1, 2, 3)) * EMU_PER_POINT)
        )

        output = self._write()
        drawing = etree.fromstring(_read_part(output, "xl/drawings/drawing1.xml"))

        self.assertIsNone(drawing.find(f"{{{NS_XDR}}}twoCellAnchor"))
        anchor = drawing.find(f"{{{NS_XDR}}}oneCellAnchor")
        self.assertIsNotNone(anchor, "受影响的悬浮图片应改为 oneCellAnchor")

        children = [etree.QName(child).localname for child in anchor]
        self.assertEqual(children[:2], ["from", "ext"])
        self.assertIn("clientData", children)

        ext = anchor.find(f"{{{NS_XDR}}}ext")
        self.assertEqual(int(ext.get("cy")), expected_cy)
        self.assertGreater(int(ext.get("cx")), 0)

        # 行高确实变了，否则这条断言没有意义
        out_sheet = etree.fromstring(_read_part(output, "xl/worksheets/sheet1.xml"))
        row2 = next(
            row for row in out_sheet.iter(f"{{{NS_MAIN}}}row") if row.get("r") == "2"
        )
        self.assertEqual(row2.get("customHeight"), "1")
        self.assertGreater(float(row2.get("ht")), original_heights.get(2, 15.0))

    def test_locked_row_height_leaves_anchor_alone(self) -> None:
        output = self._write(lock_row_height=True)
        self.assertEqual(
            _read_part(output, "xl/drawings/drawing1.xml"),
            _read_part(self.fixture, "xl/drawings/drawing1.xml"),
            "锁定行高时行高不变，图片锚点不该被动",
        )

    # ── 回填 / 标色 / 行高语义 ────────────────────────────────────────────
    def test_translations_and_review_marks(self) -> None:
        review_positions: list[dict[str, str]] = []
        output = self._write(
            translations={"施工内容": "Construction scope", "配电箱": "配电箱"},
            review_marks={"施工内容": MIXED_MARK_FOREIGN_NOISE},
            review_positions=review_positions,
        )

        workbook = load_workbook(output)
        try:
            sheet = workbook["报价"]
            self.assertEqual(sheet["A1"].value, "施工内容\nConstruction scope")
            self.assertTrue(sheet["A1"].alignment.wrap_text)
            self.assertEqual(sheet["A1"].fill.fill_type, "solid")
            # 译文与原文相同 → 自动标成未解决
            self.assertEqual(sheet["A3"].value, "配电箱")
            self.assertEqual(sheet["A3"].fill.fill_type, "solid")
            # 公式单元格原样保留
            self.assertEqual(sheet["B1"].value, "=SUM(C1:C2)")
            self.assertEqual(sheet["C1"].value, 123)
        finally:
            workbook.close()

        self.assertEqual(
            [(item["worksheet"], item["cell"], item["action"]) for item in review_positions],
            [
                ("报价", "A1", "marked_fill"),
                ("报价", "A3", "marked_fill"),
                ("报价", "C5", "marked_fill"),
                ("附表", "A1", "marked_fill"),
            ],
        )

    def test_existing_fill_policy_skip_preserves_fill(self) -> None:
        review_positions: list[dict[str, str]] = []
        source = self.root / "filled.xlsx"
        shutil.copy2(self.fixture, source)
        workbook = load_workbook(source)
        from openpyxl.styles import PatternFill

        workbook["报价"]["A1"].fill = PatternFill(fill_type="solid", fgColor="FF92D050")
        workbook.save(source)
        workbook.close()

        output = self._write(
            source_path=source,
            review_marks={"施工内容": MIXED_MARK_UNRESOLVED},
            existing_fill_policy="skip",
            review_positions=review_positions,
        )

        written = load_workbook(output)
        try:
            self.assertTrue(str(written["报价"]["A1"].fill.fgColor.rgb).endswith("92D050"))
        finally:
            written.close()
        self.assertIn(
            {
                "worksheet": "报价",
                "cell": "A1",
                "category": MIXED_MARK_UNRESOLVED,
                "action": "preserved_existing_fill",
            },
            review_positions,
        )

    # ── 原文分表克隆 ──────────────────────────────────────────────────────
    def test_keep_original_sheets_clones_images_and_content(self) -> None:
        output = self._write(keep_original_sheets=True)

        workbook = load_workbook(output)
        try:
            self.assertEqual(
                workbook.sheetnames, ["报价", "附表", "报价_原文", "附表_原文"]
            )
            self.assertEqual(workbook["报价_原文"]["A1"].value, "施工内容")
            self.assertEqual(workbook["报价"]["A1"].value, "施工内容\nConstruction scope")
        finally:
            workbook.close()

        with zipfile.ZipFile(output) as archive:
            names = set(archive.namelist())
            content_types = archive.read("[Content_Types].xml").decode()
            clone_rels = None
            for name in names:
                if name.startswith("xl/worksheets/_rels/") and "sheet1" not in name:
                    clone_rels = archive.read(name).decode()

        self.assertIn("xl/worksheets/sheet3.xml", names)
        self.assertIn("xl/drawings/drawing2.xml", names)
        self.assertIn("xl/drawings/_rels/drawing2.xml.rels", names)
        self.assertIn("/xl/worksheets/sheet3.xml", content_types)
        self.assertIn("/xl/drawings/drawing2.xml", content_types)
        self.assertIsNotNone(clone_rels)
        self.assertIn("drawing2.xml", clone_rels)

        # 克隆分表的图片保持原始 twoCellAnchor（原文分表行高也没变）
        self.assertEqual(
            _read_part(output, "xl/drawings/drawing2.xml"),
            _read_part(self.fixture, "xl/drawings/drawing1.xml"),
        )

    def test_long_sheet_name_clone_is_truncated_and_unique(self) -> None:
        source = self.root / "long.xlsx"
        workbook = Workbook()
        workbook.active.title = "A" * 30
        workbook.active["A1"] = "施工内容"
        workbook.create_sheet("B" * 30)["A1"] = "施工内容"
        workbook.save(source)
        workbook.close()

        output = self._write(source_path=source, keep_original_sheets=True)
        written = load_workbook(output)
        try:
            self.assertEqual(len(written.sheetnames), 4)
            for name in written.sheetnames:
                self.assertLessEqual(len(name), 31)
            self.assertEqual(len(set(written.sheetnames)), 4)
        finally:
            written.close()


# ══════════════════════════════════════════════════════════════════════════════
# 悬浮图片锚点冻结（整表 autofit / absoluteAnchor 漂移）
# ══════════════════════════════════════════════════════════════════════════════
# 固定几何，方便在测试里手算期望值，不依赖实现里的换算函数：
ROW_HEIGHT_PT = 20.0
ROW_EMU = 254000          # round(20pt × 12700 EMU/pt)
COL_WIDTH_CHARS = 10.0
COL_EMU = 666750          # int((256×10 + int(128/7)) / 256 × 7) = 70 像素 × 9525
ABS_EXT_CX = 1_000_000
ABS_EXT_CY = 500_000

_LONG_TEXT = "很长很长的施工说明文字需要换行才放得下所以行高会被自动调整"


def _shape_xml(name: str) -> str:
    return (
        f"<xdr:pic><xdr:nvPicPr>"
        f'<xdr:cNvPr id="1" name="{name}"/><xdr:cNvPicPr/></xdr:nvPicPr>'
        f'<xdr:blipFill><a:blip r:embed="rIdImg"/>'
        f"<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>"
        f'<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>'
        f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr></xdr:pic>'
        f"<xdr:clientData/>"
    )


def _two_cell(name: str, from_row: int, to_row: int, *, edit_as: str = "oneCell") -> str:
    """0-based 行下标；列固定跨 0→2（两列宽）。"""
    return (
        f'<xdr:twoCellAnchor editAs="{edit_as}">'
        f"<xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{from_row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
        f"<xdr:to><xdr:col>2</xdr:col><xdr:colOff>0</xdr:colOff>"
        f"<xdr:row>{to_row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
        f"{_shape_xml(name)}</xdr:twoCellAnchor>"
    )


def _absolute(name: str, x_emu: int, y_emu: int) -> str:
    return (
        f"<xdr:absoluteAnchor>"
        f'<xdr:pos x="{x_emu}" y="{y_emu}"/>'
        f'<xdr:ext cx="{ABS_EXT_CX}" cy="{ABS_EXT_CY}"/>'
        f"{_shape_xml(name)}</xdr:absoluteAnchor>"
    )


def _drawing_xml(*anchors: str) -> bytes:
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<xdr:wsDr xmlns:xdr="{NS_XDR}"'
        f' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
        f' xmlns:r="{REL_NS}">{"".join(anchors)}</xdr:wsDr>'
    ).encode()


def _rels_xml(*rels: tuple[str, str, str]) -> bytes:
    body = "".join(
        f'<Relationship Id="{rel_id}" Type="{rel_type}" Target="{target}"/>'
        for rel_id, rel_type, target in rels
    )
    return (
        f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<Relationships xmlns="{PKG_REL_NS}">{body}</Relationships>'
    ).encode()


class AnchorFixture:
    """两张分表、五个锚点的工件，几何全部写死方便手算。

    分表「图纸」（行 1-20 高 20pt，列 A-H 宽 10）：
      - hit         twoCellAnchor 行 1-2   → 命中被改高的第 1 行
      - miss        twoCellAnchor 行 10-11 → 没人动过（且带 editAs="absolute"）
      - absolute_in absoluteAnchor 落在第 3 列第 5 行内 → 上方有行变高，会漂移
      - absolute_out absoluteAnchor 横坐标越过 <dimension> 最后一列（第 8 列）→
        落进没有 <col> 定义的第 9 列，按默认列宽外推定位（0-based col=8）
    分表「附表」（行 1-5 高 20pt，列 A 宽 10）：
      - hit2        twoCellAnchor 行 1-2   → 用来验证跨分表汇总
    """

    ABS_X = 2 * COL_EMU + 12345   # → col=2（0-based）, colOff=12345
    ABS_Y = 4 * ROW_EMU + 6789    # → row=4（0-based）, rowOff=6789
    # 第 8 列（1-based，A-H 都是显式宽度 10）累计跨度之外 100 EMU，落进第 9 列
    # （没有 <col> 定义，按 sheetFormatPr 默认列宽 8.43 渲染）→ col=8（0-based）、
    # colOff=100（默认列宽换算出的列宽远大于 100，不会再进位到第 10 列）。
    ABS_OUT_X = 8 * COL_EMU + 100

    @staticmethod
    def build(root: Path) -> Path:
        base = root / "anchor_base.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "图纸"
        sheet["A1"] = _LONG_TEXT
        sheet["H20"] = 1  # 只为把 dimension 撑到 A1:H20，数字不参与翻译
        for index in range(1, 21):
            sheet.row_dimensions[index].height = ROW_HEIGHT_PT
        for letter in "ABCDEFGH":
            sheet.column_dimensions[letter].width = COL_WIDTH_CHARS

        second = workbook.create_sheet("附表")
        second["A1"] = _LONG_TEXT
        for index in range(1, 6):
            second.row_dimensions[index].height = ROW_HEIGHT_PT
        second.column_dimensions["A"].width = COL_WIDTH_CHARS
        workbook.save(base)
        workbook.close()

        target = root / "anchor_fixture.xlsx"
        AnchorFixture._inject(base, target)
        return target

    @staticmethod
    def _inject(source: Path, target: Path) -> None:
        with zipfile.ZipFile(source) as archive:
            parts = {name: archive.read(name) for name in archive.namelist()}

        for sheet_part, rel_id in (
            ("xl/worksheets/sheet1.xml", "rIdDraw1"),
            ("xl/worksheets/sheet2.xml", "rIdDraw2"),
        ):
            sheet_xml = etree.fromstring(parts[sheet_part])
            drawing_el = etree.SubElement(sheet_xml, f"{{{NS_MAIN}}}drawing")
            drawing_el.set(f"{{{REL_NS}}}id", rel_id)
            parts[sheet_part] = etree.tostring(
                sheet_xml, xml_declaration=True, encoding="UTF-8", standalone=True
            )

        parts["xl/worksheets/_rels/sheet1.xml.rels"] = _rels_xml(
            ("rIdDraw1", f"{REL_NS}/drawing", "../drawings/drawing1.xml")
        )
        parts["xl/worksheets/_rels/sheet2.xml.rels"] = _rels_xml(
            ("rIdDraw2", f"{REL_NS}/drawing", "../drawings/drawing2.xml")
        )
        parts["xl/drawings/drawing1.xml"] = _drawing_xml(
            _two_cell("hit", 0, 2),
            _two_cell("miss", 9, 11, edit_as="absolute"),
            _absolute("absolute_in", AnchorFixture.ABS_X, AnchorFixture.ABS_Y),
            _absolute("absolute_out", AnchorFixture.ABS_OUT_X, 0),
        )
        parts["xl/drawings/drawing2.xml"] = _drawing_xml(_two_cell("hit2", 0, 2))
        for index in (1, 2):
            parts[f"xl/drawings/_rels/drawing{index}.xml.rels"] = _rels_xml(
                ("rIdImg", f"{REL_NS}/image", "../media/image1.png")
            )
        parts["xl/media/image1.png"] = PNG_BYTES

        content_types = etree.fromstring(parts["[Content_Types].xml"])
        for part_name in ("/xl/drawings/drawing1.xml", "/xl/drawings/drawing2.xml"):
            override = etree.SubElement(content_types, f"{{{CT_NS}}}Override")
            override.set("PartName", part_name)
            override.set(
                "ContentType",
                "application/vnd.openxmlformats-officedocument.drawing+xml",
            )
        default = etree.SubElement(content_types, f"{{{CT_NS}}}Default")
        default.set("Extension", "png")
        default.set("ContentType", "image/png")
        parts["[Content_Types].xml"] = etree.tostring(
            content_types, xml_declaration=True, encoding="UTF-8", standalone=True
        )

        with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as archive:
            for name, data in parts.items():
                archive.writestr(name, data)


def _anchors_by_name(drawing_xml: bytes) -> dict[str, tuple[str, object]]:
    """``{形状名: (锚点标签名, 锚点元素)}``。"""
    root = etree.fromstring(drawing_xml)
    result: dict[str, tuple[str, object]] = {}
    for anchor in root:
        name_el = next(iter(anchor.iter(f"{{{NS_XDR}}}cNvPr")), None)
        if name_el is None:
            continue
        result[name_el.get("name")] = (etree.QName(anchor).localname, anchor)
    return result


class AnchorFreezeTests(unittest.TestCase):
    TRANSLATIONS = {_LONG_TEXT: "A very long construction note that needs wrapping"}

    def setUp(self) -> None:
        self._temp = tempfile.TemporaryDirectory()
        self.root = Path(self._temp.name)
        self.fixture = AnchorFixture.build(self.root)

    def tearDown(self) -> None:
        self._temp.cleanup()

    def _write(self, **overrides) -> tuple[Path, dict[str, int]]:
        stats: dict[str, int] = {}
        kwargs = dict(
            source_path=self.fixture,
            output_dir=self.root / "out",
            translations=dict(self.TRANSLATIONS),
            target_lang="en",
            source_lang="zh",
            keep_original_sheets=False,
            formula_display_value_backfill=False,
            enable_print_guard=False,
            stats=stats,
        )
        kwargs.update(overrides)
        return write_bilingual_file(**kwargs), stats

    # ── 一、Excel 整表 autofit 会绕过「只冻结改过的行」 ────────────────────
    def test_external_autofit_freezes_every_floating_anchor(self) -> None:
        output, stats = self._write(external_autofit_planned=True)
        anchors = _anchors_by_name(_read_part(output, "xl/drawings/drawing1.xml"))

        self.assertEqual(anchors["hit"][0], "oneCellAnchor")
        self.assertEqual(
            anchors["miss"][0],
            "oneCellAnchor",
            "Excel 会重排整表行高，没被我们改过的行上的图片也必须冻结",
        )
        self.assertEqual(anchors["absolute_in"][0], "oneCellAnchor")
        # editAs="absolute" 不开特例，照常冻结（属性在 oneCellAnchor 上无处可放）
        self.assertNotIn("editAs", anchors["miss"][1].attrib)

        miss_ext = anchors["miss"][1].find(f"{{{NS_XDR}}}ext")
        self.assertEqual(int(miss_ext.get("cy")), 2 * ROW_EMU)
        self.assertEqual(int(miss_ext.get("cx")), 2 * COL_EMU)

        # hit + miss + absolute_in + absolute_out（越界外推也算）+ hit2
        self.assertEqual(stats["anchor_frozen_count"], 5)

    def test_without_autofit_only_changed_rows_are_frozen(self) -> None:
        output, stats = self._write()
        anchors = _anchors_by_name(_read_part(output, "xl/drawings/drawing1.xml"))

        self.assertEqual(anchors["hit"][0], "oneCellAnchor")
        self.assertEqual(
            anchors["miss"][0],
            "twoCellAnchor",
            "不跑 Excel autofit 时，没被改过的行上的图片应保持原样",
        )
        self.assertEqual(anchors["miss"][1].get("editAs"), "absolute")

        hit_ext = anchors["hit"][1].find(f"{{{NS_XDR}}}ext")
        self.assertEqual(int(hit_ext.get("cy")), 2 * ROW_EMU)
        self.assertEqual(int(hit_ext.get("cx")), 2 * COL_EMU)

        # 图纸：hit + absolute_in；附表：hit2
        self.assertEqual(stats["anchor_frozen_count"], 3)

    def test_locked_row_height_freezes_nothing(self) -> None:
        output, stats = self._write(lock_row_height=True, external_autofit_planned=True)
        for index in (1, 2):
            self.assertEqual(
                _read_part(output, f"xl/drawings/drawing{index}.xml"),
                _read_part(self.fixture, f"xl/drawings/drawing{index}.xml"),
                "锁定行高时行高不变，任何锚点都不该被动",
            )
        self.assertEqual(stats["anchor_frozen_count"], 0)

    # ── 二、absoluteAnchor 反查坐标 ───────────────────────────────────────
    def test_absolute_anchor_becomes_one_cell_anchor(self) -> None:
        output, _ = self._write()
        anchors = _anchors_by_name(_read_part(output, "xl/drawings/drawing1.xml"))
        tag, anchor = anchors["absolute_in"]

        self.assertEqual(tag, "oneCellAnchor")
        self.assertIsNone(anchor.find(f"{{{NS_XDR}}}pos"), "pos 必须被 from 取代")

        from_node = anchor.find(f"{{{NS_XDR}}}from")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}col"), "2")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}colOff"), "12345")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}row"), "4")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}rowOff"), "6789")

        # 尺寸原样保留：absoluteAnchor 本来就不随行高缩放
        ext = anchor.find(f"{{{NS_XDR}}}ext")
        self.assertEqual(int(ext.get("cx")), ABS_EXT_CX)
        self.assertEqual(int(ext.get("cy")), ABS_EXT_CY)

        children = [etree.QName(child).localname for child in anchor]
        self.assertEqual(children[:2], ["from", "ext"])
        self.assertIn("clientData", children)

    def test_out_of_dimension_absolute_anchor_extrapolates_default_size(self) -> None:
        # 越过 <dimension> 最后一列的坐标不是没依据——Excel 自己渲染这些没有
        # <col> 定义的列时就是按默认列宽算，我们跟着算，不能当成瞎猜就撒手不管。
        output, _ = self._write(external_autofit_planned=True)
        anchors = _anchors_by_name(_read_part(output, "xl/drawings/drawing1.xml"))
        tag, anchor = anchors["absolute_out"]

        self.assertEqual(tag, "oneCellAnchor")
        from_node = anchor.find(f"{{{NS_XDR}}}from")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}col"), "8")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}colOff"), "100")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}row"), "0")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}rowOff"), "0")

        ext = anchor.find(f"{{{NS_XDR}}}ext")
        self.assertEqual(int(ext.get("cx")), ABS_EXT_CX)
        self.assertEqual(int(ext.get("cy")), ABS_EXT_CY)

    def test_cloned_original_sheets_are_frozen_too_under_autofit(self) -> None:
        """Excel autofit 跑的是整本工作簿，克隆出来的原文分表也会被重排行高。"""
        output, stats = self._write(
            keep_original_sheets=True, external_autofit_planned=True
        )
        with zipfile.ZipFile(output) as archive:
            cloned = [
                name
                for name in archive.namelist()
                if name.startswith("xl/drawings/drawing")
                and name.endswith(".xml")
                and name not in {"xl/drawings/drawing1.xml", "xl/drawings/drawing2.xml"}
            ]
        self.assertEqual(len(cloned), 2, "两张分表各克隆一份 drawing")
        for part in cloned:
            anchors = _anchors_by_name(_read_part(output, part))
            for name, (tag, _anchor) in anchors.items():
                self.assertEqual(tag, "oneCellAnchor", f"{part} 的 {name} 没被冻结")
        # 只数正表 5（图纸 4 + 附表 1）。克隆件上面那 5 个锚点上面刚断言过确实被冻结了，
        # 但它们和正表是同一批图片，计入汇总会让面向用户的「N 张悬浮图片已固定尺寸」翻倍。
        self.assertEqual(stats["anchor_frozen_count"], 5)

    def test_cloned_original_sheets_untouched_without_autofit(self) -> None:
        output, _ = self._write(keep_original_sheets=True)
        with zipfile.ZipFile(output) as archive:
            cloned = sorted(
                name
                for name in archive.namelist()
                if name.startswith("xl/drawings/drawing")
                and name.endswith(".xml")
                and name not in {"xl/drawings/drawing1.xml", "xl/drawings/drawing2.xml"}
            )
        self.assertEqual(
            [_read_part(output, part) for part in cloned],
            [
                _read_part(self.fixture, "xl/drawings/drawing1.xml"),
                _read_part(self.fixture, "xl/drawings/drawing2.xml"),
            ],
            "不跑 Excel autofit 时，原文分表行高没变，克隆件应逐字节照抄",
        )

    # ── 三、计数汇总 ──────────────────────────────────────────────────────
    def test_anchor_frozen_count_sums_across_sheets(self) -> None:
        output, stats = self._write(external_autofit_planned=True)
        # 图纸 4（hit / miss / absolute_in / absolute_out）+ 附表 1
        self.assertEqual(stats["anchor_frozen_count"], 5)

        # 第二张分表确实也被处理过，计数不是单表的
        self.assertEqual(
            _anchors_by_name(_read_part(self.fixture, "xl/drawings/drawing2.xml"))["hit2"][0],
            "twoCellAnchor",
        )
        self.assertEqual(
            _anchors_by_name(_read_part(output, "xl/drawings/drawing2.xml"))["hit2"][0],
            "oneCellAnchor",
        )


class AbsoluteAnchorGeometryTests(unittest.TestCase):
    """几何信息不全时 absoluteAnchor 必须原样不动。"""

    def _drawing_root(self):
        return etree.fromstring(_drawing_xml(_absolute("only", 100, 100)))

    def test_missing_geometry_leaves_absolute_anchor_untouched(self) -> None:
        # 既没有 <dimension> 也没有任何单元格 → 算不出整表范围
        sheet_root = etree.fromstring(
            f'<worksheet xmlns="{NS_MAIN}"><sheetData/></worksheet>'.encode()
        )
        geometry = _SheetGeometry(sheet_root)
        self.assertIsNone(geometry.extent())

        drawing_root = self._drawing_root()
        frozen = fix_drawing_anchors(drawing_root, geometry, set(), freeze_all=True)
        self.assertEqual(frozen, 0)
        self.assertEqual(etree.QName(drawing_root[0]).localname, "absoluteAnchor")

    def test_extent_falls_back_to_scanning_cells(self) -> None:
        sheet_root = etree.fromstring(
            (
                f'<worksheet xmlns="{NS_MAIN}"><sheetData>'
                f'<row r="1"><c r="A1"/><c r="C1"/></row>'
                f'<row r="4"><c r="B4"/></row>'
                f"</sheetData></worksheet>"
            ).encode()
        )
        self.assertEqual(_SheetGeometry(sheet_root).extent(), (4, 3))

    # ── 越过 <dimension> 之后按默认尺寸外推 ─────────────────────────────────
    def _sheet_root_with_dimension(self):
        """A1:B3：行高列宽都显式写死，defaultRowHeight/defaultColWidth 也写成
        好算的数，方便手算越过 ``<dimension>`` 之后的外推结果。"""
        return etree.fromstring(
            (
                f'<worksheet xmlns="{NS_MAIN}">'
                f'<dimension ref="A1:B3"/>'
                f'<sheetFormatPr defaultRowHeight="25" defaultColWidth="20"/>'
                f'<cols><col min="1" max="2" width="10" customWidth="1"/></cols>'
                f"<sheetData>"
                f'<row r="1" ht="20" customHeight="1"/>'
                f'<row r="2" ht="20" customHeight="1"/>'
                f'<row r="3" ht="20" customHeight="1"/>'
                f"</sheetData></worksheet>"
            ).encode()
        )

    def _freeze_on(self, sheet_root, x: int, y: int) -> tuple[str, object]:
        """把单个 absoluteAnchor 跑一遍冻结逻辑，返回 ``(锚点标签名, 锚点元素)``。"""
        geometry = _SheetGeometry(sheet_root)
        drawing_root = etree.fromstring(_drawing_xml(_absolute("only", x, y)))
        fix_drawing_anchors(drawing_root, geometry, set(), freeze_all=True)
        anchor = drawing_root[0]
        return etree.QName(anchor).localname, anchor

    def _freeze_only(self, x: int, y: int) -> tuple[str, object]:
        return self._freeze_on(self._sheet_root_with_dimension(), x, y)

    def test_below_last_row_extrapolates_with_default_row_height(self) -> None:
        # 行 1-3 每行 20pt → 254000 EMU/行，累计 762000。defaultRowHeight=25pt
        # → 每行外推尺寸 round(25×12700)=317500 EMU。y 落在累计之后 1000 EMU
        # 处，理应落进紧接着的第 4 行（0-based row=3），偏移就是那 1000。
        # x 取 500，落在第 1 列（列宽 10 → col_emu=666750）内部，走的是范围内
        # 的原逻辑，用来确认外推只发生在越界的那一维。
        tag, anchor = self._freeze_only(500, 3 * 254000 + 1000)
        self.assertEqual(tag, "oneCellAnchor")
        from_node = anchor.find(f"{{{NS_XDR}}}from")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}col"), "0")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}colOff"), "500")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}row"), "3")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}rowOff"), "1000")

    def test_right_of_last_col_extrapolates_with_default_col_width(self) -> None:
        # 列 A/B 宽 10 → col_emu=666750/列，累计 1333500。defaultColWidth=20 →
        # 每列外推尺寸同样算下来是 1333500 EMU（宽度翻倍、换算出的像素也翻
        # 倍）。x 落在累计之后 2000 EMU 处，理应落进紧接着的第 3 列（0-based
        # col=2），偏移就是那 2000。
        tag, anchor = self._freeze_only(2 * 666750 + 2000, 1000)
        self.assertEqual(tag, "oneCellAnchor")
        from_node = anchor.find(f"{{{NS_XDR}}}from")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}col"), "2")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}colOff"), "2000")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}row"), "0")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}rowOff"), "1000")

    def test_far_beyond_excel_row_limit_is_left_alone(self) -> None:
        # 外推出的行号一旦超过 Excel 本身的行号上限（0-based 1048575），说明
        # 坐标本身就是坏数据，不能再往下算，只能保持原样。
        tag, _anchor = self._freeze_only(500, 10**15)
        self.assertEqual(tag, "absoluteAnchor")

    def test_negative_coordinate_is_left_alone(self) -> None:
        tag, _anchor = self._freeze_only(-1, 500)
        self.assertEqual(tag, "absoluteAnchor")

    # ── <dimension> 之外仍有显式行高/列宽 ──────────────────────────────────
    def _sheet_root_sized_beyond_dimension(self):
        """数据只有 A1:B3，但为了摆图片把第 4-10 行拉高、C-J 列拉宽。

        ``<dimension>`` 只覆盖有内容的单元格，不会因为设了行高列宽而变大，所以
        这些真实尺寸落在 ``<dimension>`` 之外——最普通不过的现实写法。
        """
        return etree.fromstring(
            (
                f'<worksheet xmlns="{NS_MAIN}">'
                f'<dimension ref="A1:B3"/>'
                f'<sheetFormatPr defaultRowHeight="25" defaultColWidth="20"/>'
                f"<cols>"
                f'<col min="1" max="2" width="10" customWidth="1"/>'
                f'<col min="3" max="10" width="50" customWidth="1"/>'
                f"</cols>"
                f"<sheetData>"
                + "".join(
                    f'<row r="{r}" ht="20" customHeight="1"/>' for r in range(1, 4)
                )
                + "".join(
                    f'<row r="{r}" ht="60" customHeight="1"/>' for r in range(4, 11)
                )
                + "</sheetData></worksheet>"
            ).encode()
        )

    def test_explicit_sizes_outside_dimension_beat_default_extrapolation(self) -> None:
        # 手算目标格 E8（0-based col=4 / row=7）：
        #   列宽 10 → int((256×10 + int(128/7))/256 × 7) = 70 像素 × 9525 = 666750
        #   列宽 50 → int((256×50 + int(128/7))/256 × 7) = 350 像素 × 9525 = 3333750
        #   A+B 两列 666750 + C/D 两列 3333750 = 8001000，E 列从这里开始
        #   行高 20pt → 254000，行高 60pt → 762000
        #   第 1-3 行 254000 + 第 4-7 行 762000 = 3810000，第 8 行从这里开始
        x = 2 * 666750 + 2 * 3333750 + 12345
        y = 3 * 254000 + 4 * 762000 + 6789
        self.assertEqual((x, y), (8013345, 3816789))

        tag, anchor = self._freeze_on(self._sheet_root_sized_beyond_dimension(), x, y)
        self.assertEqual(tag, "oneCellAnchor")
        from_node = anchor.find(f"{{{NS_XDR}}}from")
        # 只累加到 <dimension>（B3）再按默认尺寸外推的话会算成 col=7 / row=12，
        # 图片被平移 3 列 5 行——这个功能的全部意义就是把图片钉在原位。
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}col"), "4")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}colOff"), "12345")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}row"), "7")
        self.assertEqual(from_node.findtext(f"{{{NS_XDR}}}rowOff"), "6789")

    def test_bogus_dimension_cannot_yield_row_beyond_excel_limit(self) -> None:
        """畸形 ``<dimension>`` 不能把逐格累加那条路径也拖到 Excel 行号上限之外。

        ``<dimension ref>`` 是外部数据，行号位数不设限；坐标又落在没有任何显式
        行高的区域，逐格累加会一路走到两百万行去。Excel 打不开这样的 from/row，
        宁可保持 absoluteAnchor 原样。
        """
        sheet_root = etree.fromstring(
            (
                f'<worksheet xmlns="{NS_MAIN}">'
                f'<dimension ref="A1:A99999999"/>'
                f'<sheetFormatPr defaultRowHeight="25" defaultColWidth="20"/>'
                f'<sheetData><row r="1" ht="20" customHeight="1"/></sheetData>'
                f"</worksheet>"
            ).encode()
        )
        # 默认行高 25pt → 317500 EMU/行，取第 2000000 行往下 100 EMU 处
        tag, anchor = self._freeze_on(sheet_root, 500, 2_000_000 * 317500 + 100)
        self.assertEqual(tag, "absoluteAnchor")
        self.assertIsNone(anchor.find(f"{{{NS_XDR}}}from"))
        self.assertIsNotNone(anchor.find(f"{{{NS_XDR}}}pos"))


class ColumnLetterTests(unittest.TestCase):
    def test_round_trip(self) -> None:
        for index, letters in ((1, "A"), (26, "Z"), (27, "AA"), (703, "AAA")):
            self.assertEqual(_column_letter(index), letters)


if __name__ == "__main__":
    unittest.main(verbosity=2)
