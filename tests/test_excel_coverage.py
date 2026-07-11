from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.excel_coverage import (
    build_excel_coverage_plan,
    write_untranslated_excel_file,
)
from core.translation_coverage import (
    COVERAGE_COVERED,
    COVERAGE_SOURCE_ONLY,
)


class ExcelCoverageTests(unittest.TestCase):
    def test_same_cell_bilingual_is_covered_and_source_only_is_written(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(
                Path(tmp),
                "source.xlsx",
                {
                    "A1": "项目名称\nProject name",
                    "A2": "施工内容",
                    "A3": "Already translated",
                },
            )

            plan = build_excel_coverage_plan(source, target_lang="en", source_lang="zh")
            by_location = {unit.location: unit for unit in plan.units}

            self.assertEqual(by_location["Sheet!A1"].status, COVERAGE_COVERED)
            self.assertEqual(by_location["Sheet!A2"].status, COVERAGE_SOURCE_ONLY)
            self.assertEqual(plan.source_texts, ["施工内容"])

            out_path = write_untranslated_excel_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                plan=plan,
                translations={"施工内容": "Construction scope"},
                target_lang="en",
                source_lang="zh",
                keep_original_sheets=True,
            )

            wb = load_workbook(out_path)
            try:
                self.assertEqual(wb["Sheet"]["A1"].value, "项目名称\nProject name")
                self.assertEqual(wb["Sheet"]["A2"].value, "施工内容\nConstruction scope")
                self.assertEqual(wb["Sheet"]["A3"].value, "Already translated")
                self.assertIn("Sheet_原文", wb.sheetnames)
            finally:
                wb.close()

            second_plan = build_excel_coverage_plan(
                out_path,
                target_lang="en",
                source_lang="zh",
            )
            self.assertEqual(second_plan.source_units, [])

    def test_duplicate_source_text_only_patches_untranslated_position(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(
                Path(tmp),
                "source.xlsx",
                {
                    "A1": "项目名称\nProject name",
                    "A2": "项目名称",
                },
            )

            plan = build_excel_coverage_plan(source, target_lang="en", source_lang="zh")

            self.assertEqual(len(plan.source_units), 1)
            self.assertEqual(plan.source_units[0].location, "Sheet!A2")

            out_path = write_untranslated_excel_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                plan=plan,
                translations={"项目名称": "Project name"},
                target_lang="en",
                source_lang="zh",
                keep_original_sheets=False,
            )

            wb = load_workbook(out_path)
            try:
                self.assertEqual(wb["Sheet"]["A1"].value, "项目名称\nProject name")
                self.assertEqual(wb["Sheet"]["A2"].value, "项目名称\nProject name")
            finally:
                wb.close()

    def test_english_to_french_does_not_treat_french_line_as_new_source(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(
                Path(tmp),
                "source.xlsx",
                {
                    "A1": "The concrete slab is ready\nLa dalle de béton est prête",
                    "A2": "The concrete wall is ready",
                    "A3": "Project name",
                },
            )

            plan = build_excel_coverage_plan(source, target_lang="fr", source_lang="en")
            by_location = {unit.location: unit for unit in plan.units}

            self.assertEqual(by_location["Sheet!A1"].status, COVERAGE_COVERED)
            self.assertEqual(by_location["Sheet!A2"].status, COVERAGE_SOURCE_ONLY)
            self.assertEqual(
                plan.source_texts,
                ["The concrete wall is ready", "Project name"],
            )

    @staticmethod
    def _workbook(root: Path, name: str, cells: dict[str, str]) -> Path:
        path = root / name
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        for coordinate, value in cells.items():
            ws[coordinate] = value
        wb.save(path)
        wb.close()
        return path


if __name__ == "__main__":
    unittest.main(verbosity=2)
