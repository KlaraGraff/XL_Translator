from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from core.bilingual_writer import build_output_dir, write_bilingual_file
from core.mixed_language import (
    MIXED_COLOR_FOREIGN_NOISE,
    MIXED_COLOR_UNRESOLVED,
    MIXED_MARK_FOREIGN_NOISE,
    MIXED_MARK_SEMANTIC,
    MIXED_MARK_UNRESOLVED,
)


class BilingualWriterTests(unittest.TestCase):
    def test_excel_review_mark_fills_unfilled_cell(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(Path(tmp), "source.xlsx", "项目")
            out_path = write_bilingual_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                translations={"项目": "Projet"},
                target_lang="fr",
                keep_original_sheets=False,
                formula_display_value_backfill=True,
                enable_print_guard=False,
                review_marks={"项目": MIXED_MARK_FOREIGN_NOISE},
            )

            wb = load_workbook(out_path)
            try:
                cell = wb.active["A1"]
                self.assertEqual(cell.value, "项目\nProjet")
                self.assertTrue(str(cell.fill.fgColor.rgb).endswith(MIXED_COLOR_FOREIGN_NOISE))
            finally:
                wb.close()

    def test_excel_review_mark_uses_configured_color_map(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(Path(tmp), "source.xlsx", "项目")
            out_path = write_bilingual_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                translations={"项目": "Projet"},
                target_lang="fr",
                keep_original_sheets=False,
                formula_display_value_backfill=True,
                enable_print_guard=False,
                review_marks={"项目": MIXED_MARK_SEMANTIC},
                review_mark_colors={MIXED_MARK_SEMANTIC: "DDEBFF"},
            )

            wb = load_workbook(out_path)
            try:
                cell = wb.active["A1"]
                self.assertEqual(cell.value, "项目\nProjet")
                self.assertTrue(str(cell.fill.fgColor.rgb).endswith("DDEBFF"))
            finally:
                wb.close()

    def test_excel_review_mark_uses_red_font_when_existing_fill_conflicts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(
                Path(tmp),
                "source.xlsx",
                "项目",
                fill_color="92D050",
                font_color="0000FF",
            )
            out_path = write_bilingual_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                translations={"项目": "Projet"},
                target_lang="fr",
                keep_original_sheets=False,
                formula_display_value_backfill=True,
                enable_print_guard=False,
                review_marks={"项目": MIXED_MARK_UNRESOLVED},
                existing_fill_policy="red_font",
            )

            wb = load_workbook(out_path)
            try:
                cell = wb.active["A1"]
                self.assertTrue(str(cell.fill.fgColor.rgb).endswith("92D050"))
                self.assertTrue(str(cell.font.color.rgb).endswith("FF0000"))
            finally:
                wb.close()

    def test_excel_review_mark_can_skip_existing_fill(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(
                Path(tmp),
                "source.xlsx",
                "项目",
                fill_color="92D050",
                font_color="0000FF",
            )
            out_path = write_bilingual_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                translations={"项目": "Projet"},
                target_lang="fr",
                keep_original_sheets=False,
                formula_display_value_backfill=True,
                enable_print_guard=False,
                review_marks={"项目": MIXED_MARK_UNRESOLVED},
                existing_fill_policy="skip",
            )

            wb = load_workbook(out_path)
            try:
                cell = wb.active["A1"]
                self.assertTrue(str(cell.fill.fgColor.rgb).endswith("92D050"))
                self.assertTrue(str(cell.font.color.rgb).endswith("0000FF"))
            finally:
                wb.close()

    def test_excel_retained_original_is_marked_as_unresolved(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(Path(tmp), "source.xlsx", "项目")
            out_path = write_bilingual_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                translations={"项目": "项目"},
                target_lang="fr",
                keep_original_sheets=False,
                formula_display_value_backfill=True,
                enable_print_guard=False,
            )

            wb = load_workbook(out_path)
            try:
                cell = wb.active["A1"]
                self.assertEqual(cell.value, "项目")
                self.assertTrue(str(cell.fill.fgColor.rgb).endswith(MIXED_COLOR_UNRESOLVED))
            finally:
                wb.close()

    def test_excel_review_mark_can_be_disabled(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(Path(tmp), "source.xlsx", "项目")
            out_path = write_bilingual_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                translations={"项目": "Projet"},
                target_lang="fr",
                keep_original_sheets=False,
                formula_display_value_backfill=True,
                enable_print_guard=False,
                review_marks={"项目": MIXED_MARK_FOREIGN_NOISE},
                mark_review_items=False,
            )

            wb = load_workbook(out_path)
            try:
                cell = wb.active["A1"]
                self.assertEqual(cell.value, "项目\nProjet")
                self.assertEqual(str(cell.fill.fgColor.rgb), "00000000")
            finally:
                wb.close()

    def test_excel_disabled_review_mark_skips_retained_original_auto_mark(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(Path(tmp), "source.xlsx", "项目")
            out_path = write_bilingual_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                translations={"项目": "项目"},
                target_lang="fr",
                keep_original_sheets=False,
                formula_display_value_backfill=True,
                enable_print_guard=False,
                mark_review_items=False,
            )

            wb = load_workbook(out_path)
            try:
                cell = wb.active["A1"]
                self.assertEqual(cell.value, "项目")
                self.assertEqual(str(cell.fill.fgColor.rgb), "00000000")
            finally:
                wb.close()

    def test_plain_text_containing_dispimg_is_not_erased(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(Path(tmp), "source.xlsx", "DISPIMG operating note")
            out_path = write_bilingual_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                translations={"DISPIMG operating note": "Note de fonctionnement DISPIMG"},
                target_lang="fr",
                source_lang="en",
                keep_original_sheets=False,
                formula_display_value_backfill=True,
                enable_print_guard=False,
            )

            wb = load_workbook(out_path)
            try:
                self.assertEqual(
                    wb.active["A1"].value,
                    "DISPIMG operating note\nNote de fonctionnement DISPIMG",
                )
            finally:
                wb.close()

    def test_wps_dispimg_formula_is_preserved(self) -> None:
        # 补丁式写入器保留 WPS 嵌入图片公式，图片不再随翻译丢失。
        with tempfile.TemporaryDirectory() as tmp:
            source = self._workbook(Path(tmp), "source.xlsx", '=DISPIMG("ID_1",1)')
            out_path = write_bilingual_file(
                source_path=source,
                output_dir=Path(tmp) / "out",
                translations={},
                target_lang="fr",
                keep_original_sheets=False,
                formula_display_value_backfill=False,
                enable_print_guard=False,
            )

            wb = load_workbook(out_path, data_only=False)
            try:
                self.assertEqual(wb.active["A1"].value, '=DISPIMG("ID_1",1)')
            finally:
                wb.close()

    @staticmethod
    def _workbook(
        root: Path,
        name: str,
        value: str,
        *,
        fill_color: str | None = None,
        font_color: str | None = None,
    ) -> Path:
        path = root / name
        wb = Workbook()
        ws = wb.active
        ws["A1"] = value
        if fill_color:
            ws["A1"].fill = PatternFill(fill_type="solid", fgColor=f"FF{fill_color}")
        if font_color:
            ws["A1"].font = Font(color=f"FF{font_color}")
        wb.save(path)
        wb.close()
        return path


if __name__ == "__main__":
    unittest.main(verbosity=2)


class OutputDirNamingTests(unittest.TestCase):
    """输出目录名要能被人读、被人复制，所以不带微秒也不带哈希。"""

    def test_name_is_source_folder_plus_timestamp_to_the_second(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "fixtures"
            source.mkdir()
            name = build_output_dir(source).name
            self.assertTrue(name.startswith("fixtures_翻译输出_"))
            stamp = name.removeprefix("fixtures_翻译输出_")
            self.assertRegex(stamp, r"^\d{8}_\d{6}$")

    def test_same_second_rerun_still_gets_a_distinct_directory(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "fixtures"
            source.mkdir()
            first = build_output_dir(source)
            first.mkdir()
            second = build_output_dir(source)
            self.assertNotEqual(first, second)
            self.assertEqual(second.name, f"{first.name}_2")

    def test_custom_root_is_honoured(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            source = Path(tmp) / "fixtures"
            source.mkdir()
            custom = Path(tmp) / "elsewhere"
            custom.mkdir()
            self.assertEqual(build_output_dir(source, custom).parent, custom)

