"""审计报告 4.11 / 4.12 的回归测试（补译覆盖检测）。

两个场景都用真实 xlsx 走完整链路：先用 ``write_bilingual_file`` 生成一份带
``_原文`` 克隆分表的产物，再把这份产物交给 ``build_excel_coverage_plan``——这正是
用户「翻一遍、再补译一遍」时的实际路径。
"""
from __future__ import annotations

import tempfile
import time
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.bilingual_writer import write_bilingual_file
from core.excel_coverage import build_excel_coverage_plan
from core.translation_coverage import COVERAGE_SOURCE_ONLY


class _CoverageCase(unittest.TestCase):
    def setUp(self) -> None:
        self._temp = tempfile.TemporaryDirectory()
        self.addCleanup(self._temp.cleanup)
        self.root = Path(self._temp.name)

    def _translate(self, source: Path, translations: dict[str, str]) -> Path:
        return write_bilingual_file(
            source_path=source,
            output_dir=self.root / "out",
            translations=translations,
            target_lang="en",
            source_lang="zh",
            keep_original_sheets=True,
            formula_display_value_backfill=False,
            enable_print_guard=False,
        )

    def _plan_sheets(self, path: Path) -> set[str]:
        """出现在补译候选里的分表名。"""
        plan = build_excel_coverage_plan(path, target_lang="en", source_lang="zh")
        return {
            str(unit.data.get("sheet"))
            for unit in plan.units
            if unit.status == COVERAGE_SOURCE_ONLY
        }


class GeneratedOriginalSheetTests(_CoverageCase):
    """4.12 克隆分表名反推错误 → 自己生成的 _原文 分表被当成待译内容。"""

    def test_truncated_clone_name_is_still_recognized(self) -> None:
        # 30 个字符的分表名：拼上「_原文」有 33 个字符，被 31 字符上限截成
        # 「……_」，尾巴上的「_原文」自己就没了，靠字符串反推必然认不出来。
        long_name = "施工项目明细与单价对照表第一分册" + "补充说明说明" * 2 + "甲"
        self.assertEqual(len(long_name), 29)
        source = self.root / "long_sheet.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = long_name
        sheet["A1"] = "施工内容"
        summary = workbook.create_sheet("汇总")
        summary["A1"] = "配电箱"
        workbook.save(source)
        workbook.close()

        output = self._translate(source, {"施工内容": "Construction scope"})

        translated = load_workbook(output)
        try:
            clone_names = [
                name
                for name in translated.sheetnames
                if name not in {long_name, "汇总"}
            ]
        finally:
            translated.close()
        self.assertEqual(len(clone_names), 2, translated.sheetnames)
        # 前提确认：克隆名确实已经被截得认不出「_原文」了。
        long_clone = next(name for name in clone_names if name.startswith(long_name[:20]))
        self.assertFalse(
            long_clone.endswith("_原文"),
            f"这个场景要的就是被截断的克隆名，实际拿到 {long_clone!r}",
        )

        sheets_in_plan = self._plan_sheets(output)
        for clone in clone_names:
            self.assertNotIn(
                clone,
                sheets_in_plan,
                f"自己生成的原文分表 {clone!r} 被当成待译内容，补译会把它整张重翻一遍",
            )
        # 别矫枉过正：没被翻译的普通分表仍然要进补译候选。
        self.assertIn("汇总", sheets_in_plan)

    def test_dedup_suffixed_clone_name_is_still_recognized(self) -> None:
        # 工作簿里已经有一张叫「报价_原文」的分表 → 克隆「报价」时重名，
        # 生成的名字带上 _2 后缀。
        source = self.root / "dedup.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "报价"
        sheet["A1"] = "施工内容"
        existing = workbook.create_sheet("报价_原文")
        existing["A1"] = "施工内容"
        workbook.save(source)
        workbook.close()

        output = self._translate(source, {"施工内容": "Construction scope"})

        translated = load_workbook(output)
        try:
            self.assertIn("报价_原文_2", translated.sheetnames)
        finally:
            translated.close()

        self.assertNotIn(
            "报价_原文_2",
            self._plan_sheets(output),
            "带去重后缀的克隆分表同样是我们自己生成的，补译必须跳过",
        )


class CoverageScanPerformanceTests(_CoverageCase):
    """4.11 公式格的显示值查询原本是 O(n²)。"""

    FORMULA_CELLS = 1600

    def _formula_workbook(self) -> Path:
        source = self.root / "formulas.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "报价"
        sheet["A1"] = "施工内容"
        for index in range(2, self.FORMULA_CELLS + 2):
            sheet.cell(row=index, column=1, value=f"=B{index}&C{index}")
        workbook.save(source)
        workbook.close()
        return source

    def test_large_formula_sheet_scans_quickly(self) -> None:
        source = self._formula_workbook()

        started = time.perf_counter()
        plan = build_excel_coverage_plan(
            source,
            target_lang="en",
            source_lang="zh",
            formula_display_value_backfill=True,
        )
        elapsed = time.perf_counter() - started

        self.assertEqual(plan.source_texts, ["施工内容"])
        # 逐格重解析整张 sheet XML 时这一步实测 7.7s；一次遍历建映射后在 0.1s 量级。
        # 阈值给得很松，只用来挡住 O(n²) 的回归。
        self.assertLess(
            elapsed,
            2.0,
            f"{self.FORMULA_CELLS} 个公式格扫了 {elapsed:.2f}s，公式显示值查询疑似退回逐格重解析",
        )

    def test_formula_display_values_are_still_resolved(self) -> None:
        """改成映射查表后，公式格的缓存显示值仍然要能被读出来。"""
        source = self.root / "cached.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "报价"
        sheet["A1"] = "=B1&C1"
        workbook.save(source)
        workbook.close()

        # openpyxl 写不出缓存值，手工把 <v> 塞进分表 XML。
        _inject_cached_value(source, "A1", "配电箱")

        plan = build_excel_coverage_plan(
            source,
            target_lang="en",
            source_lang="zh",
            formula_display_value_backfill=True,
        )
        self.assertEqual(plan.source_texts, ["配电箱"])


def _inject_cached_value(path: Path, coordinate: str, value: str) -> None:
    """给公式单元格补一个 ``t="str"`` 的缓存显示值。"""
    import shutil
    import zipfile

    from lxml import etree

    from core.xlsx_patcher import NS_MAIN

    temp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path) as archive:
        parts = {name: archive.read(name) for name in archive.namelist()}

    root = etree.fromstring(parts["xl/worksheets/sheet1.xml"])
    for cell in root.iter(f"{{{NS_MAIN}}}c"):
        if cell.get("r") != coordinate:
            continue
        cell.set("t", "str")
        existing = cell.find(f"{{{NS_MAIN}}}v")
        if existing is None:
            existing = etree.SubElement(cell, f"{{{NS_MAIN}}}v")
        existing.text = value
    parts["xl/worksheets/sheet1.xml"] = etree.tostring(
        root, xml_declaration=True, encoding="UTF-8", standalone=True
    )

    with zipfile.ZipFile(temp, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in parts.items():
            archive.writestr(name, data)
    shutil.move(str(temp), str(path))


if __name__ == "__main__":
    unittest.main()
