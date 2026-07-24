"""Phase 4 Excel contracts that must remain safe without local Office.

All task-runner cases use mocked translation/Excel automation.  The suite is
therefore suitable for the isolated macOS development gate while the real
Office and macOS 12 release checks stay separately deferred.
"""

from __future__ import annotations

import hashlib
import json
import tempfile
import unittest
from contextlib import ExitStack
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from core.api_config_check import ApiConfigCheckResult
from core import xls_converter
from core.file_scanner import FileItem, scan_excel_sources
from core.language_preflight import TranslationLanguageResult
from core.mixed_language import MIXED_MARK_UNRESOLVED
from core.model_throughput import EffectiveModelThroughput
from core.task_runner import DoneMsg, StoppedMsg, TaskRunner
from settings import AppSettings, EngineSettings


class _PreflightEngine:
    engine_name = "phase4/mock"

    def __init__(self, result_for_sample: dict[str, str] | None = None) -> None:
        self.calls: list[tuple[str, str]] = []
        self._result_for_sample = dict(result_for_sample or {})

    def chat(self, system: str, user: str) -> str:
        self.calls.append((system, user))
        samples = json.loads(user)["samples"]
        first = str(samples[0]) if samples else ""
        return self._result_for_sample.get(first, '{"source_langs":["zh"]}')


def _settings(*, source_lang: str = "zh", target_lang: str = "en") -> AppSettings:
    return AppSettings(
        engine=EngineSettings(
            mode="cloud",
            cloud_provider="custom_openai",
            cloud_model="phase4-test-model",
            cloud_base_url="https://example.invalid/v1",
            concurrency=1,
            batch_size=10,
        ),
        source_lang=source_lang,
        target_lang=target_lang,
    )


def _done_message(runner: TaskRunner) -> DoneMsg:
    messages = [message for message in list(runner._queue.queue) if isinstance(message, DoneMsg)]
    if len(messages) != 1:
        raise AssertionError(f"expected one DoneMsg, got {len(messages)}")
    return messages[0]


class ExcelWorkbookContractTests(unittest.TestCase):
    def test_writer_keeps_source_untouched_formula_merge_style_and_existing_fill(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "报价"
            ws["A1"] = "施工内容"
            ws["A1"].font = Font(name="Arial", bold=True, italic=True)
            ws["A1"].fill = PatternFill("solid", fgColor="00B050")
            ws["B1"] = "=SUM(C1:C2)"
            ws.merge_cells("A2:B2")
            ws["A2"] = "合并单元格内容"
            ws.row_dimensions[1].height = 25
            wb.save(source)
            wb.close()
            source_sha256 = hashlib.sha256(source.read_bytes()).hexdigest()
            review_positions: list[dict[str, str]] = []

            from core.bilingual_writer import write_bilingual_file

            output = write_bilingual_file(
                source_path=source,
                output_dir=root / "out",
                translations={"施工内容": "Construction scope"},
                target_lang="en",
                source_lang="zh",
                keep_original_sheets=True,
                formula_display_value_backfill=False,
                enable_print_guard=False,
                review_marks={"施工内容": MIXED_MARK_UNRESOLVED},
                review_mark_colors={MIXED_MARK_UNRESOLVED: "FF0000"},
                mark_review_items=True,
                existing_fill_policy="skip",
                review_positions=review_positions,
            )

            self.assertEqual(hashlib.sha256(source.read_bytes()).hexdigest(), source_sha256)
            written = load_workbook(output, data_only=False)
            try:
                sheet = written["报价"]
                self.assertEqual(sheet["A1"].value, "施工内容\nConstruction scope")
                self.assertEqual(sheet["B1"].value, "=SUM(C1:C2)")
                self.assertIn("A2:B2", {str(item) for item in sheet.merged_cells.ranges})
                self.assertTrue(sheet["A1"].font.bold)
                self.assertTrue(sheet["A1"].font.italic)
                self.assertEqual(sheet["A1"].fill.fgColor.rgb, "0000B050")
                self.assertIn("报价_原文", written.sheetnames)
                self.assertEqual(written["报价_原文"]["A1"].value, "施工内容")
            finally:
                written.close()
            self.assertEqual(
                review_positions,
                [
                    {
                        "worksheet": "报价",
                        "cell": "A1",
                        "category": MIXED_MARK_UNRESOLVED,
                        "action": "preserved_existing_fill",
                    }
                ],
            )


class ExcelScanningContractTests(unittest.TestCase):
    def test_scan_reports_usable_files_skips_broken_sources_and_excludes_old_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            good = root / "nested" / "source.xlsx"
            good.parent.mkdir()
            workbook = Workbook()
            workbook.active.title = "First"
            workbook.create_sheet("Second")
            workbook.save(good)
            workbook.close()
            (root / "~$source.xlsx").touch()
            (root / "broken.xlsx").write_text("not an Excel workbook", encoding="utf-8")
            old_output = root / "_翻译输出_20260724" / "old.xlsx"
            old_output.parent.mkdir()
            old_output.write_bytes(good.read_bytes())

            result = scan_excel_sources(root)

            self.assertEqual([item.path for item in result.items], [good])
            self.assertEqual(result.items[0].relative_path, "nested/source.xlsx")
            self.assertEqual(result.items[0].format, "xlsx")
            self.assertEqual(len(result.items[0].sheets), 2)
            self.assertEqual(result.summary["scanned_count"], 1)
            self.assertEqual(result.summary["selected_count"], 1)
            self.assertEqual(result.summary["sheet_count"], 2)
            self.assertEqual(result.summary["xls_count"], 0)
            self.assertEqual(len(result.skipped), 1)
            self.assertEqual(result.skipped[0].path.name, "broken.xlsx")
            self.assertIn("读取失败", result.skipped[0].reason)


class ExcelCompatibilityContractTests(unittest.TestCase):
    def test_macos_automation_help_uses_the_real_monterey_settings_path(self) -> None:
        with (
            patch.object(xls_converter.platform, "system", return_value="Darwin"),
            patch.object(xls_converter.platform, "mac_ver", return_value=("12.7.6", (), "")),
        ):
            self.assertEqual(
                xls_converter.macos_excel_automation_privacy_path(),
                "系统偏好设置 > 安全性与隐私 > 隐私 > 自动化",
            )
        with (
            patch.object(xls_converter.platform, "system", return_value="Darwin"),
            patch.object(xls_converter.platform, "mac_ver", return_value=("13.0", (), "")),
        ):
            self.assertEqual(
                xls_converter.macos_excel_automation_privacy_path(),
                "系统设置 > 隐私与安全性 > 自动化",
            )


class ExcelTaskRunnerContractTests(unittest.TestCase):
    def _run_patches(
        self,
        stack: ExitStack,
        *,
        root: Path,
        engine: _PreflightEngine,
        texts_by_path: dict[Path, list[str]],
        lookup_side_effect=None,
        writer_side_effect=None,
    ) -> MagicMock:
        writer = stack.enter_context(
            patch(
                "core.task_runner.bilingual_writer.write_bilingual_file",
                side_effect=writer_side_effect
                or (lambda **kwargs: root / "out" / f"{kwargs['source_path'].stem}.xlsx"),
            )
        )
        stack.enter_context(
            patch("core.task_runner.TaskLogger", return_value=MagicMock(task_id="phase4-contract"))
        )
        stack.enter_context(
            patch("core.task_runner.check_translation_api_config", return_value=ApiConfigCheckResult(ok=True))
        )
        stack.enter_context(patch("core.task_runner.build_engine", return_value=engine))
        stack.enter_context(patch("core.task_runner.get_system_prompt", return_value="system"))
        stack.enter_context(patch("core.task_runner.resolve_effective_model_config", return_value=object()))
        stack.enter_context(
            patch(
                "core.task_runner.get_model_throughput",
                return_value=EffectiveModelThroughput(
                    profile_key="phase4", batch_size=10, concurrency=1
                ),
            )
        )
        stack.enter_context(
            patch(
                "core.task_runner.bilingual_writer.build_output_dir",
                return_value=root / "out",
            )
        )

        def collect(path: Path, *_args, **_kwargs):
            return list(texts_by_path[path]), 1

        stack.enter_context(patch.object(TaskRunner, "_collect_texts", side_effect=collect))
        stack.enter_context(
            patch(
                "core.task_runner.tm_manager.lookup_batch",
                side_effect=lookup_side_effect
                or (lambda texts, _pair: {text: "TM translation" for text in texts}),
            )
        )
        stack.enter_context(patch("core.task_runner.tm_manager.insert_batch", return_value=0))
        stack.enter_context(patch("core.task_runner.translate_texts", return_value={}))
        return writer

    def test_auto_preflight_is_once_per_file_bounded_and_uses_actual_tm_pairs(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            zh_file = root / "zh.xlsx"
            en_file = root / "en.xlsx"
            long_secret = "仅用于验证不发送完整文件" * 800
            texts = {
                zh_file: [long_secret, "项目名称"],
                en_file: ["The concrete slab requires inspection."],
            }
            engine = _PreflightEngine(
                {
                    long_secret[:240]: '{"source_langs":["zh"]}',
                    "The concrete slab requires inspection.": '{"source_langs":["en"]}',
                }
            )
            lookup_values_by_pair: dict[str, set[str]] = {}

            def lookup(values: list[str], pair: str) -> dict[str, str]:
                lookup_values_by_pair.setdefault(pair, set()).update(values)
                return {value: "TM translation" for value in values}

            with ExitStack() as stack:
                writer = self._run_patches(
                    stack,
                    root=root,
                    engine=engine,
                    texts_by_path=texts,
                    lookup_side_effect=lookup,
                )
                runner = TaskRunner(
                    [
                        FileItem(path=zh_file, name="zh", size_kb=1.0),
                        FileItem(path=en_file, name="en", size_kb=1.0),
                    ],
                    _settings(source_lang="auto", target_lang="fr"),
                    source_root=root,
                )
                runner._run()

            self.assertEqual(len(engine.calls), 2)
            payloads = [json.loads(user) for _system, user in engine.calls]
            all_samples = {
                sample for payload in payloads for sample in payload["samples"]
            }
            self.assertIn(long_secret[:240], all_samples)
            self.assertIn("The concrete slab requires inspection.", all_samples)
            self.assertTrue(all(long_secret not in user for _system, user in engine.calls))
            self.assertTrue(all(len(user) < 3000 for _system, user in engine.calls))
            self.assertEqual(set(lookup_values_by_pair), {"zh-fr", "en-fr"})
            self.assertEqual(
                lookup_values_by_pair["zh-fr"], {long_secret, "项目名称"}
            )
            self.assertEqual(
                lookup_values_by_pair["en-fr"],
                {"The concrete slab requires inspection."},
            )
            self.assertFalse(any(pair.startswith("auto-") for pair in lookup_values_by_pair))
            self.assertEqual(writer.call_count, 2)
            done = _done_message(runner)
            self.assertEqual(len(done.file_results), 2)
            self.assertEqual([item["status"] for item in done.files], ["succeeded", "succeeded"])
            self.assertEqual(done.kpi["selected_file_count"], 2)
            self.assertEqual(done.kpi["succeeded_file_count"], 2)
            self.assertEqual(done.kpi["failed_file_count"], 0)
            self.assertEqual(done.kpi["unstarted_file_count"], 0)
            self.assertEqual(done.kpi["model_translation_text_count"], 0)
            self.assertEqual(done.language["mode"], "automatic")
            self.assertEqual(
                {item["preflight"]["source_langs"][0] for item in done.language["files"]},
                {"zh", "en"},
            )

    def test_high_fidelity_xls_permission_failure_never_silently_falls_back(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            legacy = root / "legacy.xls"
            healthy = root / "healthy.xlsx"
            legacy.touch()
            healthy.touch()
            engine = _PreflightEngine()
            with ExitStack() as stack:
                writer = self._run_patches(
                    stack,
                    root=root,
                    engine=engine,
                    texts_by_path={healthy: ["施工内容"]},
                )
                excel_app = MagicMock()
                excel_app.books.open.side_effect = RuntimeError(
                    "not authorized to send apple events"
                )
                stack.enter_context(patch("core.task_runner.initialize_excel_thread", return_value=None))
                stack.enter_context(patch("core.task_runner.finalize_excel_thread"))
                stack.enter_context(patch("core.task_runner.create_excel_app", return_value=excel_app))
                fallback = stack.enter_context(patch("core.xls_converter.convert_with_fallback"))
                runner = TaskRunner(
                    [
                        FileItem(path=legacy, name="legacy", size_kb=1.0),
                        FileItem(path=healthy, name="healthy", size_kb=1.0),
                    ],
                    _settings(),
                    source_root=root,
                    allow_xls_fallback=False,
                )
                runner._run()

            fallback.assert_not_called()
            writer.assert_called_once()
            by_source = {
                item["source_path"]: item for item in _done_message(runner).file_results
            }
            self.assertFalse(by_source[str(legacy)]["success"])
            self.assertIn("自动化", by_source[str(legacy)]["error"])
            self.assertTrue(by_source[str(healthy)]["success"])

    def test_xls_compatibility_conversion_requires_explicit_confirmation(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            legacy = root / "legacy.xls"
            converted = root / "converted.xlsx"
            legacy.touch()
            converted.touch()
            engine = _PreflightEngine()
            with ExitStack() as stack:
                writer = self._run_patches(
                    stack,
                    root=root,
                    engine=engine,
                    texts_by_path={converted: ["施工内容"]},
                )
                fallback = stack.enter_context(
                    patch("core.xls_converter.convert_with_fallback", return_value=converted)
                )
                runner = TaskRunner(
                    [FileItem(path=legacy, name="legacy", size_kb=1.0)],
                    _settings(),
                    source_root=root,
                    allow_xls_fallback=True,
                )
                runner._run()

            fallback.assert_called_once_with(legacy)
            writer.assert_called_once()
            self.assertTrue(_done_message(runner).file_results[0]["success"])

    def test_excel_review_colors_are_not_read_from_word_settings(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.xlsx"
            source.touch()
            settings = _settings()
            expected_excel_colors = {MIXED_MARK_UNRESOLVED: "123456"}
            settings_payload = settings.model_dump()
            settings_payload["excel_review"] = {
                **settings.excel_review.model_dump(),
                "mark_colors": expected_excel_colors,
            }
            settings_payload["word_review"] = {
                "mark_colors": {MIXED_MARK_UNRESOLVED: "ABCDEF"}
            }
            settings = AppSettings(**settings_payload)
            engine = _PreflightEngine()
            with ExitStack() as stack:
                writer = self._run_patches(
                    stack,
                    root=root,
                    engine=engine,
                    texts_by_path={source: ["施工内容"]},
                )
                runner = TaskRunner(
                    [FileItem(path=source, name="source", size_kb=1.0)],
                    settings,
                    source_root=root,
                )
                runner._run()

            applied_colors = writer.call_args.kwargs["review_mark_colors"]
            self.assertEqual(
                applied_colors[MIXED_MARK_UNRESOLVED],
                expected_excel_colors[MIXED_MARK_UNRESOLVED],
            )
            self.assertNotEqual(
                applied_colors[MIXED_MARK_UNRESOLVED],
                settings.word_review.mark_colors[MIXED_MARK_UNRESOLVED],
            )

    def test_auto_tm_writes_only_model_reported_language_inside_file_preflight(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "fr.xlsx"
            source.touch()
            french_one = "Bonjour le monde"
            french_two = "Texte complémentaire"
            engine = _PreflightEngine(
                {
                    french_one: '{"source_langs":["fr"]}',
                    french_two: '{"source_langs":["fr"]}',
                }
            )
            inserted: list[tuple[str, list[tuple[str, str]]]] = []

            def capture_insert(entries, pair, *_args, **_kwargs):
                inserted.append((pair, list(entries)))
                return len(entries)

            with ExitStack() as stack:
                self._run_patches(
                    stack,
                    root=root,
                    engine=engine,
                    texts_by_path={source: [french_one, french_two]},
                    lookup_side_effect=lambda texts, _pair: {text: None for text in texts},
                )
                stack.enter_context(
                    patch(
                        "core.task_runner.translate_texts_with_sources",
                        return_value={
                            french_one: TranslationLanguageResult(
                                french_one,
                                "Hello world",
                                source_lang="fr",
                                target_lang="zh",
                                tm_eligible=True,
                            ),
                            french_two: TranslationLanguageResult(
                                french_two,
                                "Supplementary text",
                                source_lang="en",
                                target_lang="zh",
                                tm_eligible=True,
                            ),
                        },
                    )
                )
                stack.enter_context(
                    patch("core.tm_manager.insert_batch", side_effect=capture_insert)
                )
                runner = TaskRunner(
                    [FileItem(path=source, name="fr", size_kb=1.0)],
                    _settings(source_lang="auto", target_lang="zh"),
                    source_root=root,
                )
                runner._run()

            self.assertEqual(inserted, [("fr-zh", [(french_one, "Hello world")])])
            self.assertTrue(_done_message(runner).file_results[0]["success"])

    def test_stop_before_scanning_preserves_sources_and_records_unstarted_files(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            first = root / "first.xlsx"
            second = root / "second.xlsx"
            first.write_bytes(b"first source bytes")
            second.write_bytes(b"second source bytes")
            source_hashes = {
                path: hashlib.sha256(path.read_bytes()).hexdigest()
                for path in (first, second)
            }
            engine = _PreflightEngine()
            with ExitStack() as stack:
                self._run_patches(
                    stack,
                    root=root,
                    engine=engine,
                    texts_by_path={},
                )
                runner = TaskRunner(
                    [
                        FileItem(path=first, name="first", size_kb=1.0),
                        FileItem(path=second, name="second", size_kb=1.0),
                    ],
                    _settings(),
                    source_root=root,
                )
                runner.stop()
                runner._run()

            stopped = [
                message
                for message in list(runner._queue.queue)
                if isinstance(message, StoppedMsg)
            ]
            self.assertEqual(len(stopped), 1)
            self.assertEqual(
                [item["status"] for item in stopped[0].files],
                ["unstarted", "unstarted"],
            )
            self.assertEqual(
                [item["source_path"] for item in stopped[0].files],
                [str(first), str(second)],
            )
            self.assertEqual(stopped[0].kpi["unstarted_file_count"], 2)
            self.assertEqual(
                {
                    path: hashlib.sha256(path.read_bytes()).hexdigest()
                    for path in (first, second)
                },
                source_hashes,
            )

    def test_result_contract_aggregates_file_review_locations(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.xlsx"
            source.touch()
            engine = _PreflightEngine()

            def writer_with_review_position(**kwargs):
                kwargs["review_positions"].append(
                    {
                        "worksheet": "报价",
                        "cell": "C7",
                        "category": MIXED_MARK_UNRESOLVED,
                        "action": "preserved_existing_fill",
                    }
                )
                return root / "out" / "source.xlsx"

            with ExitStack() as stack:
                self._run_patches(
                    stack,
                    root=root,
                    engine=engine,
                    texts_by_path={source: ["施工内容"]},
                    writer_side_effect=writer_with_review_position,
                )
                runner = TaskRunner(
                    [FileItem(path=source, name="source", size_kb=1.0)],
                    _settings(),
                    source_root=root,
                )
                runner._run()

            done = _done_message(runner)
            self.assertEqual(done.file_results[0]["review_count"], 1)
            self.assertEqual(
                done.review["items"],
                [
                    {
                        "file": "source.xlsx",
                        "worksheet": "报价",
                        "cell": "C7",
                        "category": MIXED_MARK_UNRESOLVED,
                        "action": "preserved_existing_fill",
                    }
                ],
            )
            self.assertEqual(done.review["counts"], {MIXED_MARK_UNRESOLVED: 1})


if __name__ == "__main__":
    unittest.main(verbosity=2)
