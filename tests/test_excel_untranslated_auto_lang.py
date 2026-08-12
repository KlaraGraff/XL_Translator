"""源语言选「自动识别」时，补译清单必须按识别出来的那门语言算。

自动识别的答案要等词条提出来、问过模型一轮才有；补译判定又必须先知道源语言是
哪一门。这两件事的先后一旦搞反，待补清单会是空的，任务照样「成功」，输出的却是
一份一个字都没翻的文件——这正是用户报的那个 bug。

这里特意用法语稿：只有「先识别、后建清单」才认得出来。要是又退回成「提取时先
按默认的中文判一遍」，法语单元格会被判成「不是源文」，清单立刻空掉。
"""
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook

from core.api_config_check import ApiConfigCheckResult
from core.file_scanner import FileItem
from core.language_preflight import LanguagePreflightResult
from core.model_throughput import EffectiveModelThroughput
from core.task_runner import TaskRunner
from engines.base_engine import TranslationEngine
from settings import AppSettings, EngineSettings


class _FakeEngine(TranslationEngine):
    @property
    def engine_name(self) -> str:
        return "fake/auto-lang"

    def translate_batch(
        self,
        texts: list[str],
        target_lang: str,
        system_prompt: str,
        source_lang: str = "zh",
    ) -> dict[str, str]:
        return {text: f"translated:{text}" for text in texts}


class ExcelUntranslatedAutoLangTests(unittest.TestCase):
    def test_auto_source_lang_still_produces_a_fill_in_plan(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            source = root / "source.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet"
            ws["A1"] = "La dalle de béton est prête\nThe concrete slab is ready"
            ws["A2"] = "Le mur de béton est prêt"
            wb.save(source)
            wb.close()

            settings = AppSettings(
                engine=EngineSettings(
                    mode="cloud",
                    cloud_provider="custom_openai",
                    cloud_model="fake-model",
                    cloud_base_url="https://example.invalid/v1",
                    concurrency=1,
                    batch_size=20,
                ),
                target_lang="en",
                source_lang="auto",
            )

            captured: dict[str, object] = {}

            def fake_write(**kwargs):
                captured["plan"] = kwargs["plan"]
                captured["translations"] = kwargs["translations"]
                captured["source_lang"] = kwargs["source_lang"]
                return root / "out" / "双语(英文)_source.xlsx"

            with (
                patch(
                    "core.task_runner.TaskLogger",
                    return_value=MagicMock(task_id="auto-lang"),
                ),
                patch(
                    "core.task_runner.check_translation_api_config",
                    return_value=ApiConfigCheckResult(ok=True),
                ),
                patch("core.task_runner.build_engine", return_value=_FakeEngine()),
                patch("core.task_runner.get_system_prompt", return_value="system"),
                patch(
                    "core.task_runner.resolve_effective_model_config",
                    return_value=object(),
                ),
                patch(
                    "core.task_runner.get_model_throughput",
                    return_value=EffectiveModelThroughput(
                        profile_key="test",
                        batch_size=20,
                        concurrency=1,
                    ),
                ),
                # 语言预检的样本来自阶段 1 的全量提取，这里直接给出「识别为法语」。
                patch(
                    "core.task_runner.preflight_files",
                    return_value={
                        str(source): LanguagePreflightResult(
                            source_langs=("fr",),
                            requested=True,
                        )
                    },
                ),
                patch("core.task_runner.tm_manager.lookup_batch", return_value={}),
                patch("core.task_runner.tm_manager.insert_batch", return_value=0),
                patch(
                    "core.task_runner.translate_texts",
                    return_value={"Le mur de béton est prêt": "The concrete wall is ready"},
                ),
                patch(
                    "core.task_runner.bilingual_writer.build_output_dir",
                    return_value=root / "out",
                ),
                patch(
                    "core.task_runner.write_untranslated_excel_file",
                    side_effect=fake_write,
                ),
            ):
                runner = TaskRunner(
                    [FileItem(path=source, name="source", size_kb=1.0)],
                    settings,
                    source_root=root,
                    untranslated_only=True,
                )
                runner._run()

            plan = captured.get("plan")
            self.assertIsNotNone(plan, "补译模式没有走到按位置写入这一步")
            self.assertEqual(
                [unit.location for unit in plan.source_units],
                ["Sheet!A2"],
                "自动识别下补译清单为空，等于输出一份没翻译的文件",
            )
            # 计划是在预检定下源语言之后才建的，写入时用的也得是那门语言。
            self.assertEqual(captured["source_lang"], "fr")
            # 只有待补的那一格进了翻译清单，已经双语的 A1 不重翻。
            self.assertEqual(set(captured["translations"]), {"Le mur de béton est prêt"})


if __name__ == "__main__":
    unittest.main(verbosity=2)
