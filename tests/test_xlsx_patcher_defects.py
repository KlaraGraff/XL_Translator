"""审计报告 1.2 / 1.3 / 4.8 / 4.9 / 4.10 的回归测试。

全部走真实的 openpyxl 工作簿 + ``write_bilingual_file`` 端到端路径，断言落在
输出文件的 XML 上，而不是辅助函数的返回值。补丁式写入的字节保真承诺（除被改的
分表 XML / drawing / styles.xml 外其余部件原样照抄）在每个场景里一并守住。
"""
from __future__ import annotations

import hashlib
import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from lxml import etree
from openpyxl import Workbook, load_workbook

from core import bilingual_writer
from core.bilingual_writer import write_bilingual_file
from core.xlsx_patcher import MAX_CELL_TEXT_LEN, NS_MAIN, NS_XDR

NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

# 已经明确不该被重写的部件之外，允许变动的只有这几个。
ALLOWED_MODIFIED_PARTS = {
    "xl/worksheets/sheet1.xml",
    "xl/worksheets/sheet2.xml",
    "xl/drawings/drawing1.xml",
    "xl/styles.xml",
}


def _part_hashes(path: Path) -> dict[str, str]:
    with zipfile.ZipFile(path) as archive:
        return {
            name: hashlib.sha256(archive.read(name)).hexdigest()
            for name in archive.namelist()
        }


def _read_part(path: Path, name: str) -> bytes:
    with zipfile.ZipFile(path) as archive:
        return archive.read(name)


def _sheet_rows(path: Path, part: str = "xl/worksheets/sheet1.xml") -> dict[str, object]:
    root = etree.fromstring(_read_part(path, part))
    return {row.get("r"): row for row in root.iter(f"{{{NS_MAIN}}}row")}


class _PatcherDefectCase(unittest.TestCase):
    def setUp(self) -> None:
        self._temp = tempfile.TemporaryDirectory()
        self.addCleanup(self._temp.cleanup)
        self.root = Path(self._temp.name)
        self.out_dir = self.root / "out"
        self.logs: list[str] = []

    def _write(self, source: Path, translations: dict[str, str], **overrides) -> Path:
        kwargs = dict(
            source_path=source,
            output_dir=self.out_dir,
            translations=translations,
            target_lang="en",
            source_lang="zh",
            keep_original_sheets=False,
            formula_display_value_backfill=False,
            enable_print_guard=False,
            log_callback=self.logs.append,
        )
        kwargs.update(overrides)
        return write_bilingual_file(**kwargs)

    def assert_only_expected_parts_changed(self, source: Path, output: Path) -> None:
        before = _part_hashes(source)
        after = _part_hashes(output)
        self.assertEqual(set(before) - set(after), set(), "输出包丢失了原有部件")
        changed = {name for name, digest in before.items() if after[name] != digest}
        self.assertTrue(
            changed <= ALLOWED_MODIFIED_PARTS,
            f"补丁式写入被破坏，意外重写了部件：{sorted(changed - ALLOWED_MODIFIED_PARTS)}",
        )


class CellTextGuardTests(_PatcherDefectCase):
    """1.2 超 32767 字符 / 1.3 非法 XML 字符。"""

    def _single_cell_workbook(self, value: str) -> Path:
        path = self.root / "long.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "报价"
        sheet["A1"] = value
        workbook.save(path)
        workbook.close()
        return path

    def test_oversized_cell_is_truncated_and_reported(self) -> None:
        source_text = "施" * 20000
        translation = "Construction " * 2000  # 26000 字符
        source = self._single_cell_workbook(source_text)
        stats: dict[str, object] = {}

        output = self._write(source, {source_text: translation}, stats=stats)

        with zipfile.ZipFile(output) as archive:  # 包本身仍然可读
            self.assertIn("xl/worksheets/sheet1.xml", archive.namelist())
        workbook = load_workbook(output)
        try:
            written = workbook["报价"]["A1"].value
        finally:
            workbook.close()

        self.assertEqual(
            len(written),
            MAX_CELL_TEXT_LEN,
            "写入的单元格长度必须正好卡在 Excel 上限，多一个字符文件就打不开",
        )
        self.assertTrue(written.startswith(source_text[:100]))
        self.assertTrue(written.endswith("…"), "截断处要留可见记号")

        # 截断绝不能是静默的：统计出参和任务日志都要点名分表与坐标。
        self.assertEqual(stats["truncated_cells"], 1)
        self.assertEqual(stats["truncated_positions"], ["报价!A1"])
        truncation_logs = [line for line in self.logs if "截断" in line]
        self.assertTrue(truncation_logs, f"没有任何截断提示：{self.logs}")
        self.assertTrue(any("报价!A1" in line for line in truncation_logs))
        self.assertTrue(any(str(MAX_CELL_TEXT_LEN) in line for line in truncation_logs))

        self.assert_only_expected_parts_changed(source, output)

    def test_cell_within_limit_reports_no_truncation(self) -> None:
        source = self._single_cell_workbook("施工内容")
        stats: dict[str, object] = {}
        output = self._write(source, {"施工内容": "Construction scope"}, stats=stats)

        workbook = load_workbook(output)
        try:
            self.assertEqual(
                workbook["报价"]["A1"].value, "施工内容\nConstruction scope"
            )
        finally:
            workbook.close()
        self.assertEqual(stats["truncated_cells"], 0)
        self.assertEqual(stats["truncated_positions"], [])
        self.assertFalse([line for line in self.logs if "截断" in line])

    def test_illegal_xml_characters_do_not_abort_the_write(self) -> None:
        source = self._single_cell_workbook("施工内容")
        # \x0b 是 XML 1.0 不允许的码点，lxml 赋值时会直接抛 ValueError。
        output = self._write(
            source, {"施工内容": "Construction\x0bscope\x00end\x1f!"}
        )

        workbook = load_workbook(output)
        try:
            written = workbook["报价"]["A1"].value
        finally:
            workbook.close()
        self.assertEqual(written, "施工内容\nConstructionscopeend!")
        self.assertTrue(any("XML 非法字符" in line for line in self.logs), self.logs)
        self.assert_only_expected_parts_changed(source, output)

    def test_legal_whitespace_survives_sanitizing(self) -> None:
        source = self._single_cell_workbook("施工内容")
        output = self._write(source, {"施工内容": "a\tb\nc"})
        workbook = load_workbook(output)
        try:
            self.assertEqual(workbook["报价"]["A1"].value, "施工内容\na\tb\nc")
        finally:
            workbook.close()

    def test_failed_patch_leaves_no_output_file(self) -> None:
        """补丁失败时输出目录必须干净——半成品副本的文件名和成品一模一样。"""
        source = self._single_cell_workbook("施工内容")
        self.out_dir.mkdir(parents=True, exist_ok=True)

        with mock.patch.object(
            bilingual_writer,
            "write_bilingual_workbook",
            side_effect=ValueError("boom"),
        ):
            with self.assertRaises(ValueError):
                self._write(source, {"施工内容": "Construction scope"})

        self.assertEqual(
            sorted(p.name for p in self.out_dir.iterdir()),
            [],
            "补丁失败后输出目录里不能残留任何文件（包括未翻译的副本和临时件）",
        )


class GroupedShapeTests(_PatcherDefectCase):
    """4.8 组合图形不能被整组尺寸压平。"""

    CHILD_EXTENTS = (("111", "222"), ("333", "444"))

    def _workbook_with_grouped_drawing(self) -> Path:
        base = self.root / "base.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "报价"
        # 第 1 行会被翻译 → 行高变化 → 锚点需要冻结。
        sheet["A1"] = "很长很长的施工说明文字需要换行才放得下所以行高会被自动调整"
        workbook.save(base)
        workbook.close()

        target = self.root / "grouped.xlsx"
        self._inject_drawing(base, target)
        return target

    def _inject_drawing(self, source: Path, target: Path) -> None:
        with zipfile.ZipFile(source) as archive:
            parts = {name: archive.read(name) for name in archive.namelist()}

        sheet_xml = etree.fromstring(parts["xl/worksheets/sheet1.xml"])
        drawing_el = etree.SubElement(sheet_xml, f"{{{NS_MAIN}}}drawing")
        drawing_el.set(f"{{{REL_NS}}}id", "rIdDraw")
        parts["xl/worksheets/sheet1.xml"] = etree.tostring(
            sheet_xml, xml_declaration=True, encoding="UTF-8", standalone=True
        )

        parts["xl/worksheets/_rels/sheet1.xml.rels"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<Relationships xmlns="{PKG_REL_NS}">'
            f'<Relationship Id="rIdDraw" Type="{REL_NS}/drawing"'
            f' Target="../drawings/drawing1.xml"/></Relationships>'
        ).encode()

        children = "".join(
            f'<xdr:sp><xdr:nvSpPr><xdr:cNvPr id="{index + 2}" name="Shape {index}"/>'
            f"<xdr:cNvSpPr/></xdr:nvSpPr><xdr:spPr>"
            f'<a:xfrm><a:off x="0" y="0"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm>'
            f'<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr></xdr:sp>'
            for index, (cx, cy) in enumerate(self.CHILD_EXTENTS)
        )
        parts["xl/drawings/drawing1.xml"] = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<xdr:wsDr xmlns:xdr="{NS_XDR}" xmlns:a="{NS_A}" xmlns:r="{REL_NS}">'
            f'<xdr:twoCellAnchor editAs="oneCell">'
            f"<xdr:from><xdr:col>3</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>0</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
            f"<xdr:to><xdr:col>5</xdr:col><xdr:colOff>0</xdr:colOff>"
            f"<xdr:row>3</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
            f"<xdr:grpSp><xdr:nvGrpSpPr>"
            f'<xdr:cNvPr id="1" name="Group 1"/><xdr:cNvGrpSpPr/></xdr:nvGrpSpPr>'
            f'<xdr:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="10" cy="20"/>'
            f'<a:chOff x="0" y="0"/><a:chExt cx="10" cy="20"/></a:xfrm></xdr:grpSpPr>'
            f"{children}</xdr:grpSp>"
            f"<xdr:clientData/></xdr:twoCellAnchor></xdr:wsDr>"
        ).encode()

        content_types = etree.fromstring(parts["[Content_Types].xml"])
        override = etree.SubElement(content_types, f"{{{CT_NS}}}Override")
        override.set("PartName", "/xl/drawings/drawing1.xml")
        override.set(
            "ContentType",
            "application/vnd.openxmlformats-officedocument.drawing+xml",
        )
        parts["[Content_Types].xml"] = etree.tostring(
            content_types, xml_declaration=True, encoding="UTF-8", standalone=True
        )

        with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as archive:
            for name, data in parts.items():
                archive.writestr(name, data)

    def test_child_shapes_keep_their_own_size(self) -> None:
        source = self._workbook_with_grouped_drawing()
        output = self._write(
            source,
            {
                "很长很长的施工说明文字需要换行才放得下所以行高会被自动调整": (
                    "A very long construction note that needs wrapping"
                )
            },
        )

        drawing = etree.fromstring(_read_part(output, "xl/drawings/drawing1.xml"))
        anchors = list(drawing)
        self.assertEqual(len(anchors), 1)
        self.assertEqual(etree.QName(anchors[0]).localname, "oneCellAnchor")

        anchor_ext = anchors[0].find(f"{{{NS_XDR}}}ext")
        anchor_cx = anchor_ext.get("cx")
        anchor_cy = anchor_ext.get("cy")

        group = anchors[0].find(f"{{{NS_XDR}}}grpSp")
        group_xfrm = group.find(f"{{{NS_XDR}}}grpSpPr/{{{NS_A}}}xfrm")
        group_ext = group_xfrm.find(f"{{{NS_A}}}ext")
        self.assertEqual(
            (group_ext.get("cx"), group_ext.get("cy")),
            (anchor_cx, anchor_cy),
            "组合图形的外框仍要跟着锚点走，否则渲染器读到 0 尺寸",
        )
        # chExt 保持原值，整组才是等比缩放而不是被压平。
        child_extent = group_xfrm.find(f"{{{NS_A}}}chExt")
        self.assertEqual((child_extent.get("cx"), child_extent.get("cy")), ("10", "20"))

        child_sizes = [
            (
                shape.find(f"{{{NS_XDR}}}spPr/{{{NS_A}}}xfrm/{{{NS_A}}}ext").get("cx"),
                shape.find(f"{{{NS_XDR}}}spPr/{{{NS_A}}}xfrm/{{{NS_A}}}ext").get("cy"),
            )
            for shape in group.findall(f"{{{NS_XDR}}}sp")
        ]
        self.assertEqual(
            child_sizes,
            list(self.CHILD_EXTENTS),
            "组内子形状各自的尺寸必须原样保留，被覆盖成整组尺寸就是图形被毁",
        )


class RowHeightScopeTests(_PatcherDefectCase):
    """4.9 未翻译的行不许改行高 / 4.10 公式源码不参与行高估算。"""

    LONG_SOURCE = "很长很长的施工说明文字需要换行才放得下所以行高会被自动调整"
    LONG_TARGET = "A very long construction note that definitely needs wrapping"
    UNTRANSLATED = "这一行本次一个字都没翻但是它同样很长会被估算成需要换很多行"

    def _workbook(self) -> Path:
        path = self.root / "rows.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "报价"
        sheet["A1"] = self.LONG_SOURCE
        sheet["A2"] = self.UNTRANSLATED
        sheet.row_dimensions[2].height = 20  # openpyxl 会自动带上 customHeight
        workbook.save(path)
        workbook.close()
        return path

    def test_untranslated_row_keeps_its_manual_height(self) -> None:
        source = self._workbook()
        output = self._write(source, {self.LONG_SOURCE: self.LONG_TARGET})

        rows = _sheet_rows(output)
        translated_row = rows["1"]
        self.assertEqual(translated_row.get("customHeight"), "1")
        self.assertGreater(
            float(translated_row.get("ht")),
            15.0,
            "被翻译的行仍然要按双语内容加高",
        )

        untouched_row = rows["2"]
        self.assertEqual(
            float(untouched_row.get("ht")),
            20.0,
            "本次一字未翻的行，手工设的行高必须原样保留",
        )
        self.assert_only_expected_parts_changed(source, output)

    def test_nothing_translated_leaves_the_sheet_byte_identical(self) -> None:
        source = self._workbook()
        output = self._write(source, {})

        before = _part_hashes(source)
        after = _part_hashes(output)
        self.assertEqual(
            after["xl/worksheets/sheet1.xml"],
            before["xl/worksheets/sheet1.xml"],
            "一个单元格都没改时分表 XML 不该被重写",
        )

    def test_formula_source_does_not_inflate_row_height(self) -> None:
        path = self.root / "formula.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "报价"
        sheet["A1"] = "配电箱"
        # 长嵌套公式：源码 400+ 字符，显示出来只是一个数字。
        sheet["B1"] = "=IF(" + "+".join(f"C{i}" for i in range(1, 120)) + ",1,2)"
        workbook.save(path)
        workbook.close()

        output = self._write(path, {"配电箱": "Distribution box"})

        row = _sheet_rows(output)["1"]
        height = float(row.get("ht"))
        # "配电箱\nDistribution box" 在默认列宽下约 3 行；公式源码若参与估算
        # 会算出几十行、把行高顶到 400pt 以上。
        self.assertLess(
            height,
            80.0,
            f"行高被公式源码撑到了 {height}pt，公式不该按源码文本参与换行估算",
        )
        self.assertGreater(height, 15.0)


if __name__ == "__main__":
    unittest.main()
