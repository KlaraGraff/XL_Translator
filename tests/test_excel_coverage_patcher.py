"""core.excel_coverage 补译写入路径迁移到补丁式写入器后的契约测试。

覆盖点：
  1. ``allowed_positions`` 坐标限定：计划外的同文本单元格不被改写；
  2. 文本匹配守卫：单元格实际文本与计划记录不符时跳过；
  3. 嵌入图片部件（WPS cellimages.xml / Excel vm= 富数据）补译后逐字节保留；
  4. ``keep_original_sheets`` 克隆的 ``_原文`` 分表存在且内容为原文。
"""
from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from lxml import etree
from openpyxl import load_workbook

from core.excel_coverage import ExcelCoveragePlan, write_untranslated_excel_file
from core.translation_coverage import COVERAGE_SOURCE_ONLY, CoverageUnit
from core.xlsx_patcher import NS_MAIN
from tests.test_xlsx_patcher import (
    PNG_BYTES,
    RICH_VALUE_PART,
    WPS_CELLIMAGES,
    XlsxPatcherFixture,
    _read_part,
)


def _plan(units: list[CoverageUnit], *, sheet_count: int = 2) -> ExcelCoveragePlan:
    return ExcelCoveragePlan(path=Path("unused"), units=units, sheet_count=sheet_count)


def _source_only(sheet: str, coordinate: str, text: str) -> CoverageUnit:
    return CoverageUnit(
        source_text=text,
        status=COVERAGE_SOURCE_ONLY,
        location=f"{sheet}!{coordinate}",
        reason="test",
        data={"sheet": sheet, "coordinate": coordinate},
    )


class ExcelCoveragePatcherTests(unittest.TestCase):
    def setUp(self) -> None:
        self._temp = tempfile.TemporaryDirectory()
        self.root = Path(self._temp.name)
        # 复用 test_xlsx_patcher 的工件构造：两张分表都有 "配电箱"，
        # 外加一个 DISPIMG 公式单元格与一个 vm= 富数据单元格（同样是 "配电箱"）。
        self.fixture = XlsxPatcherFixture.build(self.root)

    def tearDown(self) -> None:
        self._temp.cleanup()

    # ── 坐标限定 ──────────────────────────────────────────────────────────
    def test_allowed_positions_excludes_duplicate_text_outside_plan(self) -> None:
        plan = _plan([_source_only("报价", "A3", "配电箱")])
        logs: list[str] = []

        out_path = write_untranslated_excel_file(
            source_path=self.fixture,
            output_dir=self.root / "out",
            plan=plan,
            translations={"配电箱": "Distribution box"},
            target_lang="en",
            source_lang="zh",
            keep_original_sheets=False,
            log_callback=logs.append,
        )

        wb = load_workbook(out_path)
        try:
            self.assertEqual(wb["报价"]["A3"].value, "配电箱\nDistribution box")
            # 附表!A1 文本完全相同，但不在计划限定坐标内，不应被改写。
            self.assertEqual(wb["附表"]["A1"].value, "配电箱")
        finally:
            wb.close()

        self.assertTrue(any("补译 1 个单元格" in msg for msg in logs), logs)

    # ── 文本匹配守卫 ──────────────────────────────────────────────────────
    def test_stale_cell_text_guard_skips_write(self) -> None:
        # 计划记录 A1 的原文是"配电箱"，但 A1 的实际内容是"施工内容"——
        # 模拟计划构建之后单元格发生了漂移，写入时必须重新核对并跳过。
        plan = _plan([_source_only("报价", "A1", "配电箱")])

        out_path = write_untranslated_excel_file(
            source_path=self.fixture,
            output_dir=self.root / "out",
            plan=plan,
            translations={"配电箱": "Distribution box"},
            target_lang="en",
            source_lang="zh",
            keep_original_sheets=False,
        )

        wb = load_workbook(out_path)
        try:
            self.assertEqual(wb["报价"]["A1"].value, "施工内容")
        finally:
            wb.close()

    # ── 嵌入图片保真 ──────────────────────────────────────────────────────
    def test_embedded_image_parts_survive_supplementary_write(self) -> None:
        # 手工构造计划，故意把 DISPIMG 公式单元格（A5）和 vm= 富数据单元格（B5）
        # 的坐标也列进"待补译"范围——补丁写入器必须整格跳过它们，不能因为坐标
        # 在计划限定范围内就当成普通文本单元格改写，这正是主翻译路径此前修复的
        # 图片丢失 bug 在补译路径上的等价场景。
        plan = _plan(
            [
                _source_only("报价", "A1", "施工内容"),
                _source_only(
                    "报价", "A2", "很长很长的施工说明文字需要换行才放得下所以行高会被自动调整"
                ),
                _source_only("报价", "A3", "配电箱"),
                _source_only("报价", "A5", "配电箱"),
                _source_only("报价", "B5", "配电箱"),
                _source_only("报价", "C5", "配电箱"),
            ]
        )

        expected_shared_strings = _read_part(self.fixture, "xl/sharedStrings.xml")

        out_path = write_untranslated_excel_file(
            source_path=self.fixture,
            output_dir=self.root / "out",
            plan=plan,
            translations={
                "施工内容": "Construction scope",
                "很长很长的施工说明文字需要换行才放得下所以行高会被自动调整": (
                    "A very long construction note"
                ),
                "配电箱": "Distribution box",
            },
            target_lang="en",
            source_lang="zh",
            keep_original_sheets=False,
        )

        self.assertEqual(_read_part(out_path, "xl/cellimages.xml"), WPS_CELLIMAGES)
        self.assertEqual(_read_part(out_path, "xl/richData/rdrichvalue.xml"), RICH_VALUE_PART)
        self.assertEqual(_read_part(out_path, "xl/media/image1.png"), PNG_BYTES)
        self.assertEqual(
            _read_part(out_path, "xl/sharedStrings.xml"),
            expected_shared_strings,
            "回填走 inlineStr，共享字符串表不应被改写",
        )

        root = etree.fromstring(_read_part(out_path, "xl/worksheets/sheet1.xml"))
        cells = {cell.get("r"): cell for cell in root.iter(f"{{{NS_MAIN}}}c")}

        dispimg = cells["A5"]
        self.assertEqual(
            dispimg.findtext(f"{{{NS_MAIN}}}f"), '_xlfn.DISPIMG("ID_1",1)'
        )
        self.assertEqual(dispimg.findtext(f"{{{NS_MAIN}}}v"), "配电箱")

        rich = cells["B5"]
        self.assertEqual(rich.get("vm"), "1")
        self.assertEqual(rich.get("t"), "s")
        self.assertEqual(rich.findtext(f"{{{NS_MAIN}}}v"), "0")

        # 同样的原文在普通共享字符串单元格里必须照常补译。
        self.assertEqual(cells["C5"].get("t"), "inlineStr")
        self.assertEqual(
            "".join(cells["C5"].itertext()).strip(), "配电箱\nDistribution box"
        )

    # ── 原文分表克隆 ──────────────────────────────────────────────────────
    def test_keep_original_sheets_clones_original_text(self) -> None:
        plan = _plan([_source_only("报价", "A1", "施工内容")])

        out_path = write_untranslated_excel_file(
            source_path=self.fixture,
            output_dir=self.root / "out",
            plan=plan,
            translations={"施工内容": "Construction scope"},
            target_lang="en",
            source_lang="zh",
            keep_original_sheets=True,
        )

        wb = load_workbook(out_path)
        try:
            self.assertIn("报价_原文", wb.sheetnames)
            self.assertEqual(wb["报价_原文"]["A1"].value, "施工内容")
            self.assertEqual(wb["报价"]["A1"].value, "施工内容\nConstruction scope")
        finally:
            wb.close()


if __name__ == "__main__":
    unittest.main(verbosity=2)
