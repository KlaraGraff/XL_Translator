"""core.file_scanner 对嵌入图片 / 含文字文本框的扫描期计数测试。

背景：Excel 翻译流水线只翻译单元格文字，工作簿里的嵌入图片和文本框/形状文字
从未被翻译过，写入器逐字节原样保留。这组测试只验证扫描阶段能不能把它们数
出来（为后续界面提示提供数据），不涉及翻译或写回。

覆盖：
  - 含浮动图片的 .xlsx（drawingN.xml 里的 <xdr:pic>）
  - 含带文字文本框的 .xlsx（drawingN.xml 里的 <xdr:sp> + 非空 <a:t>），
    纯装饰性形状（无文字）不应计入
  - WPS xl/cellimages.xml + Excel 富数据图片（只算 _localImage，股票/地理位置
    等非图片富数据不能算；结构表缺失时退回 richValueRel.xml 的 <rel> 数）
  - 干净的 .xlsx（没有任何图片/文本框）计数应为 0，而不是 None
  - .xls：扫描阶段拿不到这些信息，必须是 None，不能猜成 0
  - 解析失败（drawing 部件损坏）不应让整个文件的扫描失败
"""
from __future__ import annotations

import tempfile
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from core.file_scanner import scan_excel_sources
from core.xlsx_patcher import NS_A, NS_XDR

REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
RICH_DATA_NS = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"


def _build_base_xlsx(path: Path) -> None:
    """用 openpyxl 生成一个最普通的单表工作簿，作为注入自定义部件的底子。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    wb.save(path)
    wb.close()


def _inject_parts(path: Path, parts: dict[str, bytes]) -> None:
    """往已有 .xlsx 里新增/覆盖若干 zip 部件（openpyxl 本身不认识这些部件）。"""
    with zipfile.ZipFile(path) as archive:
        existing = {name: archive.read(name) for name in archive.namelist()}
    existing.update(parts)
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as archive:
        for name, data in existing.items():
            archive.writestr(name, data)


def _anchor(idx: int, inner: str) -> str:
    return (
        '<xdr:twoCellAnchor editAs="oneCell">'
        "<xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff>"
        "<xdr:row>0</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>"
        "<xdr:to><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff>"
        "<xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>"
        f"{inner}"
        "<xdr:clientData/></xdr:twoCellAnchor>"
    )


def _pic_anchor(idx: int) -> str:
    inner = (
        "<xdr:pic><xdr:nvPicPr>"
        f'<xdr:cNvPr id="{idx}" name="Picture {idx}"/><xdr:cNvPicPr/></xdr:nvPicPr>'
        f'<xdr:blipFill><a:blip r:embed="rIdImg{idx}"/>'
        "<a:stretch><a:fillRect/></a:stretch></xdr:blipFill>"
        '<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>'
        '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr></xdr:pic>'
    )
    return _anchor(idx, inner)


def _text_shape_anchor(idx: int) -> str:
    inner = (
        '<xdr:sp macro="" textlink="">'
        f'<xdr:nvSpPr><xdr:cNvPr id="{idx}" name="TextBox {idx}"/>'
        '<xdr:cNvSpPr txBox="1"/></xdr:nvSpPr>'
        '<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>'
        '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
        "<xdr:txBody><a:bodyPr/><a:p><a:r><a:t>需要翻译的文字</a:t></a:r></a:p></xdr:txBody>"
        "</xdr:sp>"
    )
    return _anchor(idx, inner)


def _decorative_shape_anchor(idx: int) -> str:
    """无文字的纯装饰形状：没有 txBody，不应计入 shape_text_count。"""
    inner = (
        '<xdr:sp macro="" textlink="">'
        f'<xdr:nvSpPr><xdr:cNvPr id="{idx}" name="Decoration {idx}"/>'
        "<xdr:cNvSpPr/></xdr:nvSpPr>"
        '<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/></a:xfrm>'
        '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
        "</xdr:sp>"
    )
    return _anchor(idx, inner)


def _drawing_xml(*, pics: int = 0, text_shapes: int = 0, empty_shapes: int = 0) -> bytes:
    anchors: list[str] = []
    idx = 1
    for _ in range(pics):
        anchors.append(_pic_anchor(idx))
        idx += 1
    for _ in range(text_shapes):
        anchors.append(_text_shape_anchor(idx))
        idx += 1
    for _ in range(empty_shapes):
        anchors.append(_decorative_shape_anchor(idx))
        idx += 1
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<xdr:wsDr xmlns:xdr="{NS_XDR}" xmlns:a="{NS_A}" xmlns:r="{REL_NS}">'
        + "".join(anchors)
        + "</xdr:wsDr>"
    ).encode()


def _cellimages_xml(count: int) -> bytes:
    entries = "".join(f"<etc:cellImage>ID_{i}</etc:cellImage>" for i in range(1, count + 1))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<etc:cellImages xmlns:etc="http://www.wps.cn/officeDocument/2017/etCustomData">'
        + entries
        + "</etc:cellImages>"
    ).encode()


def _rich_value_parts(*, images: int = 0, non_images: int = 0) -> dict[str, bytes]:
    """一组富数据部件：结构表 0 号是本地图片，1 号是股票之类的非图片富数据。

    <rv s="N"> 的 N 指向结构表下标，只有 _localImage 才是「置于单元格内」的图片。
    """
    structure = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<rvStructures xmlns="{RICH_DATA_NS}" count="2">'
        '<s t="_localImage"><k n="_rvRel:LocalImageIdentifier" t="i"/></s>'
        '<s t="_linkedEntity2"><k n="_Self" t="spr"/></s>'
        "</rvStructures>"
    ).encode()
    entries = "".join(f'<rv s="0"><v>{i}</v></rv>' for i in range(images))
    entries += "".join(f'<rv s="1"><v>{i}</v></rv>' for i in range(non_images))
    values = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<rvData xmlns="{RICH_DATA_NS}" count="{images + non_images}">'
        + entries
        + "</rvData>"
    ).encode()
    return {
        "xl/richData/rdrichvalue.xml": values,
        "xl/richData/rdrichvaluestructure.xml": structure,
    }


def _rich_value_rel_xml(count: int) -> bytes:
    entries = "".join(f'<rel r:id="rId{i + 1}"/>' for i in range(count))
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<richValueRels xmlns="{RICH_DATA_NS}" xmlns:r="{REL_NS}">'
        + entries
        + "</richValueRels>"
    ).encode()


SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
THREADED_COMMENT_NS = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"


def _legacy_comments_xml(refs: list[str]) -> bytes:
    """传统批注部件（Excel 里叫「注释」）：<comment ref="B2"> 一个格子一条。"""
    entries = "".join(
        f'<comment ref="{ref}" authorId="0"><text><r><t>note {ref}</t></r></text></comment>'
        for ref in refs
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<comments xmlns="{SPREADSHEET_NS}"><authors><author>tester</author></authors>'
        f"<commentList>{entries}</commentList></comments>"
    ).encode()


def _threaded_comments_xml(entries: list[tuple[str, bool]]) -> bytes:
    """新版对话式批注部件。第二项为 True 表示这条是回复（挂在同一个格子上）。"""
    body = "".join(
        f'<threadedComment ref="{ref}" dT="2026-01-01T00:00:00Z" personId="{{p}}" '
        f'id="{{c{index}}}"' + (' parentId="{c0}"' if is_reply else "") + ">"
        f"<text>comment {index}</text></threadedComment>"
        for index, (ref, is_reply) in enumerate(entries)
    )
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<ThreadedComments xmlns="{THREADED_COMMENT_NS}">{body}</ThreadedComments>'
    ).encode()


class _FakeXlsSheet:
    nrows = 0

    def row_values(self, row: int) -> list[str]:
        return []


class _FakeXlsWorkbook:
    """最小 xlrd workbook 桩：只提供 _build_file_item 的 .xls 分支需要的接口。"""

    def sheet_names(self) -> list[str]:
        return ["Sheet1"]

    @property
    def nsheets(self) -> int:
        return 1

    def sheet_by_index(self, index: int) -> _FakeXlsSheet:
        return _FakeXlsSheet()

    def unload_sheet(self, index: int) -> None:
        pass

    def release_resources(self) -> None:
        pass


class ExcelScanEmbeddedContentTests(unittest.TestCase):
    def setUp(self) -> None:
        self._temp = tempfile.TemporaryDirectory()
        self.root = Path(self._temp.name)

    def tearDown(self) -> None:
        self._temp.cleanup()

    def _scan_single(self, path: Path):
        result = scan_excel_sources(path)
        self.assertEqual(result.skipped, [], f"文件不应被判定为损坏：{result.skipped}")
        self.assertEqual(len(result.items), 1)
        return result.items[0], result

    def test_floating_images_are_counted(self) -> None:
        path = self.root / "images.xlsx"
        _build_base_xlsx(path)
        _inject_parts(path, {"xl/drawings/drawing1.xml": _drawing_xml(pics=2)})

        item, result = self._scan_single(path)
        self.assertEqual(item.image_count, 2)
        self.assertEqual(item.shape_text_count, 0)
        self.assertEqual(result.summary["image_count"], 2)
        self.assertEqual(result.summary["image_count_unknown_files"], 0)

    def test_text_shapes_are_counted_and_decorative_ones_excluded(self) -> None:
        path = self.root / "textboxes.xlsx"
        _build_base_xlsx(path)
        _inject_parts(
            path,
            {"xl/drawings/drawing1.xml": _drawing_xml(text_shapes=2, empty_shapes=1)},
        )

        item, result = self._scan_single(path)
        self.assertEqual(item.image_count, 0)
        self.assertEqual(item.shape_text_count, 2)
        self.assertEqual(result.summary["shape_text_count"], 2)
        self.assertEqual(result.summary["shape_text_count_unknown_files"], 0)

    def test_wps_cellimages_and_rich_data_images_are_counted(self) -> None:
        path = self.root / "cellimages.xlsx"
        _build_base_xlsx(path)
        parts: dict[str, bytes] = {"xl/cellimages.xml": _cellimages_xml(2)}
        # 4 条股票之类的非图片富数据混在里面，一条都不能算进图片数。
        parts.update(_rich_value_parts(images=3, non_images=4))
        _inject_parts(path, parts)

        item, _result = self._scan_single(path)
        # cellimages(2 张) + 富数据里的本地图片(3 张)
        self.assertEqual(item.image_count, 5)
        self.assertEqual(item.shape_text_count, 0)

    def test_non_image_rich_data_alone_is_not_counted_as_images(self) -> None:
        path = self.root / "stocks.xlsx"
        _build_base_xlsx(path)
        _inject_parts(path, _rich_value_parts(non_images=200))

        item, _result = self._scan_single(path)
        self.assertEqual(item.image_count, 0)

    def test_rich_data_without_structure_part_falls_back_to_rels(self) -> None:
        path = self.root / "rels_fallback.xlsx"
        _build_base_xlsx(path)
        # 缺结构表（或结构表解析失败）时，退回按 richValueRel.xml 的 <rel> 数计。
        parts = _rich_value_parts(images=3, non_images=4)
        parts.pop("xl/richData/rdrichvaluestructure.xml")
        parts["xl/richData/richValueRel.xml"] = _rich_value_rel_xml(2)
        _inject_parts(path, parts)

        item, _result = self._scan_single(path)
        self.assertEqual(item.image_count, 2)

    def test_clean_workbook_has_zero_not_none(self) -> None:
        path = self.root / "clean.xlsx"
        _build_base_xlsx(path)

        item, result = self._scan_single(path)
        self.assertEqual(item.image_count, 0)
        self.assertEqual(item.shape_text_count, 0)
        self.assertEqual(result.summary["image_count"], 0)
        self.assertEqual(result.summary["shape_text_count"], 0)
        self.assertEqual(result.summary["image_count_unknown_files"], 0)
        self.assertEqual(result.summary["shape_text_count_unknown_files"], 0)

    def test_xls_counts_are_unknown_none_not_zero(self) -> None:
        path = self.root / "legacy.xls"
        path.write_bytes(b"legacy-xls-placeholder")

        with patch("xlrd.open_workbook", return_value=_FakeXlsWorkbook()):
            item, result = self._scan_single(path)

        self.assertIsNone(item.image_count)
        self.assertIsNone(item.shape_text_count)
        # 汇总不能把 None 当 0 加；要能看出「1 个文件数不出来」。
        self.assertEqual(result.summary["image_count"], 0)
        self.assertEqual(result.summary["image_count_unknown_files"], 1)
        self.assertEqual(result.summary["shape_text_count"], 0)
        self.assertEqual(result.summary["shape_text_count_unknown_files"], 1)

    def test_legacy_comments_are_counted_per_cell(self) -> None:
        path = self.root / "notes.xlsx"
        _build_base_xlsx(path)
        _inject_parts(path, {"xl/comments1.xml": _legacy_comments_xml(["B2", "C7"])})

        item, result = self._scan_single(path)
        self.assertEqual(item.comment_count, 2)
        self.assertEqual(result.summary["comment_count"], 2)
        self.assertEqual(result.summary["comment_count_unknown_files"], 0)

    def test_a_threaded_comment_and_its_legacy_mirror_count_once(self) -> None:
        # Excel 存新版对话式批注时会同时写一份传统批注做兼容。两边都数就会把
        # 用户表里的 1 条批注报成 2 条——这个数字要出现在界面上，不能虚报。
        path = self.root / "threaded.xlsx"
        _build_base_xlsx(path)
        _inject_parts(
            path,
            {
                "xl/comments1.xml": _legacy_comments_xml(["B2"]),
                "xl/threadedComments/threadedComment1.xml": _threaded_comments_xml(
                    [("B2", False), ("B2", True), ("D4", False)],
                ),
            },
        )

        item, _result = self._scan_single(path)
        # B2 一条（含一条回复）+ D4 一条 = 2
        self.assertEqual(item.comment_count, 2)

    def test_clean_workbook_has_zero_comments_not_none(self) -> None:
        path = self.root / "clean_comments.xlsx"
        _build_base_xlsx(path)

        item, result = self._scan_single(path)
        self.assertEqual(item.comment_count, 0)
        self.assertEqual(result.summary["comment_count_unknown_files"], 0)

    def test_xls_comment_count_is_unknown_none_not_zero(self) -> None:
        path = self.root / "legacy_comments.xls"
        path.write_bytes(b"legacy-xls-placeholder")

        with patch("xlrd.open_workbook", return_value=_FakeXlsWorkbook()):
            item, result = self._scan_single(path)

        self.assertIsNone(item.comment_count)
        self.assertEqual(result.summary["comment_count"], 0)
        self.assertEqual(result.summary["comment_count_unknown_files"], 1)

    def test_corrupt_comments_part_does_not_fail_the_scan(self) -> None:
        path = self.root / "corrupt_comments.xlsx"
        _build_base_xlsx(path)
        _inject_parts(path, {"xl/comments1.xml": b"<not-well-formed-xml"})

        item, _result = self._scan_single(path)
        self.assertEqual(item.comment_count, 0)

    def test_corrupt_drawing_part_does_not_fail_the_scan(self) -> None:
        path = self.root / "corrupt_drawing.xlsx"
        _build_base_xlsx(path)
        _inject_parts(path, {"xl/drawings/drawing1.xml": b"<not-well-formed-xml"})

        item, _result = self._scan_single(path)
        self.assertEqual(item.image_count, 0)
        self.assertEqual(item.shape_text_count, 0)


if __name__ == "__main__":
    unittest.main()
