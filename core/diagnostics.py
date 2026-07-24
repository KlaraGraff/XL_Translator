"""Lightweight diagnostic archives for failed translation tasks."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import platform
import re
import shutil
import sys
import zipfile
from dataclasses import asdict, is_dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from config import APP_DATA_DIR, APP_VERSION, LOG_PATH

DIAGNOSTICS_DIR = APP_DATA_DIR / "diagnostics"
DIAGNOSTIC_RECORDS_DIR = DIAGNOSTICS_DIR / "records"

_SECRET_KEY_RE = re.compile(
    r"(api[_-]?key|token|secret|password|authorization|credential)",
    re.IGNORECASE,
)
_BEARER_RE = re.compile(r"(Bearer\s+)[A-Za-z0-9._~+/=-]{8,}", re.IGNORECASE)
_ASSIGNMENT_SECRET_RE = re.compile(
    r"((?:api[_-]?key|token|secret|password)\s*[=:]\s*)[^\s,;]+",
    re.IGNORECASE,
)
_URL_SECRET_RE = re.compile(
    r"([?&](?:api[_-]?key|key|token|secret|password)=)[^&\s]+",
    re.IGNORECASE,
)
_SOURCE_EXCERPT_LIMIT = 220
_APP_LOG_TAIL_CHARS = 160_000
_DIAGNOSTIC_MAX_RECORDS = 80
_DIAGNOSTIC_MAX_TOTAL_BYTES = 256 * 1024 * 1024
_WORD_RUNTIME_EVENT_MARKERS = (
    "正在单段重试",
    "单段重试恢复",
    "单段重试未恢复",
    "正在语义仲裁",
    "语义仲裁接受",
    "语义仲裁未接受",
    "保留原文，需复核",
)


def archive_task_diagnostics(
    *,
    surface: str,
    phase: str,
    task_id: str,
    settings: Any,
    selected_files: list[Any],
    logs: list[dict[str, Any]],
    done: Any | None = None,
    error_message: str = "",
    source_root: str | Path | None = None,
    status: str = "",
    progress: Any | None = None,
    task_artifacts: dict[str, Any] | None = None,
) -> Path:
    """Persist a strictly anonymous support record.

    Diagnostics deliberately do not contain document names, paths, cell or
    paragraph locations, source/target content, prompts, model responses, API
    keys, or a copied ``app.log``.  User-facing reports remain the only local
    files that may retain document-specific positioning information.
    """
    del source_root, task_artifacts
    record_dir = _unique_record_dir(surface=surface, task_id=task_id)
    (record_dir / "environment").mkdir(parents=True, exist_ok=True)

    done_payload = _json_safe(done)
    file_results = list(done_payload.get("file_results") or []) if isinstance(done_payload, dict) else []
    quality_issues = list(done_payload.get("issues") or []) if isinstance(done_payload, dict) else []
    failure_count = sum(1 for item in file_results if not bool(item.get("success")))
    metrics = {
        "file_count": len(selected_files),
        "failed_file_count": failure_count,
        "quality_issue_count": len(quality_issues),
        "runtime_log_count": len(logs),
        "progress_event_count": 1 if progress is not None else 0,
        "log_levels": _count_log_levels(logs),
    }
    manifest = {
        "record_id": record_dir.name,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "surface": _safe_category(surface, fallback="task"),
        "phase": _safe_category(phase, fallback="unknown"),
        "anonymous_locator": _anonymous_locator(task_id),
        "status": _safe_category(status, fallback="unknown"),
        "error_code": _safe_error_code(error_message),
        **metrics,
        "content_categories": [
            "runtime",
            "task_type_and_stage",
            "anonymous_counts",
            "redacted_connection_summary",
        ],
    }

    _write_json(record_dir / "manifest.json", manifest)
    _write_json(record_dir / "metrics.json", metrics)
    _write_json(
        record_dir / "environment" / "runtime.json",
        {
            **_build_runtime_payload(),
            "connection_summary": _connection_summary(settings),
        },
    )

    prune_diagnostic_records()

    return record_dir


def list_diagnostic_records() -> list[dict[str, Any]]:
    """Return archived diagnostic records, newest first."""
    records: list[dict[str, Any]] = []
    if not DIAGNOSTIC_RECORDS_DIR.exists():
        return records

    for manifest_path in DIAGNOSTIC_RECORDS_DIR.glob("*/manifest.json"):
        try:
            payload = json.loads(manifest_path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if not _is_current_safe_manifest(payload):
            # A prior format may contain document-specific locations.  Keep
            # it on disk until the user explicitly clears diagnostics, but
            # never surface or export it through the current application.
            continue
        payload["record_dir"] = str(manifest_path.parent)
        payload["size_bytes"] = estimate_record_size(manifest_path.parent)
        records.append(payload)

    records.sort(key=lambda item: str(item.get("created_at") or ""), reverse=True)
    return records


def count_diagnostic_records() -> int:
    """Return the number of persisted diagnostic records."""
    return len(list_diagnostic_records())


def public_diagnostic_records() -> list[dict[str, Any]]:
    """Return diagnostic summaries that are safe to render or expose over API."""
    allowed = {
        "record_id",
        "created_at",
        "surface",
        "phase",
        "anonymous_locator",
        "status",
        "error_code",
        "file_count",
        "failed_file_count",
        "quality_issue_count",
        "runtime_log_count",
        "size_bytes",
        "content_categories",
    }
    return [
        {key: value for key, value in record.items() if key in allowed}
        for record in list_diagnostic_records()
    ]


def diagnostic_overview() -> dict[str, Any]:
    records = list_diagnostic_records()
    return {
        "count": len(records),
        "total_size_bytes": sum(int(item.get("size_bytes") or 0) for item in records),
        "max_records": _DIAGNOSTIC_MAX_RECORDS,
        "max_total_bytes": _DIAGNOSTIC_MAX_TOTAL_BYTES,
        "content_categories": [
            "runtime",
            "task_type_and_stage",
            "anonymous_counts",
            "redacted_connection_summary",
        ],
    }


def find_diagnostic_record(record_id: str) -> dict[str, Any] | None:
    return next(
        (item for item in list_diagnostic_records() if item.get("record_id") == record_id),
        None,
    )


def delete_diagnostic_record(record_id: str) -> bool:
    record = find_diagnostic_record(record_id)
    if record is None:
        return False
    record_dir = Path(str(record.get("record_dir") or ""))
    if not _is_record_directory(record_dir):
        return False
    shutil.rmtree(record_dir, ignore_errors=False)
    return True


def clear_diagnostic_records() -> int:
    removed = 0
    if not DIAGNOSTIC_RECORDS_DIR.exists():
        return removed
    for candidate in DIAGNOSTIC_RECORDS_DIR.iterdir():
        if not candidate.is_dir() or not _is_record_directory(candidate):
            continue
        shutil.rmtree(candidate, ignore_errors=False)
        removed += 1
    return removed


def _is_current_safe_manifest(payload: object) -> bool:
    if not isinstance(payload, dict):
        return False
    required = {"record_id", "anonymous_locator", "content_categories"}
    forbidden = {
        "source_root",
        "output_dir",
        "error_message",
        "task_id",
        "pdf_manifest_path",
        "pdf_report_path",
    }
    return required.issubset(payload) and not any(key in payload for key in forbidden)


def record_system_diagnostic(*, phase: str, error_code: str) -> Path:
    """Record a non-task failure such as a release check without raw details."""
    return archive_task_diagnostics(
        surface="system",
        phase=phase,
        task_id=f"system-{phase}",
        settings={},
        selected_files=[],
        logs=[],
        status="error",
        error_message=error_code,
    )


def prune_diagnostic_records(
    *,
    max_records: int = _DIAGNOSTIC_MAX_RECORDS,
    max_total_bytes: int = _DIAGNOSTIC_MAX_TOTAL_BYTES,
) -> None:
    """Keep diagnostic history bounded while preserving the newest record."""
    records = list_diagnostic_records()
    kept_size = 0
    for index, record in enumerate(records):
        record_dir = Path(str(record.get("record_dir") or ""))
        size = int(record.get("size_bytes") or 0)
        keep_by_count = index < max_records
        keep_by_size = kept_size + size <= max_total_bytes
        if index == 0 or (keep_by_count and keep_by_size):
            kept_size += size
            continue
        if record_dir.exists():
            shutil.rmtree(record_dir, ignore_errors=True)


def build_diagnostic_zip_bytes(record_dir: str | Path) -> tuple[bytes, str]:
    """Build one diagnostic record zip in memory."""
    root = Path(record_dir)
    if not root.exists():
        raise FileNotFoundError(f"诊断归档不存在：{root}")
    filename = f"{root.name}.zip"
    return _zip_directory(root, prefix=root.name), filename


def build_diagnostics_history_zip_bytes() -> tuple[bytes, str, int]:
    """Build a single zip containing all persisted diagnostic records."""
    records = list_diagnostic_records()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"xl_translator_diagnostics_history_{timestamp}.zip"
    buffer = BytesIO()

    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        summary_rows = []
        for record in records:
            record_dir = Path(str(record.get("record_dir") or ""))
            if not record_dir.exists():
                continue
            summary_rows.append(
                {
                    "record_id": record.get("record_id", record_dir.name),
                    "created_at": record.get("created_at", ""),
                    "surface": record.get("surface", ""),
                    "phase": record.get("phase", ""),
                    "failed_file_count": record.get("failed_file_count", 0),
                    "quality_issue_count": record.get("quality_issue_count", 0),
                    "size_bytes": record.get("size_bytes", 0),
                }
            )
            _add_directory_to_zip(zf, record_dir, Path("records") / record_dir.name)

        zf.writestr(
            "history_summary.csv",
            _csv_text(
                summary_rows,
                [
                    "record_id",
                    "created_at",
                    "surface",
                    "phase",
                    "failed_file_count",
                    "quality_issue_count",
                    "size_bytes",
                ],
            ),
        )

    return buffer.getvalue(), filename, len(records)


def estimate_record_size(record_dir: str | Path) -> int:
    """Return the total size of files under a diagnostic record directory."""
    root = Path(record_dir)
    if not root.exists():
        return 0
    return sum(path.stat().st_size for path in root.rglob("*") if path.is_file())


def format_size(num_bytes: int) -> str:
    """Format byte counts for UI captions."""
    value = float(max(0, int(num_bytes or 0)))
    for unit in ("B", "KB", "MB", "GB"):
        if value < 1024 or unit == "GB":
            if unit == "B":
                return f"{int(value)} {unit}"
            return f"{value:.1f} {unit}"
        value /= 1024
    return f"{value:.1f} GB"


def _unique_record_dir(*, surface: str, task_id: str) -> Path:
    DIAGNOSTIC_RECORDS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_surface = _safe_filename(surface or "task")
    base = f"{timestamp}_{safe_surface}_{_anonymous_locator(task_id)}"
    candidate = DIAGNOSTIC_RECORDS_DIR / base
    suffix = 2
    while candidate.exists():
        candidate = DIAGNOSTIC_RECORDS_DIR / f"{base}_{suffix}"
        suffix += 1
    candidate.mkdir(parents=True, exist_ok=False)
    return candidate


def _safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", str(value or "").strip())
    return cleaned.strip("._")[:80] or "diagnostic"


def _safe_category(value: object, *, fallback: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value or "").strip().lower())
    return cleaned.strip("_")[:48] or fallback


def _anonymous_locator(task_id: object) -> str:
    raw = str(task_id or "").encode("utf-8", errors="ignore")
    return hashlib.sha256(raw).hexdigest()[:16]


def _safe_error_code(message: object) -> str:
    """Persist a bounded category instead of an exception or provider response."""
    raw = str(message or "").strip().lower()
    if not raw:
        return "none"
    if "429" in raw or "rate" in raw:
        return "rate_limited"
    if "timeout" in raw or "timed out" in raw:
        return "timeout"
    if "auth" in raw or "401" in raw or "403" in raw:
        return "authentication"
    if "network" in raw or "connect" in raw:
        return "network"
    if "schema" in raw or "json" in raw:
        return "data_validation"
    return "task_failed"


def _count_log_levels(logs: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for log in logs:
        level = _safe_category(log.get("level") if isinstance(log, dict) else "", fallback="info")
        counts[level] = counts.get(level, 0) + 1
    return counts


def _connection_summary(settings: Any) -> list[dict[str, str]]:
    """Keep provider role labels, never model names, endpoints, or credentials."""
    result: list[dict[str, str]] = []
    roles = (
        ("translation", getattr(settings, "engine", None)),
        ("cleaner", getattr(settings, "cleaner_model_role", None)),
        ("image", getattr(settings, "image_model_role", None)),
        ("pdf_review", getattr(settings, "pdf_review_model_role", None)),
    )
    for role, value in roles:
        if value is None:
            continue
        result.append(
            {
                "role": role,
                "mode": _safe_category(getattr(value, "mode", "cloud"), fallback="unknown"),
                "provider": _safe_category(
                    getattr(value, "cloud_provider", ""),
                    fallback="unknown",
                ),
            }
        )
    return result


def _is_record_directory(path: Path) -> bool:
    try:
        return path.resolve().parent == DIAGNOSTIC_RECORDS_DIR.resolve()
    except OSError:
        return False


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Path):
        return str(value)
    if is_dataclass(value):
        return _json_safe(asdict(value))
    if hasattr(value, "model_dump"):
        try:
            return _json_safe(value.model_dump(mode="json"))
        except TypeError:
            return _json_safe(value.model_dump())
    if isinstance(value, dict):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]
    return str(value)


def _redact(value: Any) -> Any:
    if isinstance(value, dict):
        result = {}
        for key, item in value.items():
            if _SECRET_KEY_RE.search(str(key)):
                result[key] = "***" if item else ""
            else:
                result[key] = _redact(item)
        return result
    if isinstance(value, list):
        return [_redact(item) for item in value]
    if isinstance(value, str):
        text = _BEARER_RE.sub(r"\1***", value)
        text = _ASSIGNMENT_SECRET_RE.sub(r"\1***", text)
        return _URL_SECRET_RE.sub(r"\1***", text)
    return value


def _serialize_file_item(item: Any) -> dict[str, Any]:
    path = Path(getattr(item, "path", ""))
    payload = {
        "name": str(getattr(item, "name", path.stem or "")),
        "filename": path.name,
        "path": str(path),
        "suffix": path.suffix.lower(),
        "exists": path.exists() if str(path) else False,
        "size_kb": getattr(item, "size_kb", None),
        "sheets": list(getattr(item, "sheets", []) or []),
        "paragraph_count": getattr(item, "paragraph_count", None),
        "table_count": getattr(item, "table_count", None),
        "translatable_count": getattr(item, "translatable_count", None),
    }
    return {key: value for key, value in payload.items() if value not in (None, [], "")}


def _build_pdf_diagnostic_summary(
    *,
    output_dir: str,
    file_results: list[dict[str, Any]],
    quality_issues: list[dict[str, Any]],
    artifacts: dict[str, Any],
) -> dict[str, Any]:
    manifest_path = str(artifacts.get("manifest_path") or "")
    report_path = str(artifacts.get("report_path") or "")
    if not manifest_path and output_dir:
        candidate = Path(output_dir) / "pdf_translation_manifest.json"
        if candidate.exists():
            manifest_path = str(candidate)
    if not report_path and output_dir:
        candidate = Path(output_dir) / "pdf_translation_report.md"
        if candidate.exists():
            report_path = str(candidate)

    summary: dict[str, Any] = {
        "output_dir": output_dir,
        "manifest_path": manifest_path,
        "report_path": report_path,
        "generated_pdf_count": sum(1 for item in file_results if item.get("success")),
        "placeholder_page_count": sum(
            int(item.get("placeholder_page_count") or 0) for item in file_results
        ),
        "emergency_ratio_normalized_count": sum(
            int(item.get("emergency_ratio_normalized_count") or 0)
            for item in file_results
        ),
        "quality_issue_count": len(quality_issues),
    }
    if manifest_path and Path(manifest_path).exists():
        try:
            package_manifest = json.loads(Path(manifest_path).read_text(encoding="utf-8"))
        except Exception:
            package_manifest = {}
        for key in (
            "status",
            "file_count",
            "total_page_count",
            "generated_pdf_count",
            "placeholder_page_count",
            "emergency_ratio_normalized_count",
            "retry_count",
            "rate_limit_reduction_count",
            "partial_artifacts_available",
        ):
            if key in package_manifest:
                summary[key] = package_manifest[key]
    return _redact(summary)


def _build_runtime_payload() -> dict[str, Any]:
    return {
        "app_version": APP_VERSION,
        "platform": _runtime_platform_label(),
        "python_version": platform.python_version(),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }


def _runtime_platform_label() -> str:
    if os.name == "nt":
        get_windows_version = getattr(sys, "getwindowsversion", None)
        if get_windows_version is None:
            return "Windows"
        version = get_windows_version()
        return f"Windows-{version.major}.{version.minor}.{version.build}"
    return platform.platform()


def _build_summary(
    manifest: dict[str, Any],
    file_results: list[dict[str, Any]],
    quality_issues: list[dict[str, Any]],
) -> str:
    surface_label = {
        "excel": "Excel",
        "word": "Word",
        "pdf": "PDF",
    }.get(str(manifest.get("surface") or ""), "任务")
    lines = [
        f"# {surface_label} 翻译诊断摘要",
        "",
        f"- 归档时间：{manifest.get('created_at', '')}",
        f"- 任务 ID：{manifest.get('task_id', '')}",
        f"- 阶段：{manifest.get('phase', '')}",
        f"- 源路径：{manifest.get('source_root', '')}",
        f"- 输出目录：{manifest.get('output_dir', '')}",
        f"- 文件数：{manifest.get('file_count', 0)}",
        f"- 失败文件数：{manifest.get('failed_file_count', 0)}",
        f"- 质量/API 提示数：{manifest.get('quality_issue_count', 0)}",
    ]
    if manifest.get("error_message"):
        lines.append(f"- 最近错误：{manifest['error_message']}")
    lines.extend(["", "## 失败文件", ""])
    failed = [item for item in file_results if not item.get("success")]
    if failed:
        for item in failed:
            lines.append(f"- {item.get('name') or item.get('file') or '未知文件'}：{item.get('error') or ''}")
    else:
        lines.append("无失败文件。")
    lines.extend(["", "## 质量/API 提示", ""])
    if quality_issues:
        for issue in quality_issues:
            message = issue.get("message") or issue.get("problem") or "未记录"
            lines.append(f"- {message}")
    else:
        lines.append("无质量/API 提示。")
    lines.append("")
    return "\n".join(str(_redact(line)) for line in lines)


def _sanitize_quality_issues(issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    sanitized = []
    for issue in issues:
        item = _redact(_json_safe(issue))
        failed_sources = []
        for source_item in item.get("failed_sources") or []:
            source_text = str(source_item.get("source") or "")
            failed_sources.append(
                {
                    "source_hash": _source_hash(source_text),
                    "source_excerpt": _source_excerpt(source_text),
                    "error": _redact(str(source_item.get("error") or "")),
                }
            )
        if failed_sources:
            item["failed_sources"] = failed_sources
        sanitized.append(item)
    return sanitized


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(_redact(_json_safe(payload)), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(_redact(_json_safe(row)), ensure_ascii=False))
            handle.write("\n")


def _write_text(path: Path, text: str) -> None:
    path.write_text(str(_redact(text)), encoding="utf-8")


def _write_failed_items_csv(
    path: Path,
    file_results: list[dict[str, Any]],
    quality_issues: list[dict[str, Any]],
) -> None:
    rows: list[dict[str, Any]] = []
    for item in file_results:
        if item.get("success"):
            continue
        rows.append(
            {
                "item_type": "file",
                "file": item.get("name", ""),
                "location": "",
                "location_label": "",
                "section_path": "",
                "severity": "failed",
                "problem": "文件处理失败",
                "status": "",
                "error": item.get("error", ""),
                "source_hash": "",
                "source_excerpt": "",
                "output": item.get("output", ""),
            }
        )

    for issue in quality_issues:
        for failed_source in issue.get("failed_sources") or []:
            source_text = str(failed_source.get("source") or "")
            rows.append(
                {
                    "item_type": "api",
                    "file": issue.get("file", ""),
                    "location": "",
                    "location_label": issue.get("location_label", ""),
                    "section_path": issue.get("section_path", ""),
                    "severity": issue.get("severity", "needs_action"),
                    "problem": issue.get("message") or issue.get("problem", ""),
                    "status": issue.get("status", ""),
                    "error": failed_source.get("error", ""),
                    "source_hash": _source_hash(source_text),
                    "source_excerpt": _source_excerpt(source_text),
                    "output": "",
                }
            )
        if not issue.get("failed_sources"):
            rows.append(
                {
                    "item_type": "quality",
                    "file": issue.get("file", ""),
                    "location": issue.get("location", ""),
                    "location_label": issue.get("location_label", ""),
                    "section_path": issue.get("section_path", ""),
                    "severity": issue.get("severity", ""),
                    "problem": issue.get("message") or issue.get("problem", ""),
                    "status": issue.get("status", ""),
                    "error": "",
                    "source_hash": "",
                    "source_excerpt": issue.get("snippet", ""),
                    "output": "",
                }
            )

    _write_csv(
        path,
        rows,
        [
            "item_type",
            "file",
            "location",
            "location_label",
            "section_path",
            "severity",
            "problem",
            "status",
            "error",
            "source_hash",
            "source_excerpt",
            "output",
        ],
    )


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.write_text(_csv_text(rows, fieldnames), encoding="utf-8-sig")


def _csv_text(rows: list[dict[str, Any]], fieldnames: list[str]) -> str:
    buffer = BytesIO()
    text_buffer = _TextBuffer(buffer)
    writer = csv.DictWriter(text_buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({key: _redact(_json_safe(row.get(key, ""))) for key in fieldnames})
    return buffer.getvalue().decode("utf-8")


class _TextBuffer:
    """Small text adapter over BytesIO for csv.DictWriter."""

    def __init__(self, buffer: BytesIO):
        self._buffer = buffer

    def write(self, text: str) -> int:
        data = text.encode("utf-8")
        self._buffer.write(data)
        return len(text)


def _read_app_log_excerpt(*, task_id: str) -> str:
    if not LOG_PATH.exists():
        return "未找到应用日志文件。"

    candidates = sorted(LOG_PATH.parent.glob(f"{LOG_PATH.name}*"))
    lines: list[str] = []
    for path in candidates:
        try:
            text = _read_tail_text(path, _APP_LOG_TAIL_CHARS)
        except Exception as exc:
            lines.append(f"[WARN] 读取日志失败 {path.name}: {exc}")
            continue
        if task_id:
            matched = [line for line in text.splitlines() if f"[task:{task_id}]" in line]
            lines.extend(matched)

    if not lines and task_id:
        return "未找到匹配当前任务 ID 的应用日志片段。"
    if not lines:
        try:
            return _read_tail_text(LOG_PATH, _APP_LOG_TAIL_CHARS)
        except Exception as exc:
            return f"读取应用日志失败：{exc}"
    return "\n".join(_redact(line) for line in lines[-2000:])


def _read_tail_text(path: Path, max_chars: int) -> str:
    text = path.read_text(encoding="utf-8", errors="replace")
    if len(text) <= max_chars:
        return text
    return text[-max_chars:]


def _extract_failed_source_texts(quality_issues: list[dict[str, Any]]) -> list[str]:
    texts: list[str] = []
    seen: set[str] = set()
    for issue in quality_issues:
        for item in issue.get("failed_sources") or []:
            source = str(item.get("source") or "").strip()
            if source and source not in seen:
                texts.append(source)
                seen.add(source)
    return texts


def _collect_excel_locations(file_items: list[Any], source_texts: list[str]) -> list[dict[str, Any]]:
    source_set = set(source_texts)
    rows: list[dict[str, Any]] = []
    for item in file_items:
        path = Path(getattr(item, "path", ""))
        if not path.exists():
            continue
        try:
            if path.suffix.lower() == ".xlsx":
                rows.extend(_collect_xlsx_locations(path, source_set))
            elif path.suffix.lower() == ".xls":
                rows.extend(_collect_xls_locations(path, source_set))
        except Exception as exc:  # noqa: BLE001 - diagnostics must not block export
            rows.append(
                {
                    "file": path.name,
                    "sheet": "",
                    "cell": "",
                    "row": "",
                    "column": "",
                    "source_hash": "",
                    "source_excerpt": f"定位扫描失败：{exc}",
                }
            )
    return rows


def _collect_xlsx_locations(path: Path, source_set: set[str]) -> list[dict[str, Any]]:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    rows: list[dict[str, Any]] = []
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                for column_index, value in enumerate(row, start=1):
                    text = str(value).strip() if isinstance(value, str) else ""
                    if text not in source_set:
                        continue
                    rows.append(
                        {
                            "file": path.name,
                            "sheet": worksheet.title,
                            "cell": f"{get_column_letter(column_index)}{row_index}",
                            "row": row_index,
                            "column": column_index,
                            "source_hash": _source_hash(text),
                            "source_excerpt": _source_excerpt(text),
                        }
                    )
    finally:
        workbook.close()
    return rows


def _collect_xls_locations(path: Path, source_set: set[str]) -> list[dict[str, Any]]:
    import xlrd

    rows: list[dict[str, Any]] = []
    workbook = xlrd.open_workbook(str(path), on_demand=True)
    try:
        for sheet in workbook.sheets():
            for row_index in range(sheet.nrows):
                for column_index in range(sheet.ncols):
                    value = sheet.cell_value(row_index, column_index)
                    text = str(value).strip() if isinstance(value, str) else ""
                    if text not in source_set:
                        continue
                    rows.append(
                        {
                            "file": path.name,
                            "sheet": sheet.name,
                            "cell": f"R{row_index + 1}C{column_index + 1}",
                            "row": row_index + 1,
                            "column": column_index + 1,
                            "source_hash": _source_hash(text),
                            "source_excerpt": _source_excerpt(text),
                        }
                    )
    finally:
        workbook.release_resources()
    return rows


def _extract_word_locations(quality_issues: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for issue in quality_issues:
        rows.append(
            {
                "file": issue.get("file", ""),
                "kind": issue.get("kind", ""),
                "location": issue.get("location", ""),
                "location_label": issue.get("location_label", ""),
                "section_path": issue.get("section_path", ""),
                "severity": issue.get("severity", ""),
                "problem": issue.get("problem", ""),
                "status": issue.get("status", ""),
                "snippet": issue.get("snippet", ""),
            }
        )
    return rows


def _extract_word_runtime_events(logs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in logs:
        message = str(item.get("message") or "")
        if " · " not in message:
            continue
        parts = message.split(" · ", 2)
        if len(parts) != 3:
            continue
        file_name, section_path, location_and_event = parts
        marker_match = None
        marker_index = -1
        for marker in _WORD_RUNTIME_EVENT_MARKERS:
            index = location_and_event.find(marker)
            if index >= 0 and (marker_index < 0 or index < marker_index):
                marker_match = marker
                marker_index = index
        if marker_match is None:
            continue
        rows.append(
            {
                "ts": item.get("ts", ""),
                "level": item.get("level", ""),
                "file": file_name,
                "section_path": section_path,
                "location_label": location_and_event[:marker_index].strip(),
                "event": location_and_event[marker_index:].strip(),
                "message": message,
            }
        )
    return rows


def _source_hash(text: str) -> str:
    return hashlib.sha256(str(text or "").encode("utf-8")).hexdigest()[:16]


def _source_excerpt(text: str) -> str:
    normalized = " ".join(str(text or "").split())
    if len(normalized) <= _SOURCE_EXCERPT_LIMIT:
        return normalized
    half = max(20, (_SOURCE_EXCERPT_LIMIT - 3) // 2)
    return f"{normalized[:half]}...{normalized[-half:]}"


def _zip_directory(root: Path, *, prefix: str) -> bytes:
    buffer = BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        _add_directory_to_zip(zf, root, Path(prefix))
    return buffer.getvalue()


def _add_directory_to_zip(zf: zipfile.ZipFile, root: Path, prefix: Path) -> None:
    for path in sorted(root.rglob("*")):
        if not path.is_file():
            continue
        zf.write(path, prefix / path.relative_to(root))
