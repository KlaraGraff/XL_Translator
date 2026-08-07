"""Excel coverage detection and position-based untranslated-only writing."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from core import bilingual_writer, xlsx_patcher
from core.language_registry import get_target_lang_display
from core.translation_coverage import (
    COVERAGE_AMBIGUOUS,
    COVERAGE_COVERED,
    COVERAGE_IGNORED,
    COVERAGE_SOURCE_ONLY,
    CoverageUnit,
    clean_coverage_text,
    coverage_summary,
    looks_like_source_text,
    looks_like_target_text,
    split_existing_bilingual_text,
)


@dataclass
class ExcelCoveragePlan:
    path: Path
    units: list[CoverageUnit]
    sheet_count: int

    @property
    def source_units(self) -> list[CoverageUnit]:
        return [unit for unit in self.units if unit.status == COVERAGE_SOURCE_ONLY]

    @property
    def source_texts(self) -> list[str]:
        seen: set[str] = set()
        texts: list[str] = []
        for unit in self.source_units:
            source = unit.source_text.strip()
            if source and source not in seen:
                seen.add(source)
                texts.append(source)
        return texts

    @property
    def summary(self) -> dict[str, int]:
        return coverage_summary(self.units)


def build_excel_coverage_plan(
    path: str | Path,
    *,
    target_lang: str,
    source_lang: str = "zh",
    formula_display_value_backfill: bool = True,
) -> ExcelCoveragePlan:
    """Classify app-style bilingual Excel cells by coverage status."""
    from openpyxl import load_workbook

    source_path = Path(path)
    wb = load_workbook(str(source_path), read_only=True, data_only=False)
    wb_values = (
        load_workbook(str(source_path), read_only=True, data_only=True)
        if formula_display_value_backfill
        else None
    )
    units: list[CoverageUnit] = []
    try:
        workbook_sheet_names = set(wb.sheetnames)
        for ws in wb.worksheets:
            if _is_generated_original_sheet(ws.title, workbook_sheet_names):
                continue
            # 公式格的显示值一次性读成内存映射。read_only 工作表的
            # ``ws[coord]`` 每次都要重新解析整张 sheet XML，逐格去问是 O(n²)：
            # 实测 1600 个公式格就要 7.7s，万级公式表直接卡到分钟级。
            values_by_coordinate = (
                _DisplayValues(wb_values[ws.title]) if wb_values is not None else None
            )
            for row in ws.iter_rows():
                for cell in row:
                    raw = _resolve_cell_text(
                        cell, values_by_coordinate, formula_display_value_backfill
                    )
                    if raw is None:
                        continue
                    unit = _classify_excel_cell(
                        raw,
                        sheet_name=ws.title,
                        coordinate=cell.coordinate,
                        source_lang=source_lang,
                        target_lang=target_lang,
                    )
                    if unit is not None:
                        units.append(unit)
        return ExcelCoveragePlan(
            path=source_path,
            units=units,
            sheet_count=len(wb.worksheets),
        )
    finally:
        wb.close()
        if wb_values is not None:
            wb_values.close()


def write_untranslated_excel_file(
    *,
    source_path: str | Path,
    output_dir: str | Path,
    plan: ExcelCoveragePlan,
    translations: dict[str, str],
    target_lang: str,
    source_lang: str = "zh",
    keep_original_sheets: bool = True,
    formula_display_value_backfill: bool = True,
    lock_row_height: bool = False,
    log_callback=None,
    original_path: Path | None = None,
    external_autofit_planned: bool = False,
    stats: dict[str, object] | None = None,
) -> Path:
    """Copy an Excel file and patch translations only at plan-limited source-only positions.

    坐标限定 + 文本匹配双重守卫都交给 ``xlsx_patcher.write_bilingual_workbook`` 的
    ``allowed_positions`` 机制完成：只有 ``plan.source_units`` 里的 (sheet, coordinate)
    才允许改写，且写入时会用当前单元格的实际文本重新核对是否仍匹配 ``translations``
    的键，防止误伤已经变化的单元格或同文本的其它位置。
    """
    source_path = Path(source_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    lang_display = bilingual_writer._sanitize_filename_fragment(
        get_target_lang_display(target_lang, include_optional=True)
    )
    basename = original_path.name if original_path else source_path.name
    if basename.lower().endswith(".xls"):
        basename = basename[:-4] + ".xlsx"
    out_path = output_dir / f"双语({lang_display})_{basename}"

    allowed_positions: dict[str, set[str]] = {}
    scoped_translations: dict[str, str] = {}
    for unit in plan.source_units:
        sheet_name = str(unit.data.get("sheet") or "")
        coordinate = str(unit.data.get("coordinate") or "")
        if not sheet_name or not coordinate:
            continue
        source_key = unit.source_text.strip()
        translation = str(translations.get(source_key) or "").strip()
        if not translation:
            continue
        allowed_positions.setdefault(sheet_name, set()).add(coordinate)
        scoped_translations[source_key] = translation

    # 调用方传了 stats 就直接写进去，省一次拷贝；没传就用本地临时字典。
    if stats is None:
        stats = {}
    # 补丁失败时不能在输出目录里留下一个「看起来正常」的未翻译副本，
    # 所以先在临时文件上打补丁，成功了再改名。见 bilingual_writer.patch_into_output。
    bilingual_writer.patch_into_output(
        source_path,
        out_path,
        lambda staging: xlsx_patcher.write_bilingual_workbook(
            staging,
            translations=scoped_translations,
            target_lang=target_lang,
            source_lang=source_lang,
            keep_original_sheets=keep_original_sheets,
            formula_display_value_backfill=formula_display_value_backfill,
            lock_row_height=lock_row_height,
            mark_review_items=False,
            log_callback=log_callback,
            allowed_positions=allowed_positions,
            external_autofit_planned=external_autofit_planned,
            stats=stats,
        ),
    )
    write_count = stats.get("mutated_cells", 0)

    if log_callback:
        log_callback(f"[OK] 已输出：{out_path.name}（补译 {write_count} 个单元格）")
    return out_path


def _classify_excel_cell(
    raw: str,
    *,
    sheet_name: str,
    coordinate: str,
    source_lang: str,
    target_lang: str,
) -> CoverageUnit | None:
    text = clean_coverage_text(raw)
    if not text:
        return None

    location = f"{sheet_name}!{coordinate}"
    data = {"sheet": sheet_name, "coordinate": coordinate}
    split = split_existing_bilingual_text(
        text,
        source_lang=source_lang,
        target_lang=target_lang,
    )
    if split is not None:
        source, target = split
        return CoverageUnit(
            source_text=source,
            target_text=target,
            status=COVERAGE_COVERED,
            location=location,
            kind="cell",
            reason="同一单元格已包含源文和目标语言译文。",
            data=data,
        )

    if looks_like_source_text(text, source_lang=source_lang, target_lang=target_lang):
        return CoverageUnit(
            source_text=text,
            status=COVERAGE_SOURCE_ONLY,
            location=location,
            kind="cell",
            reason="单元格包含源语言文本，未识别到目标语言译文。",
            data=data,
        )

    if looks_like_target_text(text, source_lang=source_lang, target_lang=target_lang):
        return CoverageUnit(
            source_text="",
            target_text=text,
            status=COVERAGE_IGNORED,
            location=location,
            kind="cell",
            reason="单元格看起来是目标语言译文，默认跳过。",
            data=data,
        )

    if len(text.splitlines()) > 1:
        return CoverageUnit(
            source_text=text,
            status=COVERAGE_AMBIGUOUS,
            location=location,
            kind="cell",
            reason="多行单元格无法可靠拆分源文和译文。",
            data=data,
        )

    return CoverageUnit(
        source_text=text,
        status=COVERAGE_IGNORED,
        location=location,
        kind="cell",
        reason="单元格不符合补译候选规则。",
        data=data,
    )


class _DisplayValues:
    """一张分表的公式显示值，整表读一次、之后按坐标查表。

    openpyxl 的 read_only 工作表不缓存单元格：``ws["B7"]`` 每次都从头重新解析
    这张分表的 XML。逐个公式格去问就成了 O(n²)。这里在第一次真正需要显示值时
    整表遍历一遍，之后都是字典查询。没有公式格的分表连这一遍都不会跑。
    """

    def __init__(self, worksheet) -> None:
        self._worksheet = worksheet
        self._values: dict[str, str] | None = None

    def get(self, coordinate: str) -> str | None:
        if self._values is None:
            self._values = {
                cell.coordinate: cell.value
                for row in self._worksheet.iter_rows()
                for cell in row
                if isinstance(cell.value, str)
            }
        return self._values.get(coordinate)


def _resolve_cell_text(
    cell,
    display_values: _DisplayValues | None,
    formula_display_value_backfill: bool,
) -> str | None:
    value = cell.value
    if not isinstance(value, str):
        return None
    if not formula_display_value_backfill or getattr(cell, "data_type", None) != "f":
        return value
    if display_values is None:
        return None
    return display_values.get(cell.coordinate)


def _is_generated_original_sheet(sheet_name: str, workbook_sheet_names: set[str]) -> bool:
    """这张分表是不是我们自己上一轮克隆出来的「_原文」分表（补译时要跳过）。

    判据交给生成方 ``xlsx_patcher``，它才知道分表名是怎么被 31 字符上限截断、
    怎么加去重后缀的。这里不做任何字符串反推——原名超 27 字符时「_原文」会被
    截没，反推必然落空，结果就是把自己生成的原文分表当成待译内容重翻一遍。
    """
    return xlsx_patcher.is_generated_original_sheet_title(
        sheet_name, workbook_sheet_names
    )
