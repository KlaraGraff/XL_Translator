"""
后台任务线程管理器（全局聚合流水线架构）。

三阶段全局处理模型：
  阶段 1（收集）：依次扫描所有表格并提取词条，全局去重汇聚。
  阶段 2（翻译）：对去重词汇池统一查询 TM + 批量并发 API 翻译。
  阶段 3（写入）：翻译数据就绪后，逐文件串行回填写入。

优势：打破文件壁垒，最大化利用 API 批次并发能力；跨文件去重减少重复请求。
"""
import os
import queue
import re
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from loguru import logger

from core import bilingual_writer
from core.api_concurrency_control import ApiKeyTemporarilyUnavailableError
from core.api_scheduler import API_REQUEST_CATEGORY_NORMAL, WeightedApiScheduler
from core.api_config_check import check_translation_api_config
from core.excel_coverage import build_excel_coverage_plan, write_untranslated_excel_file
from core.file_scanner import FileItem
from core.language_registry import (
    build_lang_pair,
    get_default_source_lang,
    get_tm_language_pairs,
    is_auto_source_lang,
)
from core.language_preflight import (
    LANGUAGE_PREFLIGHT_SYSTEM_PROMPT,
    build_language_preflight_prompt,
    preflight_files,
)
from core.mixed_language import (
    DEFAULT_MIXED_MAX_BATCH_CHARS,
    MIXED_ACTION_EXISTING_BILINGUAL,
    MIXED_ACTION_FOREIGN_NOISE,
    MIXED_ACTION_TRANSLATE,
    MIXED_ACTION_UNCERTAIN,
    MIXED_MARK_FOREIGN_NOISE,
    MIXED_MARK_SEMANTIC,
    MIXED_MARK_UNRESOLVED,
    MixedLanguageRunStats,
    split_mixed_language_sources,
    translate_mixed_language_texts,
)
from core.translation_filter import should_translate
from core.engine_dispatcher import (
    TranslationBatchRunStats,
    build_engine,
    build_role_engine,
    get_system_prompt,
    translate_texts,
    translate_texts_with_sources,
)
from core.model_roles import ROLE_TRANSLATION, resolve_effective_model_config
from core.model_throughput import get_model_throughput
from core.translation_protocol import should_store_translation_in_tm
from core import tm_manager
from core.excel_automation import (
    create_excel_app,
    finalize_excel_thread,
    get_excel_process_pid,
    initialize_excel_thread,
    terminate_process_tree,
)
from core.task_logger import TaskLogger
from core.user_facing_errors import humanize_error
from settings import AppSettings, provider_key_overrides

# 「给用户看的一句话」的判定标准，写在这里是因为三个 runner 都要用同一条标准。
# humanize_error 认识的原因会被换成中文；它认不出的照原样返回——那是库该有的
# 默认，却正是我们不能端到界面上的东西。所以这里再加一道闸：句子里必须有中文，
# 且不能带请求地址或 JSON 错误体，否则就用调用点给的兜底句，原文只留给 debug 日志。
_USER_REASON_CJK_RE = re.compile(r"[\u3400-\u4dbf\u4e00-\u9fff]")
_USER_REASON_URL_RE = re.compile(r"\b(?:https?|wss?)://", re.IGNORECASE)
_USER_REASON_JSON_RE = re.compile(r"\{\s*[\"']")


def user_facing_reason(value: object, *, fallback: str) -> str:
    """Return one plain-Chinese sentence describing ``value`` to a user."""
    sentence = humanize_error(value)
    if not _USER_REASON_CJK_RE.search(sentence):
        return fallback
    if _USER_REASON_URL_RE.search(sentence) or _USER_REASON_JSON_RE.search(sentence):
        return fallback
    return sentence


AUTOFIT_STALL_TIMEOUT_SECONDS = 180
AUTOFIT_MONITOR_POLL_SECONDS = 0.5
_EXCEL_REVIEW_MARK_PRIORITY = {
    MIXED_MARK_SEMANTIC: 10,
    MIXED_MARK_UNRESOLVED: 20,
    MIXED_MARK_FOREIGN_NOISE: 30,
}


# ── 消息类型 ──────────────────────────────────────────────────────────────────

@dataclass
class ProgressMsg:
    """全局进度消息。"""
    phase_index:  int      # 当前阶段序号（1/2/3/4）
    phase_total:  int      # 总阶段数（3）
    phase_name:   str      # 阶段名称（如"全局扫描"、"云端翻译"、"生成文件"）
    step_done:    int      # 当前阶段已完成步数
    step_total:   int      # 当前阶段总步数


@dataclass
class StatusMsg:
    phase_desc: str


@dataclass
class WordRecoveryStatusMsg:
    """Word recovery summary for the non-scrolling execution monitor."""
    retry_round: int = 0
    retry_total: int = 0
    retry_processing_count: int = 0
    retry_recovered_count: int = 0
    retry_unresolved_count: int = 0
    semantic_processing_count: int = 0
    semantic_checked_count: int = 0
    semantic_accepted_count: int = 0
    semantic_uncertain_count: int = 0


@dataclass
class PdfReviewStatusMsg:
    """PDF page-review summary for the non-scrolling execution monitor."""
    enabled: bool = False
    review_round: int = 0
    review_total: int = 0
    review_processing_count: int = 0
    review_passed_count: int = 0
    review_failed_count: int = 0


@dataclass
class PdfPageRecoveryStatusMsg:
    """PDF page retry/recovery summary for the non-scrolling execution monitor."""
    total_pages: int = 0
    completed_pages: int = 0
    submitted_page_count: int = 0
    pending_submitted_page_count: int = 0
    retrying_page_count: int = 0
    retried_page_count: int = 0
    recovered_page_count: int = 0
    placeholder_page_count: int = 0


@dataclass
class LogMsg:
    level:   str           # INFO / OK / WARN / ERROR
    message: str
    ts:      str = field(default_factory=lambda: datetime.now().strftime("%H:%M:%S"))
    visible: bool = True


@dataclass
class DoneMsg:
    output_dir:   str
    file_results: list[dict]
    elapsed_sec:  float
    tm_hit_count: int
    api_call_count: int
    issues: list[dict] = field(default_factory=list)
    report_path: str = ""
    report_warning: str = ""
    files: list[dict] = field(default_factory=list)
    kpi: dict[str, object] = field(default_factory=dict)
    recovery: dict[str, object] = field(default_factory=dict)
    review: dict[str, object] = field(default_factory=dict)
    # Runtime failover history: who took over mid-run, and why.
    connections: dict[str, object] = field(default_factory=dict)
    language: dict[str, object] = field(default_factory=dict)
    error: dict[str, object] = field(default_factory=dict)
    # 用户按了停止、但任务还是正常跑完时留下的痕迹（requested / truncated）。没有它，
    # 小结会写「已完成 · 全部通过」，用户看不出自己那一下停止到底起没起作用。
    stop: dict[str, object] = field(default_factory=dict)


@dataclass
class ErrorMsg:
    message: str
    output_dir: str = ""
    report_path: str = ""
    report_warning: str = ""
    manifest_path: str = ""
    files: list[dict] = field(default_factory=list)
    kpi: dict[str, object] = field(default_factory=dict)
    recovery: dict[str, object] = field(default_factory=dict)
    review: dict[str, object] = field(default_factory=dict)
    # Runtime failover history: who took over mid-run, and why.
    connections: dict[str, object] = field(default_factory=dict)
    language: dict[str, object] = field(default_factory=dict)
    error: dict[str, object] = field(default_factory=dict)


@dataclass
class StoppedMsg:
    message: str
    output_dir: str = ""
    report_path: str = ""
    report_warning: str = ""
    manifest_path: str = ""
    files: list[dict] = field(default_factory=list)
    kpi: dict[str, object] = field(default_factory=dict)
    recovery: dict[str, object] = field(default_factory=dict)
    review: dict[str, object] = field(default_factory=dict)
    # Runtime failover history: who took over mid-run, and why.
    connections: dict[str, object] = field(default_factory=dict)
    language: dict[str, object] = field(default_factory=dict)
    error: dict[str, object] = field(default_factory=dict)


class TaskStopped(Exception):
    """后台任务收到停止信号时抛出，用于统一收尾。"""


def _set_excel_review_mark(review_marks: dict[str, str], source: str, mark: str) -> None:
    source_key = str(source or "").strip()
    if not source_key:
        return
    existing = review_marks.get(source_key)
    if existing is None or _EXCEL_REVIEW_MARK_PRIORITY.get(mark, 0) > _EXCEL_REVIEW_MARK_PRIORITY.get(existing, 0):
        review_marks[source_key] = mark


# ── TaskRunner ────────────────────────────────────────────────────────────────

class TaskRunner:
    """
    封装后台翻译任务（全局聚合流水线）。
    UI 层通过轮询 .get_message() 获取进度/日志/完成消息。
    """

    def __init__(
        self,
        file_items: list[FileItem],
        settings: AppSettings,
        source_root: Path | str | None = None,
        allow_xls_fallback: bool = False,
        source_lang: str | None = None,
        key_overrides: dict[str, str] | None = None,
        api_scheduler=None,
        untranslated_only: bool = False,
        connection_chain: tuple[str, ...] | list[str] | None = None,
    ):
        self._files       = file_items
        self._settings    = settings
        self._source_root = Path(source_root) if source_root else None
        self._allow_xls_fallback = allow_xls_fallback
        self._source_lang = str(source_lang or settings.source_lang or "zh").strip() or "zh"
        self._key_overrides = dict(key_overrides or {})
        self._api_scheduler = api_scheduler
        self._untranslated_only = bool(untranslated_only)
        # The pool entries this task may fall back to, frozen at start.
        self._connection_chain = tuple(connection_chain or ())
        self._queue: queue.Queue = queue.Queue()
        self._thread: threading.Thread | None = None
        self._stop_event = threading.Event()
        self._task_logger = TaskLogger(enabled=True)
        # Runtime connection switches, in order. A switch changes who translated
        # the rest of the run, so it belongs in the terminal record and the
        # report — not only in a WARN line that scrolls away.
        self._connection_switches: list[dict[str, str]] = []
        self._switch_lock = threading.Lock()

    def start(self) -> None:
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._run_with_overrides, daemon=True)
        self._thread.start()

    @property
    def task_id(self) -> str:
        return self._task_logger.task_id

    def stop(self) -> None:
        self._stop_event.set()

    def stop_requested(self) -> bool:
        return self._stop_event.is_set()

    def is_running(self) -> bool:
        return self._thread is not None and self._thread.is_alive()

    def needs_poll(self) -> bool:
        """线程存活 OR 队列中仍有待读消息（防止线程退出后 DoneMsg 被遗漏）。"""
        return self.is_running() or not self._queue.empty()

    def get_message(self, timeout: float = 0.05):
        """非阻塞获取消息；无消息时返回 None。"""
        try:
            return self._queue.get(timeout=timeout)
        except queue.Empty:
            return None

    def _log(self, level: str, msg: str) -> None:
        self._queue.put(LogMsg(level=level, message=msg))
        logger.info(f"[{level}] {msg}")

    def _report_connection_switch(self, previous, current, failure_kind, error) -> None:
        """Surface a runtime connection switch in the task's own event log."""
        reason = {
            "endpoint": "服务端不可用",
            "credential": "密钥被拒绝或额度耗尽",
        }.get(failure_kind, failure_kind)
        # Batches run on a thread pool, so two threads can report a switch at once.
        with self._switch_lock:
            self._connection_switches.append(
                {
                    "from_label": str(getattr(previous, "display_label", "") or ""),
                    "to_label": str(getattr(current, "display_label", "") or ""),
                    "reason": reason,
                    "at": datetime.now().strftime("%H:%M:%S"),
                }
            )
        self._log(
            "WARN",
            f"已切换连接：{previous.display_label} → {current.display_label}（{reason}）",
        )

    # ── 全局聚合流水线主入口 ──────────────────────────────────────────────

    def _run_with_overrides(self) -> None:
        with provider_key_overrides(self._key_overrides):
            self._run()

    def _run(self) -> None:
        start_ts     = datetime.now()
        settings     = self._settings
        excel_output = settings.excel_output
        source_lang  = self._source_lang
        auto_source_lang = is_auto_source_lang(source_lang)
        target_lang  = settings.target_lang
        # ``auto`` is a selector state only; it is resolved after each file's
        # one-shot preflight and never becomes a TM pair.
        lang_pair    = (
            None
            if auto_source_lang
            else build_lang_pair(target_lang, source_lang=source_lang)
        )
        max_len      = settings.tm.max_len
        tm_hit_count = 0
        api_call_count = 0
        file_results: list[dict] = []
        stopped_message: str | None = None
        fatal_error_message: str | None = None
        quality_issues: list[dict] = []
        file_language_preflights: dict = {}
        file_texts: list[set[str]] = []
        normal_api_language_results: dict = {}
        excel_review_marks: dict[str, str] = {}

        try:
            config_check = check_translation_api_config(settings)
            if not config_check.ok:
                detail = f"（{config_check.detail}）" if config_check.detail else ""
                self._queue.put(ErrorMsg(message=f"{config_check.message}{detail}"))
                return
            # Only take the failover path when there is somewhere to fail over
            # to; a single-connection setup keeps exactly its previous engine.
            engine        = (
                build_role_engine(
                    settings,
                    ROLE_TRANSLATION,
                    connection_ids=self._connection_chain,
                    on_switch=self._report_connection_switch,
                )
                if len(self._connection_chain) > 1
                else build_engine(settings)
            )
            system_prompt = get_system_prompt(
                settings,
                target_lang=target_lang,
                source_lang=source_lang if not auto_source_lang else get_default_source_lang(),
                page_key="excel",
            )
            model_config = resolve_effective_model_config(settings, ROLE_TRANSLATION)
            throughput = get_model_throughput(settings, model_config)
            batch_size = throughput.batch_size or 1
            concurrency = throughput.concurrency
        except Exception as e:
            logger.debug(f"引擎初始化失败原始错误：{e!r}")
            self._queue.put(
                ErrorMsg(
                    message="引擎初始化失败："
                    + user_facing_reason(
                        e,
                        fallback="翻译设置里还有没填好的项，请检查所选连接和模型。",
                    )
                )
            )
            return

        def _raise_if_stopped(message: str = "任务已停止") -> None:
            if self._stop_event.is_set():
                raise TaskStopped(message)

        root_for_output = self._source_root if self._source_root else self._files[0].path.parent
        custom_output_dir = (
            excel_output.custom_output_dir
            if excel_output.use_custom_output_dir
            else None
        )
        try:
            if excel_output.use_custom_output_dir:
                output_dir_error = bilingual_writer.get_custom_output_dir_error(custom_output_dir)
                if output_dir_error is not None:
                    raise ValueError(output_dir_error)

            output_dir = bilingual_writer.build_output_dir(root_for_output, custom_output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            logger.debug(f"输出目录初始化失败原始错误：{e!r}")
            self._queue.put(
                ErrorMsg(
                    message="输出目录初始化失败："
                    + user_facing_reason(
                        e,
                        fallback="输出目录不可用，请在设置里另选一个可写的目录。",
                    )
                )
            )
            return

        # ── 任务日志：记录启动信息 ────────────────────────────────────
        self._task_logger.task_start(
            files                = self._files,
            engine_name          = engine.engine_name,
            target_lang          = target_lang,
            keep_original_sheets = excel_output.keep_original_sheets,
            formula_display_value_backfill = excel_output.formula_display_value_backfill,
            enable_excel_autofit = excel_output.enable_excel_autofit,
            lock_row_height      = excel_output.lock_row_height,
        )

        self._log("INFO", f"[诊断] source_root={self._source_root} | custom_output_dir={custom_output_dir} | output_dir={output_dir}")
        self._log("INFO", f"扫描到 {len(self._files)} 个文件")
        if self._untranslated_only:
            self._log("INFO", "[补译模式] 只补未译内容：已有译文位置将保持不变。")

        need_autofit = excel_output.enable_excel_autofit and not excel_output.lock_row_height
        phase_total = 4 if need_autofit else 3

        excel_app = None
        excel_thread_state = None
        excel_policy = "split"
        excel_policy_reason = "not_evaluated"
        reuse_excel_for_autofit = False
        xls_file_count = 0
        total_sheet_count = 0
        raw_text_count = 0
        def _configure_excel_app(app) -> None:
            """Tune Excel for unattended automation and large batch runs."""
            try:
                app.display_alerts = False
            except Exception:
                pass
            for attr, value in (
                ("screen_updating", False),
                ("enable_events", False),
                ("calculation", "manual"),
            ):
                try:
                    setattr(app, attr, value)
                except Exception:
                    continue

        def _kill_excel_pid(pid: int | None, reason: str) -> None:
            if not pid:
                return
            try:
                terminated = terminate_process_tree(pid, force=True)
                if terminated:
                    self._log("WARN", f"{reason}，已强制结束 Excel 进程 PID={pid}")
                else:
                    self._log("WARN", f"{reason}，但未能确认 Excel 进程已退出 PID={pid}")
            except Exception as e:
                logger.debug(f"强制结束 Excel 进程失败原始错误 PID={pid}：{e!r}")
                self._log(
                    "WARN",
                    f"{reason}，但强制结束 Excel 进程失败 PID={pid}："
                    + user_facing_reason(e, fallback="请手动退出 Excel 后重试。"),
                )

        def _cleanup_excel_app(status_msg: str | None = None, *, force: bool = False) -> None:
            nonlocal excel_app, excel_thread_state
            if excel_app is None:
                return

            pid = get_excel_process_pid(excel_app)
            if status_msg:
                self._queue.put(StatusMsg(phase_desc=status_msg))

            try:
                if force:
                    excel_app.kill()
                else:
                    excel_app.quit()
            except Exception as e:
                logger.debug(f"Excel 进程清理异常原始错误：{e!r}")
                self._log(
                    "WARN",
                    "Excel 进程清理异常："
                    + user_facing_reason(e, fallback="程序会尝试强制结束它。"),
                )
                _kill_excel_pid(pid, "常规退出 Excel 失败")
            finally:
                excel_app = None
                finalize_excel_thread(excel_thread_state)
                excel_thread_state = None

        def _run_autofit_with_guard(file_paths: list[Path], progress_callback) -> bool:
            """Run AutoFit in a dedicated Excel process so it can be stopped safely."""
            worker_state = {
                "pid": None,
                "last_progress_ts": time.monotonic(),
                "error": None,
                "result": False,
            }
            done_event = threading.Event()

            def _worker():
                thread_state = None
                app = None
                try:
                    thread_state = initialize_excel_thread()
                    app = create_excel_app(visible=False, add_book=False)
                    _configure_excel_app(app)
                    worker_state["pid"] = get_excel_process_pid(app)

                    def _progress(done, total, current_file):
                        worker_state["last_progress_ts"] = time.monotonic()
                        progress_callback(done, total, current_file)

                    worker_state["result"] = bilingual_writer.autofit_files_batch(
                        file_paths,
                        app=app,
                        log_callback=lambda msg: self._log(
                            "WARN" if msg.startswith("[WARN]") else "INFO", msg
                        ),
                        progress_callback=_progress,
                    )
                except Exception as e:
                    worker_state["error"] = e
                finally:
                    if app is not None:
                        pid = get_excel_process_pid(app)
                        try:
                            app.quit()
                        except Exception:
                            _kill_excel_pid(pid, "AutoFit 线程退出时清理 Excel 失败")
                    finalize_excel_thread(thread_state)
                    done_event.set()

            worker = threading.Thread(target=_worker, daemon=True)
            worker.start()

            while not done_event.wait(timeout=AUTOFIT_MONITOR_POLL_SECONDS):
                if self._stop_event.is_set():
                    self._task_logger.warning("收到停止请求，正在终止 Excel 精调。")
                    _kill_excel_pid(worker_state["pid"], "收到停止请求")
                    done_event.wait(timeout=5)
                    raise TaskStopped("任务已停止，Excel 精调已终止，已保留当前已生成结果。")

                stalled_for = time.monotonic() - worker_state["last_progress_ts"]
                if stalled_for >= AUTOFIT_STALL_TIMEOUT_SECONDS:
                    msg = (
                        f"Excel AutoFit 连续 {AUTOFIT_STALL_TIMEOUT_SECONDS}s 无进度，"
                        "已终止 Excel 精调并保留 Python 估算行高。"
                    )
                    self._log("WARN", msg)
                    self._task_logger.warning(msg)
                    _kill_excel_pid(worker_state["pid"], "Excel 精调长时间无进度")
                    done_event.wait(timeout=5)
                    return False

            if worker_state["error"] is not None:
                raise worker_state["error"]

            return bool(worker_state["result"])

        def _get_excel_app():
            nonlocal excel_app, excel_thread_state
            if excel_app is None:
                self._queue.put(StatusMsg(phase_desc="状态：正在唤醒底层 Excel 引擎，请稍候..."))
                try:
                    excel_thread_state = initialize_excel_thread()
                    self._log("INFO", "开始启动全局 Excel 进程...")
                    excel_app = create_excel_app(visible=False, add_book=False)
                    _configure_excel_app(excel_app)
                except Exception as e:
                    finalize_excel_thread(excel_thread_state)
                    excel_thread_state = None
                    logger.debug(f"启动全局 Excel 进程失败原始错误：{e!r}")
                    start_reason = user_facing_reason(
                        e,
                        fallback="本机 Excel 没能启动，可能没安装，或没有授权本程序控制它。",
                    )
                    self._log("WARN", f"启动全局 Excel 进程失败：{start_reason}")
                    raise Exception(
                        "无法启动本地 Excel，请确认已正确安装并允许自动化控制："
                        f"{start_reason}"
                    )
            return excel_app

        try:
            _raise_if_stopped()
            # ══════════════════════════════════════════════════════════
            # 阶段 1：全局词汇提取（扫描 + .xls 转换 + 收集词条）
            # ══════════════════════════════════════════════════════════
            self._queue.put(StatusMsg(phase_desc=f"状态：[阶段 1/{phase_total}] 正在扫描所有文件提取词汇..."))

            # process_paths[i] 对应 self._files[i]，若 .xls 则指向转换后的临时 .xlsx
            process_paths: list[Path] = []
            # file_texts[i] 对应 self._files[i] 的本文件词条集合
            file_texts = []
            coverage_plans = []
            global_unique_texts: set[str] = set()
            file_language_preflights = {}
            tm_language_pairs: list[str] = []
            text_source_scopes: dict[str, list[frozenset[str]]] = {}
            file_conversion_modes: dict[str, str] = {
                str(item.path): "native_xlsx"
                for item in self._files
            }

            t_phase1 = datetime.now()

            for fi, file_item in enumerate(self._files):
                _raise_if_stopped()

                self._queue.put(ProgressMsg(
                    phase_index=1, phase_total=phase_total, phase_name="全局扫描",
                    step_done=fi, step_total=len(self._files),
                ))

                process_path = file_item.path
                source_is_xls = file_item.path.suffix.lower() == ".xls"

                if source_is_xls:
                    xls_file_count += 1

                # .xls 格式转换（在阶段 1 顺便完成）
                if source_is_xls:
                    self._queue.put(StatusMsg(phase_desc=f"状态：[阶段 1/{phase_total}] 正在转换 .xls 文件：{file_item.name}"))
                    from core.xls_converter import (
                        convert_with_excel,
                        convert_with_fallback,
                    )
                    t_conv = datetime.now()
                    try:
                        if not self._allow_xls_fallback:
                            app = _get_excel_app()
                            process_path = convert_with_excel(app, process_path)
                            file_conversion_modes[str(file_item.path)] = "excel_automation"
                        else:
                            process_path = convert_with_fallback(process_path)
                            file_conversion_modes[str(file_item.path)] = "compatibility_fallback"
                        self._log("INFO", f"格式转换完成 {file_item.name}，耗时 {(datetime.now() - t_conv).total_seconds():.2f}s")
                    except Exception as e:
                        logger.debug(f"源文件转换失败 {file_item.name} 原始错误：{e!r}")
                        conversion_reason = user_facing_reason(
                            e,
                            fallback="这个 .xls 文件没能转换成 .xlsx，请改用兼容转换后重试。",
                        )
                        self._log("ERROR", f"源文件转换失败 {file_item.name}: {conversion_reason}")
                        self._task_logger.file_error(file_item.name, conversion_reason)
                        file_results.append({
                            "name": file_item.name,
                            "source_path": str(file_item.path),
                            "source_relative_path": self._relative_source_path(file_item.path),
                            "format": file_item.format,
                            "conversion_mode": file_conversion_modes.get(
                                str(file_item.path), "high_fidelity_failed"
                            ),
                            "status": "failed",
                            "success": False,
                            "error": f"源文件转换失败: {conversion_reason}",
                        })
                        process_paths.append(process_path)
                        file_texts.append(set())
                        coverage_plans.append(None)
                        continue

                process_paths.append(process_path)

                # 收集词条
                self._log("INFO", f"[阶段 1] 提取词汇：{file_item.name}（{fi+1}/{len(self._files)}）")
                t0 = datetime.now()
                try:
                    if self._untranslated_only and not auto_source_lang:
                        coverage_plan = build_excel_coverage_plan(
                            process_path,
                            target_lang=target_lang,
                            source_lang=source_lang,
                            formula_display_value_backfill=(
                                excel_output.formula_display_value_backfill
                            ),
                        )
                        coverage_plans.append(coverage_plan)
                        texts = coverage_plan.source_texts
                        sheet_count = coverage_plan.sheet_count
                        self._log_excel_coverage_plan(file_item.name, coverage_plan)
                    else:
                        texts, sheet_count = self._collect_texts(
                            process_path,
                            file_item.name,
                            target_lang=target_lang,
                            source_lang=source_lang,
                        )
                        coverage_plans.append(None)
                except Exception as e:
                    logger.debug(f"源文件读取失败原始错误 {file_item.name}：{e!r}")
                    read_reason = user_facing_reason(
                        e,
                        fallback="这个文件打不开，可能已损坏或不是真正的 Excel 文件。",
                    )
                    self._log("ERROR", f"源文件读取失败 {file_item.name}：{read_reason}")
                    self._task_logger.file_error(file_item.name, f"源文件读取失败: {read_reason}")
                    file_results.append({
                        "name": file_item.name,
                        "source_path": str(file_item.path),
                        "source_relative_path": self._relative_source_path(file_item.path),
                        "format": file_item.format,
                        "conversion_mode": file_conversion_modes.get(str(file_item.path), "native_xlsx"),
                        "status": "failed",
                        "success": False,
                        "error": f"源文件读取失败: {read_reason}",
                    })
                    if len(coverage_plans) < len(process_paths):
                        coverage_plans.append(None)
                    file_texts.append(set())
                    if process_path != file_item.path:
                        try:
                            os.remove(process_path)
                        except Exception as cleanup_error:
                            self._log(
                                "WARN",
                                f"临时文件清理失败 {process_path.name}: "
                                f"{user_facing_reason(cleanup_error, fallback='临时文件删不掉。')}",
                            )
                    continue
                text_set = set(texts)
                file_texts.append(text_set)
                total_sheet_count += sheet_count
                raw_text_count += len(text_set)
                collect_elapsed = (datetime.now() - t0).total_seconds()

                if self._untranslated_only and auto_source_lang:
                    # 这一遍取的是全量候选，只拿来给语言预检当样本。真正要补译的
                    # 清单要等识别出源语言之后才算得出来（见下面的补译识别重建）。
                    self._log(
                        "INFO",
                        f"  → {file_item.name}：{len(text_set)} 处候选文本，"
                        f"先用于识别源语言（{collect_elapsed:.3f}s）",
                    )
                else:
                    self._log("INFO", f"  → {file_item.name}：{len(text_set)} 处待翻译文本（{collect_elapsed:.3f}s）")
                self._task_logger.file_collected(file_item.name, len(text_set), collect_elapsed)

                global_unique_texts.update(text_set)

            # 阶段 1 收尾：发送最终进度
            self._queue.put(ProgressMsg(
                phase_index=1, phase_total=phase_total, phase_name="全局扫描",
                step_done=len(self._files), step_total=len(self._files),
            ))

            phase1_elapsed = (datetime.now() - t_phase1).total_seconds()
            self._log("OK", f"[阶段 1 完成] 全部文件合计 {len(global_unique_texts)} 处待翻译文本，相同内容只翻一次（{phase1_elapsed:.2f}s）")
            self._task_logger.global_collected(
                total_unique=len(global_unique_texts),
                file_count=len(self._files),
                elapsed=phase1_elapsed,
            )

            if auto_source_lang:
                self._queue.put(
                    StatusMsg(
                        phase_desc=(
                            "状态：自动识别模式，正在对每个有候选文本的文件执行一次语言预检..."
                        )
                    )
                )

                detector_calls = {"count": 0}

                def _detect_file_language(samples, detected_target):
                    detector_calls["count"] += 1
                    raw = engine.chat(
                        LANGUAGE_PREFLIGHT_SYSTEM_PROMPT,
                        build_language_preflight_prompt(
                            samples,
                            target_lang=detected_target,
                        ),
                    )
                    return raw

                file_payloads = {
                    str(file_item.path): texts
                    for file_item, texts in zip(self._files, file_texts)
                }
                file_language_preflights = preflight_files(
                    file_payloads,
                    _detect_file_language,
                    target_lang=target_lang,
                )
                detected_sources: list[str] = []
                for result in file_language_preflights.values():
                    for detected in result.source_langs:
                        if detected not in detected_sources:
                            detected_sources.append(detected)
                if detected_sources:
                    source_lang = detected_sources[0]
                    system_prompt = get_system_prompt(
                        settings,
                        target_lang=target_lang,
                        source_lang=source_lang,
                        page_key="excel",
                    )
                else:
                    source_lang = get_default_source_lang()
                tm_language_pairs = get_tm_language_pairs(detected_sources, target_lang)
                for file_item, text_set in zip(self._files, file_texts):
                    result = file_language_preflights.get(str(file_item.path))
                    if result is None:
                        continue
                    for text in text_set:
                        text_source_scopes.setdefault(text, []).append(
                            frozenset(result.source_langs)
                        )
                self._log(
                    "INFO",
                    (
                        f"自动语言预检完成：{detector_calls['count']} 次请求，"
                        f"实际源语言={','.join(detected_sources) or '未确定'}，"
                        f"TM 语言对={','.join(tm_language_pairs) or '无'}"
                    ),
                )
                if self._untranslated_only:
                    raw_text_count = self._rebuild_coverage_plans_after_preflight(
                        process_paths=process_paths,
                        coverage_plans=coverage_plans,
                        file_texts=file_texts,
                        file_results=file_results,
                        file_conversion_modes=file_conversion_modes,
                        global_unique_texts=global_unique_texts,
                        target_lang=target_lang,
                        source_lang=source_lang,
                        formula_display_value_backfill=(
                            excel_output.formula_display_value_backfill
                        ),
                    )
            else:
                tm_language_pairs = [lang_pair] if lang_pair else []

            excel_policy, excel_policy_reason = self._decide_excel_policy(
                need_autofit=need_autofit,
                xls_file_count=xls_file_count,
                total_sheet_count=total_sheet_count,
                raw_text_count=raw_text_count,
            )
            if excel_policy == "reuse" and excel_app is None:
                excel_policy = "split"
                excel_policy_reason = "no_reusable_excel_process"

            reuse_excel_for_autofit = (
                need_autofit
                and xls_file_count > 0
                and excel_policy == "reuse"
                and excel_app is not None
            )
            excel_policy_log = (
                f"excel_policy={excel_policy} | "
                f"xls_file_count={xls_file_count} | "
                f"total_sheet_count={total_sheet_count} | "
                f"raw_text_count={raw_text_count} | "
                f"reason={excel_policy_reason}"
            )
            self._log("INFO", f"[Excel策略] {excel_policy_log}")
            self._task_logger.excel_policy_decided(
                excel_policy=excel_policy,
                xls_file_count=xls_file_count,
                total_sheet_count=total_sheet_count,
                raw_text_count=raw_text_count,
                reason=excel_policy_reason,
            )

            if excel_app is not None and not reuse_excel_for_autofit:
                self._log("INFO", "按策略释放阶段 1 的 Excel 进程，阶段 4 将使用干净进程")
                _cleanup_excel_app(force=True)
            elif reuse_excel_for_autofit:
                self._log("INFO", "按策略保留阶段 1 的 Excel 进程，阶段 4 将直接复用")

            _raise_if_stopped()

            # ══════════════════════════════════════════════════════════
            # 阶段 2：全局统一翻译（TM 查询 + API 并发）
            # ══════════════════════════════════════════════════════════
            self._queue.put(StatusMsg(phase_desc=f"状态：[阶段 2/{phase_total}] 正在比对翻译记忆库..."))

            t_phase2 = datetime.now()

            # 混合语言先分流，避免旧 TM 命中污染已双语/夹杂外文内容。
            all_texts_list = list(global_unique_texts)
            normal_texts, mixed_texts = split_mixed_language_sources(
                all_texts_list,
                target_lang=target_lang,
                source_lang=source_lang,
            )
            if mixed_texts:
                self._log(
                    "INFO",
                    f"混合语言路径命中 {len(mixed_texts)} 处内容，已从记忆库查询中分流。",
                )

            # TM 查询按“文件预检 -> 该文件文本 -> 实际语言对”分组。不能
            # 将全局词池交叉查到每一个语言对，否则 zh 文件会误命中 en-*。
            text_tm_pairs: dict[str, set[str]] = {}
            if auto_source_lang:
                for file_item, text_set in zip(self._files, file_texts):
                    result = file_language_preflights.get(str(file_item.path))
                    allowed_pairs = (
                        set(result.tm_lang_pairs(target_lang)) if result is not None else set()
                    )
                    for text in text_set:
                        if text in normal_texts and allowed_pairs:
                            text_tm_pairs.setdefault(text, set()).update(allowed_pairs)
            elif lang_pair:
                text_tm_pairs = {text: {lang_pair} for text in normal_texts}

            # Same-value hits across two permitted pairs are usable, but a
            # conflicting value remains a model miss and is never guessed.
            tm_values_by_text: dict[str, list[str]] = {
                text: [] for text in normal_texts
            }
            tm_texts_by_pair: dict[str, list[str]] = {}
            for text, pairs in text_tm_pairs.items():
                for pair in pairs:
                    tm_texts_by_pair.setdefault(pair, []).append(text)
            for pair, pair_texts in tm_texts_by_pair.items():
                tm_result = tm_manager.lookup_batch(pair_texts, pair)
                for text, value in tm_result.items():
                    if value is not None and str(value) not in tm_values_by_text.setdefault(text, []):
                        tm_values_by_text[text].append(str(value))
            hits = {
                text: values[0]
                for text, values in tm_values_by_text.items()
                if len(values) == 1
            }
            misses = [text for text in normal_texts if text not in hits]

            tm_hit_count   = len(hits)
            api_call_count = len(misses) + len(mixed_texts)

            self._log(
                "INFO",
                f"[阶段 2] TM 命中：{tm_hit_count}  普通待 API：{len(misses)}  混合语言：{len(mixed_texts)}",
            )
            self._task_logger.global_tm_result(hits=tm_hit_count, misses=api_call_count)
            self._queue.put(ProgressMsg(
                phase_index=2,
                phase_total=phase_total,
                phase_name="云端翻译",
                step_done=0 if api_call_count else 1,
                step_total=max(api_call_count, 1),
            ))

            # API 翻译未命中词条
            api_translations: dict[str, str] = {}
            normal_api_translations: dict[str, str] = {}
            normal_api_language_results = {}
            excel_review_marks = {}
            if (misses or mixed_texts) and not self._stop_event.is_set():
                self._queue.put(StatusMsg(phase_desc=f"状态：[阶段 2/{phase_total}] 正在请求大模型翻译未命中词汇..."))
                self._log("INFO", f"发送 API 请求，共 {api_call_count} 词条")

                progress_lock = threading.Lock()
                progress_done = {"normal": 0, "mixed": 0}

                def _emit_api_progress(kind: str, done: int) -> None:
                    with progress_lock:
                        progress_done[kind] = max(0, int(done or 0))
                        total_done = min(progress_done["normal"], len(misses)) + min(
                            progress_done["mixed"],
                            len(mixed_texts),
                        )
                    self._queue.put(ProgressMsg(
                        phase_index=2,
                        phase_total=phase_total,
                        phase_name="云端翻译",
                        step_done=min(total_done, max(api_call_count, 1)),
                        step_total=max(api_call_count, 1),
                    ))

                def progress_cb(done, total):
                    _emit_api_progress("normal", done)

                def api_error_cb(msg: str) -> None:
                    level = "ERROR" if "未能翻译" in msg else "WARN"
                    self._log(level, msg)

                t0 = datetime.now()
                batch_stats = TranslationBatchRunStats()
                shared_scheduler = None
                if settings.engine.mode != "local":
                    shared_scheduler = self._api_scheduler or WeightedApiScheduler(concurrency)

                mixed_stats = MixedLanguageRunStats()
                mixed_results = {}

                def run_normal_api_translation() -> dict[str, str]:
                    if not misses:
                        return {}
                    if auto_source_lang:
                        language_results = translate_texts_with_sources(
                            misses,
                            engine,
                            target_lang,
                            system_prompt,
                            batch_size,
                            concurrency,
                            progress_cb,
                            api_error_cb,
                            should_stop=self.stop_requested,
                            api_scheduler=shared_scheduler,
                            stats=batch_stats,
                        )
                        normal_api_language_results.update(language_results)
                        return {
                            source: result.translation
                            for source, result in language_results.items()
                        }
                    return translate_texts(
                        misses,
                        engine,
                        target_lang,
                        system_prompt,
                        batch_size,
                        concurrency,
                        progress_cb,
                        api_error_cb,
                        should_stop=self.stop_requested,
                        source_lang=source_lang,
                        api_scheduler=shared_scheduler,
                        stats=batch_stats,
                    )

                def run_mixed_api_translation():
                    if not mixed_texts:
                        return {}
                    self._queue.put(StatusMsg(phase_desc=f"状态：[阶段 2/{phase_total}] 正在处理混合语言内容..."))

                    def mixed_progress_cb(done, total):
                        _emit_api_progress("mixed", done)

                    return translate_mixed_language_texts(
                        mixed_texts,
                        engine=engine,
                        target_lang=target_lang,
                        system_prompt=system_prompt,
                        source_lang=source_lang,
                        concurrency=concurrency,
                        max_items_per_batch=batch_size,
                        max_chars_per_batch=DEFAULT_MIXED_MAX_BATCH_CHARS,
                        retry_attempts=3,
                        progress_callback=mixed_progress_cb,
                        error_callback=api_error_cb,
                        should_stop=self.stop_requested,
                        api_scheduler=shared_scheduler,
                        request_category=API_REQUEST_CATEGORY_NORMAL,
                        stats=mixed_stats,
                    )

                if misses and mixed_texts and shared_scheduler is not None:
                    with ThreadPoolExecutor(max_workers=2) as main_executor:
                        normal_future = main_executor.submit(run_normal_api_translation)
                        mixed_future = main_executor.submit(run_mixed_api_translation)
                        normal_api_translations = normal_future.result()
                        mixed_results = mixed_future.result()
                else:
                    normal_api_translations = run_normal_api_translation()
                    mixed_results = run_mixed_api_translation()

                api_translations.update(normal_api_translations)
                for source, translation in normal_api_translations.items():
                    if str(source or "").strip().lower() == str(translation or "").strip().lower():
                        _set_excel_review_mark(
                            excel_review_marks,
                            source,
                            MIXED_MARK_UNRESOLVED,
                        )

                if mixed_texts:
                    for source, result in mixed_results.items():
                        if result.action in {MIXED_ACTION_TRANSLATE, MIXED_ACTION_FOREIGN_NOISE} and result.translation.strip():
                            api_translations[source] = result.translation.strip()
                        elif result.action == MIXED_ACTION_UNCERTAIN:
                            api_translations[source] = source
                        if result.mark_kind:
                            _set_excel_review_mark(
                                excel_review_marks,
                                source,
                                result.mark_kind,
                            )
                    self._log(
                        "INFO",
                        (
                            "混合语言路径："
                            f"命中 {mixed_stats.input_count} 条，"
                            f"已双语 {mixed_stats.action_counts.get(MIXED_ACTION_EXISTING_BILINGUAL, 0)}，"
                            f"疑似原文错误 {mixed_stats.action_counts.get(MIXED_ACTION_FOREIGN_NOISE, 0)}，"
                            f"不确定 {mixed_stats.action_counts.get(MIXED_ACTION_UNCERTAIN, 0)}，"
                            f"语义校验接受 {mixed_stats.semantic_accepted_count}"
                        ),
                    )
                _raise_if_stopped("任务已停止，未写入剩余翻译结果。")
                api_elapsed = (datetime.now() - t0).total_seconds()
                self._log(
                    "INFO",
                    (
                        "Excel 实际请求："
                        f"{batch_stats.batch_count} 批，"
                        f"缩小重试 {batch_stats.retry_count} 次，"
                        f"未翻译 {batch_stats.untranslated_count} 条，"
                        f"最大请求权重 {batch_stats.max_request_weight}"
                        + (
                            f"，自适应降并发 {batch_stats.adaptive_concurrency_reductions} 次，"
                            f"最低并发 {batch_stats.adaptive_lowest_concurrency}"
                            if batch_stats.adaptive_concurrency_reductions
                            else ""
                        )
                    ),
                )
                untranslated_count = (
                    batch_stats.untranslated_count or batch_stats.failed_batch_count
                )
                if untranslated_count:
                    # The file still gets written with the source text in place,
                    # so the count has to travel with the result: an unflagged
                    # run must never hand back source text dressed as translation.
                    quality_issues.append(
                        {
                            "type": "api_unavailable",
                            "severity": "needs_action",
                            "count": untranslated_count,
                            "message": (
                                f"有 {untranslated_count} 条内容未能从 API 获得译文，"
                                "文件中这些位置保留的是原文，不是译文。"
                                "请检查 API Key、Base URL、模型名称或服务状态，重新配置后再试。"
                            ),
                            "failed_sources": list(batch_stats.failed_items),
                        }
                    )
                    self._log(
                        "ERROR",
                        f"本次有 {untranslated_count} 条内容未获得译文，已在结果中标记为未翻译。",
                    )
                self._log("OK", f"API 翻译完成，返回 {len(api_translations)} 条（{api_elapsed:.2f}s）")
                self._task_logger.global_api_done(returned=len(api_translations), elapsed=api_elapsed)

                # 将 API 结果写入 TM
                written, tm_write_error = self.store_api_results_in_tm(
                    auto_source_lang=auto_source_lang,
                    normal_api_language_results=normal_api_language_results,
                    normal_api_translations=normal_api_translations,
                    text_source_scopes=text_source_scopes,
                    target_lang=target_lang,
                    lang_pair=lang_pair,
                    max_len=max_len,
                    engine_name=engine.engine_name,
                )
                if tm_write_error:
                    self._log("WARN", tm_write_error)
                    self._task_logger.warning(tm_write_error)
                    quality_issues.append(
                        {
                            "type": "tm_write_failed",
                            "severity": "info",
                            "message": tm_write_error,
                        }
                    )
                if written:
                    self._log("INFO", f"新增 TM 词条：{written} 条")

            # 汇聚全局翻译词典：TM 命中覆盖 API 结果（TM 优先）
            global_translations = {**api_translations, **hits}
            for source, translation in global_translations.items():
                if str(source or "").strip().lower() == str(translation or "").strip().lower():
                    _set_excel_review_mark(
                        excel_review_marks,
                        source,
                        MIXED_MARK_UNRESOLVED,
                    )

            phase2_elapsed = (datetime.now() - t_phase2).total_seconds()
            self._log("OK", f"[阶段 2 完成] 翻译数据就绪（{phase2_elapsed:.2f}s）")

            _raise_if_stopped()

            # ══════════════════════════════════════════════════════════
            # 阶段 3：逐文件串行回填写入
            # ══════════════════════════════════════════════════════════
            self._queue.put(StatusMsg(phase_desc=f"状态：[阶段 3/{phase_total}] 正在生成双语表格..."))
            self._queue.put(ProgressMsg(
                phase_index=3,
                phase_total=phase_total,
                phase_name="生成文件",
                step_done=0,
                step_total=max(len(self._files), 1),
            ))

            t_phase3 = datetime.now()
            source_root = self._source_root if self._source_root else self._files[0].path.parent

            for fi, file_item in enumerate(self._files):
                _raise_if_stopped()

                # 跳过阶段 1 已失败的文件
                already_failed = any(
                    r.get("source_path") == str(file_item.path) and not r.get("success")
                    for r in file_results
                )
                if already_failed:
                    continue

                self._queue.put(ProgressMsg(
                    phase_index=3, phase_total=phase_total, phase_name="生成文件",
                    step_done=fi, step_total=len(self._files),
                ))

                self._log("INFO", f"[阶段 3] 写入文件：{file_item.name}（{fi+1}/{len(self._files)}）")

                process_path = process_paths[fi]

                try:
                    rel_subdir = file_item.path.parent.relative_to(source_root)
                except ValueError:
                    rel_subdir = Path()

                try:
                    t0 = datetime.now()
                    file_review_positions: list[dict[str, str]] = []
                    # 写入器要知道后面还会不会跑 Excel 整表 AutoFit：会跑就得把整张表的
                    # 悬浮图片锚点全部固定，否则 Excel 重排行高会把没冻结的图片拉变形。
                    write_stats: dict[str, object] = {}
                    # KNOWN-ISSUE-VAL-006:
                    # The current write path intentionally stays text-only.
                    # See docs/KNOWN_ISSUES.md before reintroducing image flow.
                    if self._untranslated_only:
                        coverage_plan = coverage_plans[fi] if fi < len(coverage_plans) else None
                        if coverage_plan is None:
                            raise ValueError("缺少补译识别计划，无法安全按位置写入。")
                        out_path = write_untranslated_excel_file(
                            source_path=process_path,
                            output_dir=output_dir / rel_subdir,
                            plan=coverage_plan,
                            translations=global_translations,
                            target_lang=target_lang,
                            source_lang=source_lang,
                            keep_original_sheets=excel_output.keep_original_sheets,
                            formula_display_value_backfill=(
                                excel_output.formula_display_value_backfill
                            ),
                            lock_row_height=excel_output.lock_row_height,
                            log_callback=lambda msg: self._log(
                                "OK" if msg.startswith("[OK]") else "INFO", msg
                            ),
                            original_path=file_item.original_path,
                            external_autofit_planned=need_autofit,
                            stats=write_stats,
                        )
                    else:
                        out_path = bilingual_writer.write_bilingual_file(
                            source_path          = process_path,
                            output_dir           = output_dir / rel_subdir,
                            translations         = global_translations,
                            target_lang          = target_lang,
                            source_lang          = source_lang,
                            keep_original_sheets = excel_output.keep_original_sheets,
                            formula_display_value_backfill = excel_output.formula_display_value_backfill,
                            # E4B-10 deliberately removes the legacy print-guard
                            # toggle from Excel behavior.
                            enable_print_guard   = False,
                            lock_row_height      = excel_output.lock_row_height,
                            review_marks         = excel_review_marks,
                            review_mark_colors   = self._settings.excel_review.mark_colors,
                            mark_review_items    = self._settings.excel_review.mark_review_items,
                            existing_fill_policy = self._settings.excel_review.existing_fill_policy,
                            review_positions     = file_review_positions,
                            log_callback         = lambda msg: self._log(
                                "OK" if msg.startswith("[OK]") else "INFO", msg
                            ),
                            original_path        = file_item.original_path,
                            external_autofit_planned = need_autofit,
                            stats                = write_stats,
                        )
                    write_elapsed = (datetime.now() - t0).total_seconds()
                    self._task_logger.file_write_done(file_item.name, write_elapsed)

                    # 统计该文件对应的 TM/API 使用情况
                    this_file_texts = file_texts[fi]
                    this_tm = sum(1 for t in this_file_texts if t in hits)
                    this_api = sum(1 for t in this_file_texts if t in misses or t in mixed_texts)
                    self._task_logger.file_done(
                        filename=file_item.name,
                        elapsed=write_elapsed,
                        tm_hits=this_tm,
                        api_calls=this_api,
                    )

                    file_results.append({
                        "name":    file_item.name,
                        "source_path": str(file_item.path),
                        "source_relative_path": self._relative_source_path(file_item.path),
                        "format": file_item.format,
                        "conversion_mode": file_conversion_modes.get(str(file_item.path), "native_xlsx"),
                        "output":  str(out_path),
                        "output_path": str(out_path),
                        "status": "succeeded",
                        "review_count": len(file_review_positions),
                        "review_items": file_review_positions,
                        # 悬浮图片/形状里被我们固定住锚点的数量（0 表示没动过）
                        "anchor_frozen_count": int(
                            write_stats.get("anchor_frozen_count", 0) or 0
                        ),
                        # 超过 Excel 32767 字符上限、被截断的单元格。用户必须看得见，
                        # 否则那几格译文是残缺的而界面上毫无痕迹。
                        "truncated_cells": int(
                            write_stats.get("truncated_cells", 0) or 0
                        ),
                        "truncated_positions": list(
                            write_stats.get("truncated_positions") or []
                        ),
                        "success": True,
                    })
                    truncated_positions = list(
                        write_stats.get("truncated_positions") or []
                    )
                    if truncated_positions:
                        # 截断的单元格里译文是残缺的。只写日志不够——日志会被滚过去，
                        # 用户拿到的是一份看起来正常的文件。
                        shown = "、".join(truncated_positions[:5])
                        more = (
                            f" 等 {len(truncated_positions)} 处"
                            if len(truncated_positions) > 5
                            else ""
                        )
                        quality_issues.append(
                            {
                                "type": "cell_text_truncated",
                                "severity": "needs_action",
                                "count": len(truncated_positions),
                                "file": file_item.name,
                                "message": (
                                    f"{file_item.name} 有 {len(truncated_positions)} 个单元格"
                                    "超过 Excel 的 32767 字符上限，译文已被截断"
                                    f"（{shown}{more}）。这些格子的内容是不完整的。"
                                ),
                                "positions": truncated_positions,
                            }
                        )
                    self._log("OK", f"文件完成：{file_item.name}（{write_elapsed:.2f}s）")
                except Exception as e:
                    logger.debug(f"文件写入失败 {file_item.name} 原始错误：{e!r}")
                    write_reason = user_facing_reason(
                        e,
                        fallback="译文文件没能写出，请检查输出目录的剩余空间和写入权限。",
                    )
                    self._log("ERROR", f"文件写入失败 {file_item.name}：{write_reason}")
                    self._task_logger.file_error(file_item.name, write_reason)
                    file_results.append({
                        "name": file_item.name,
                        "source_path": str(file_item.path),
                        "source_relative_path": self._relative_source_path(file_item.path),
                        "format": file_item.format,
                        "conversion_mode": file_conversion_modes.get(str(file_item.path), "native_xlsx"),
                        "status": "failed",
                        "success": False,
                        "error": write_reason,
                    })
                finally:
                    # 清理 .xls 转换后的临时 .xlsx 文件
                    if process_path != file_item.path:
                        try:
                            os.remove(process_path)
                        except Exception as e:
                            self._log(
                                "WARN",
                                f"临时文件清理失败 {process_path.name}: "
                                f"{user_facing_reason(e, fallback='临时文件删不掉。')}",
                            )

            # 阶段 3 收尾进度
            self._queue.put(ProgressMsg(
                phase_index=3, phase_total=phase_total, phase_name="生成文件",
                step_done=len(self._files), step_total=len(self._files),
            ))

            phase3_elapsed = (datetime.now() - t_phase3).total_seconds()
            self._log("OK", f"[阶段 3 完成] 文件写入完毕（{phase3_elapsed:.2f}s）")

            elapsed = (datetime.now() - start_ts).total_seconds()
            if need_autofit:
                self._queue.put(StatusMsg(phase_desc=f"状态：[阶段 4/{phase_total}] 正在准备 Excel 精调..."))
            else:
                self._queue.put(StatusMsg(phase_desc="状态：[收尾中] 正在整理任务结果..."))

            # ── 批量 AutoFit：仅在未锁定行高时启用（且开关打开）──────────
            if need_autofit:
                out_paths = [
                    Path(r["output"]) for r in file_results
                    if r.get("success") and r.get("output")
                ]
                if out_paths:
                    if reuse_excel_for_autofit:
                        self._queue.put(StatusMsg(phase_desc=f"状态：[阶段 4/{phase_total}] 正在复用阶段 1 的 Excel 进程精调行高..."))
                    else:
                        self._queue.put(StatusMsg(phase_desc=f"状态：[阶段 4/{phase_total}] 正在启动干净 Excel 进程精调行高..."))
                    self._queue.put(ProgressMsg(
                        phase_index=4,
                        phase_total=phase_total,
                        phase_name="Excel 精调",
                        step_done=0,
                        step_total=len(out_paths),
                    ))
                    t0 = datetime.now()
                    self._log(
                        "INFO",
                        f"开始 Excel AutoFit，共 {len(out_paths)} 个文件 | policy={excel_policy}",
                    )

                    def autofit_progress_cb(done, total, current_file):
                        self._queue.put(ProgressMsg(
                            phase_index=4,
                            phase_total=phase_total,
                            phase_name="Excel 精调",
                            step_done=done,
                            step_total=total,
                        ))
                        if current_file is not None and done < total:
                            self._queue.put(StatusMsg(
                                phase_desc=f"状态：[阶段 4/{phase_total}] 正在精调 Excel 行高：{current_file.name}"
                            ))

                    if reuse_excel_for_autofit:
                        app = _get_excel_app()
                        autofit_success = bilingual_writer.autofit_files_batch(
                            out_paths,
                            app=app,
                            log_callback=lambda msg: self._log(
                                "WARN" if msg.startswith("[WARN]") else "INFO", msg
                            ),
                            progress_callback=autofit_progress_cb,
                        )
                    else:
                        autofit_success = _run_autofit_with_guard(out_paths, autofit_progress_cb)

                    self._queue.put(ProgressMsg(
                        phase_index=4,
                        phase_total=phase_total,
                        phase_name="Excel 精调",
                        step_done=len(out_paths),
                        step_total=len(out_paths),
                    ))
                    autofit_elapsed = (datetime.now() - t0).total_seconds()
                    if autofit_success:
                        self._log("INFO", f"Excel AutoFit 完成，耗时 {autofit_elapsed:.2f}s | policy={excel_policy}")
                        self._task_logger.info(
                            f"批量AutoFit完成 | 文件数={len(out_paths)} | 耗时={autofit_elapsed:.3f}s | excel_policy={excel_policy}"
                        )
                    else:
                        self._log("WARN", f"Excel AutoFit 未完全完成，已保留 Python 估算行高 | policy={excel_policy}")
                        self._task_logger.warning(
                            f"批量AutoFit未完全完成 | 文件数={len(out_paths)} | 耗时={autofit_elapsed:.3f}s | excel_policy={excel_policy}"
                        )
                else:
                    self._queue.put(StatusMsg(phase_desc=f"状态：[阶段 4/{phase_total}] 无可精调文件，已跳过 Excel 精调。"))
                    self._queue.put(ProgressMsg(
                        phase_index=4,
                        phase_total=phase_total,
                        phase_name="Excel 精调",
                        step_done=1,
                        step_total=1,
                    ))

            if excel_output.lock_row_height and excel_output.enable_excel_autofit:
                self._log("INFO", '已启用"锁定行高，缩小字号"，跳过 Excel AutoFit。')
        except TaskStopped as e:
            stopped_message = str(e)
        except ApiKeyTemporarilyUnavailableError as e:
            fatal_error_message = str(e)
        finally:
            if excel_app is not None:
                self._log("INFO", "清理全局 Excel 进程...")
                _cleanup_excel_app(status_msg="状态：[收尾中] 正在清理 Excel 进程...")

        if stopped_message is not None:
            elapsed = (datetime.now() - start_ts).total_seconds()
            self._log("WARN", stopped_message)
            self._task_logger.warning(stopped_message)
            self._task_logger.task_end(
                elapsed_sec=elapsed,
                file_results=file_results,
            )
            contract = self._build_result_contract(
                file_results=file_results,
                output_dir=str(output_dir),
                tm_hit_count=tm_hit_count,
                api_text_count=api_call_count,
                elapsed_sec=elapsed,
                stopped=True,
                source_lang=source_lang,
                target_lang=target_lang,
                preflights=file_language_preflights,
                file_texts=file_texts,
                actual_results=normal_api_language_results,
                review_marks=excel_review_marks,
            )
            self._queue.put(
                StoppedMsg(
                    message=stopped_message,
                    output_dir=str(output_dir),
                    **contract,
                )
            )
            return

        if fatal_error_message is not None:
            elapsed = (datetime.now() - start_ts).total_seconds()
            self._log("ERROR", fatal_error_message)
            self._task_logger.error(fatal_error_message)
            self._task_logger.task_end(
                elapsed_sec=elapsed,
                file_results=file_results,
            )
            contract = self._build_result_contract(
                file_results=file_results,
                output_dir=str(output_dir),
                tm_hit_count=tm_hit_count,
                api_text_count=api_call_count,
                elapsed_sec=elapsed,
                stopped=False,
                source_lang=source_lang,
                target_lang=target_lang,
                preflights=file_language_preflights,
                file_texts=file_texts,
                actual_results=normal_api_language_results,
                review_marks=excel_review_marks,
                error_message=fatal_error_message,
            )
            self._queue.put(
                ErrorMsg(
                    message=fatal_error_message,
                    output_dir=str(output_dir),
                    **contract,
                )
            )
            return

        elapsed = (datetime.now() - start_ts).total_seconds()

        # ── 任务日志：记录结束信息 ────────────────────────────────────
        self._task_logger.task_end(
            elapsed_sec  = elapsed,
            file_results = file_results,
        )
        contract = self._build_result_contract(
            file_results=file_results,
            output_dir=str(output_dir),
            tm_hit_count=tm_hit_count,
            api_text_count=api_call_count,
            elapsed_sec=elapsed,
            stopped=False,
            source_lang=source_lang,
            target_lang=target_lang,
            preflights=file_language_preflights,
            file_texts=file_texts,
            actual_results=normal_api_language_results,
            review_marks=excel_review_marks,
        )
        self._queue.put(DoneMsg(
            output_dir     = str(output_dir),
            file_results   = file_results,
            elapsed_sec    = elapsed,
            tm_hit_count   = tm_hit_count,
            api_call_count = api_call_count,
            issues         = quality_issues,
            **contract,
        ))

    def store_api_results_in_tm(self, **kwargs) -> tuple[int, str | None]:
        """Write TM without ever letting that write kill a finished translation.

        TM insertion is bookkeeping that runs *after* every API call has been
        paid for and *before* the translated files are written.  A locked
        database — someone running a long write on the glossary page — used to
        raise straight out of the worker thread and take the whole task with
        it, so any failure here becomes a warning plus a result note instead.
        """
        try:
            return self._write_api_results_to_tm(**kwargs), None
        except Exception as tm_error:  # noqa: BLE001 - never lose finished work
            message = (
                "翻译记忆库写入失败，本次译文不受影响，"
                f"但这批词条没有存入词库：{tm_error}"
            )
            logger.warning(message)
            return 0, message

    def _write_api_results_to_tm(
        self,
        *,
        auto_source_lang: bool,
        normal_api_language_results: dict,
        normal_api_translations: dict,
        text_source_scopes: dict,
        target_lang: str,
        lang_pair,
        max_len: int,
        engine_name: str,
    ) -> int:
        """Store this run's API results in TM and return the entry count."""
        if auto_source_lang:
            # The model-reported item source language is the final gate.
            # ``mixed``/``und``/``auto`` and anything outside that file's
            # preflight scope are rejected by TM manager.
            tm_entries = []
            for source_text, item in normal_api_language_results.items():
                translation = item.translation
                source_scopes = text_source_scopes.get(source_text, [])
                source_in_every_file_scope = bool(source_scopes) and all(
                    item.source_lang in scope for scope in source_scopes
                )
                tm_entries.append(
                    {
                        "source_text": source_text,
                        "translation": translation,
                        "source_lang": item.source_lang,
                        "tm_eligible": item.tm_eligible
                        and source_in_every_file_scope
                        and should_store_translation_in_tm(source_text, translation),
                    }
                )
            return tm_manager.insert_auto_entries(
                tm_entries,
                target_lang,
                max_len,
                engine_name,
                task_id=self.task_id,
            )

        new_pairs = [
            (k, v)
            for k, v in normal_api_translations.items()
            if should_store_translation_in_tm(k, v)
        ]
        return tm_manager.insert_batch(
            new_pairs,
            lang_pair,
            max_len,
            engine_name,
            sync_reverse=False,
        )

    def _log_excel_coverage_plan(self, file_name: str, plan) -> None:
        """Report one file's untranslated-only plan, including the empty case."""
        summary = plan.summary
        self._log(
            "INFO",
            (
                "  → 补译识别："
                f"待补 {summary.get('source_only', 0)}，"
                f"已覆盖 {summary.get('covered', 0)}，"
                f"不确定跳过 {summary.get('ambiguous', 0)}"
            ),
        )
        # 「一条都不用补」是合法结果，但那样输出文件会和原文一模一样。不明说的话，
        # 用户看到的就是一份没翻译的文件，看不出是「本来就不用翻」还是程序没干活。
        if not summary.get("source_only", 0):
            self._log(
                "WARN",
                f"  → {file_name}：没有找到需要补译的内容，输出文件会和原文一致。"
                "如果这份文件其实还没翻译过，请关掉「仅补译未翻译内容」再跑一次。",
            )

    def _rebuild_coverage_plans_after_preflight(
        self,
        *,
        process_paths: list[Path],
        coverage_plans: list,
        file_texts: list,
        file_results: list[dict],
        file_conversion_modes: dict[str, str],
        global_unique_texts: set[str],
        target_lang: str,
        source_lang: str,
        formula_display_value_backfill: bool,
    ) -> int:
        """Build the untranslated-only plans now that auto-detect settled the language.

        补译判定得先知道源语言是哪一门，才分得清「原文」和「已经翻好的译文」；而
        自动识别要等词条提出来、发给模型问过一轮才有答案。所以自动识别 + 补译时，
        阶段 1 先按普通方式取一遍全量候选当预检样本，真正的待补清单留到这里重算。
        返回重算后的候选文本总数（Excel 进程策略要用）。
        """
        global_unique_texts.clear()
        raw_text_count = 0
        self._log("INFO", f"[补译模式] 源语言已识别为 {source_lang}，开始按它重算待补清单。")
        for fi, file_item in enumerate(self._files):
            if fi >= len(process_paths) or fi >= len(coverage_plans):
                continue
            if any(
                r.get("source_path") == str(file_item.path) and not r.get("success")
                for r in file_results
            ):
                continue
            try:
                coverage_plan = build_excel_coverage_plan(
                    process_paths[fi],
                    target_lang=target_lang,
                    source_lang=source_lang,
                    formula_display_value_backfill=formula_display_value_backfill,
                )
            except Exception as e:  # noqa: BLE001 - 单个文件读失败不该带走整批
                logger.debug(f"补译识别失败原始错误 {file_item.name}：{e!r}")
                reason = user_facing_reason(
                    e,
                    fallback="这个文件打不开，可能已损坏或不是真正的 Excel 文件。",
                )
                self._log("ERROR", f"补译识别失败 {file_item.name}：{reason}")
                self._task_logger.file_error(file_item.name, f"补译识别失败: {reason}")
                file_results.append({
                    "name": file_item.name,
                    "source_path": str(file_item.path),
                    "source_relative_path": self._relative_source_path(file_item.path),
                    "format": file_item.format,
                    "conversion_mode": file_conversion_modes.get(
                        str(file_item.path), "native_xlsx"
                    ),
                    "status": "failed",
                    "success": False,
                    "error": f"补译识别失败: {reason}",
                })
                file_texts[fi] = set()
                continue
            coverage_plans[fi] = coverage_plan
            text_set = set(coverage_plan.source_texts)
            file_texts[fi] = text_set
            raw_text_count += len(text_set)
            global_unique_texts.update(text_set)
            self._log("INFO", f"[补译识别] {file_item.name}")
            self._log_excel_coverage_plan(file_item.name, coverage_plan)
        self._log(
            "OK",
            f"[补译识别完成] 全部文件合计 {len(global_unique_texts)} 处待补译文本。",
        )
        return raw_text_count

    @staticmethod
    def _decide_excel_policy(
        need_autofit: bool,
        xls_file_count: int,
        total_sheet_count: int,
        raw_text_count: int,
    ) -> tuple[str, str]:
        """Decide whether to reuse the stage-1 Excel process or start a clean one."""
        if not need_autofit:
            return "split", "autofit_disabled"
        if xls_file_count <= 0:
            return "split", "no_xls_conversion"

        if xls_file_count <= 3 and total_sheet_count <= 15 and raw_text_count <= 800:
            return "reuse", "light_load"

        if xls_file_count >= 10 or total_sheet_count >= 40 or raw_text_count >= 10000:
            return "split", "heavy_load"

        risk_votes = 0
        if xls_file_count >= 4:
            risk_votes += 1
        if total_sheet_count >= 20:
            risk_votes += 1
        if raw_text_count >= 2000:
            risk_votes += 1

        if risk_votes >= 2:
            return "split", "mid_load_vote"
        return "reuse", "mid_load_vote"

    def _relative_source_path(self, path: Path) -> str:
        """Return a stable, user-facing relative path for a file result."""
        root = self._source_root
        if root is not None:
            try:
                return str(path.relative_to(root))
            except ValueError:
                pass
        return path.name

    def _build_result_contract(
        self,
        *,
        file_results: list[dict],
        output_dir: str,
        tm_hit_count: int,
        api_text_count: int,
        elapsed_sec: float,
        stopped: bool,
        source_lang: str,
        target_lang: str,
        preflights: dict,
        file_texts: list[set[str]],
        actual_results: dict,
        review_marks: dict[str, str],
        error_message: str = "",
    ) -> dict[str, object]:
        """Build the Phase 4 result contract without exposing document text."""
        files = [dict(item) for item in file_results]
        terminal_sources = {str(item.get("source_path") or "") for item in files}
        if stopped:
            for item in self._files:
                if str(item.path) in terminal_sources:
                    continue
                files.append(
                    {
                        "name": item.name,
                        "source_path": str(item.path),
                        "source_relative_path": self._relative_source_path(item.path),
                        "format": getattr(item, "format", item.path.suffix.lstrip(".")),
                        "conversion_mode": "not_started",
                        "status": "unstarted",
                        "success": False,
                        "error": "任务在开始该文件前已停止。",
                    }
                )

        actual_by_file: list[dict[str, object]] = []
        for index, item in enumerate(self._files):
            preflight = preflights.get(str(item.path))
            counts: dict[str, int] = {}
            for text in file_texts[index] if index < len(file_texts) else ():
                actual = actual_results.get(text)
                code = str(getattr(actual, "source_lang", "") or "")
                if code:
                    counts[code] = counts.get(code, 0) + 1
            actual_by_file.append(
                {
                    "source_path": self._relative_source_path(item.path),
                    "preflight": (
                        preflight.to_dict(target_lang)
                        if preflight is not None
                        else {
                            "source_langs": [source_lang]
                            if source_lang != "auto"
                            else [],
                            "requested": False,
                            "request_count": 0,
                        }
                    ),
                    "actual_source_counts": counts,
                }
            )

        review_counts: dict[str, int] = {}
        review_items: list[dict[str, object]] = []
        for file_item in files:
            for item in file_item.get("review_items", []) or []:
                if not isinstance(item, dict):
                    continue
                category = str(item.get("category") or "")
                if category:
                    review_counts[category] = review_counts.get(category, 0) + 1
                review_items.append(
                    {
                        "file": file_item.get("source_relative_path", ""),
                        "worksheet": item.get("worksheet", ""),
                        "cell": item.get("cell", ""),
                        "category": category,
                        "action": item.get("action", ""),
                    }
                )
        if not review_items:
            for mark in review_marks.values():
                review_counts[mark] = review_counts.get(mark, 0) + 1
        succeeded = sum(1 for item in files if item.get("status") == "succeeded")
        failed = sum(1 for item in files if item.get("status") == "failed")
        unstarted = sum(1 for item in files if item.get("status") == "unstarted")
        return {
            "files": files,
            "kpi": {
                "selected_file_count": len(self._files),
                "succeeded_file_count": succeeded,
                "failed_file_count": failed,
                "unstarted_file_count": unstarted,
                "output_dir": output_dir,
                "elapsed_sec": elapsed_sec,
                "tm_hit_count": tm_hit_count,
                "model_translation_text_count": api_text_count,
            },
            "review": {
                "counts": review_counts,
                "total_count": sum(review_counts.values()),
                "items": review_items,
            },
            # Which connection actually did the work. A run that failed over
            # halfway was attributed entirely to the connection frozen at start,
            # so a user comparing two runs' quality had no way to see that the
            # second half came from a different provider.
            "connections": self._connection_switch_summary(),
            "language": {
                "mode": "automatic" if self._source_lang == "auto" else "manual",
                "selected_source_lang": self._source_lang,
                "target_lang": target_lang,
                "files": actual_by_file,
            },
            "error": {"message": error_message} if error_message else {},
        }

    def _connection_switch_summary(self) -> dict[str, object]:
        """The failover history, in the shape the task center and report read."""
        with self._switch_lock:
            switches = [dict(item) for item in self._connection_switches]
        summary: dict[str, object] = {
            "switch_count": len(switches),
            "switches": switches,
        }
        if switches:
            summary["final_label"] = switches[-1].get("to_label", "")
            summary["started_with_label"] = switches[0].get("from_label", "")
        return summary

    @staticmethod
    def _collect_texts(
        real_path: Path,
        item_name: str,
        *,
        target_lang: str = "",
        source_lang: str = "zh",
    ) -> tuple[list[str], int]:
        """用 openpyxl 快速读取文件中需要翻译的所有唯一词条，并返回工作表数量。"""
        from openpyxl import load_workbook
        seen: set[str] = set()
        wb = load_workbook(str(real_path), read_only=True, data_only=True)
        try:
            worksheets = wb.worksheets
            for ws in worksheets:
                for row in ws.iter_rows(values_only=True):
                    for val in row:
                        if isinstance(val, str):
                            t = val.strip()
                            if (
                                t
                                and t not in seen
                                and should_translate(
                                    t,
                                    target_lang=target_lang,
                                    source_lang=source_lang,
                                )
                            ):
                                seen.add(t)
        finally:
            wb.close()
        return list(seen), len(worksheets)
