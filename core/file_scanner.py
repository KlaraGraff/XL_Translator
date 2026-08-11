"""
文件扫描模块。
支持两种输入：
1) 文件夹路径：递归扫描目录下所有 .xlsx / .xls
2) 单文件路径：若为支持类型则仅返回该文件

嵌入图片 / 文本框计数：`core.image_detector` 那套按 openpyxl 对象锚点解析图片的
方案已下线（见 KNOWN-ISSUE-VAL-006），这里改用与 `core/xlsx_patcher.py` 相同的
思路——把 .xlsx 当 zip 直接解析目标部件的 XML，只统计「有多少」，不读取、不搬运
图片本身。用于扫描阶段给界面提示「这些内容当前不会被翻译」，不接入翻译主流程。
"""
# KNOWN-ISSUE-VAL-006:
# core.image_detector（按 openpyxl 对象锚点解析图片）仍保留但未启用。
# 本模块的图片/文本框计数走独立的 zip+XML 解析路径，与 image_detector 无关。
# See docs/KNOWN_ISSUES.md for why image_detector source is retained but offline.
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

from loguru import logger
from lxml import etree

from core.user_facing_errors import humanize_error
from core.xlsx_patcher import NS_A, NS_XDR

SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xls"}
# Translator owns output folders with this marker.  Keeping the exclusion at
# the scanner boundary prevents a recursive scan from treating a previous
# bilingual result as new source input.
GENERATED_OUTPUT_DIR_MARKER = "_翻译输出_"


@dataclass
class FileItem:
    path: Path
    name: str                          # 不含扩展名
    size_kb: float
    sheets: list[str] = field(default_factory=list)
    original_path: Path | None = None  # 用于记录原 .xls 路径
    relative_path: str = ""
    format: str = "xlsx"
    risk: dict[str, object] = field(default_factory=dict)
    # 非空文本单元格总数，与 Word 的 paragraph_count 同为不去重、不过滤的结构计数；
    # 运行时真正送模型的词条数会在全局去重和语种过滤后小得多。
    text_cell_count: int = 0
    # 嵌入图片数量：常规浮动图片（xl/drawings/drawingN.xml 里的 <xdr:pic>）
    # + WPS 单元格图片（xl/cellimages.xml 的 <etc:cellImage>，供 DISPIMG 公式引用）
    # + Excel「置于单元格内」的图片（xl/richData/rdrichvalue.xml 里结构类型为
    #   _localImage 的 <rv> 条目；股票、地理位置等非图片富数据已排除）。
    # None 表示扫描阶段无法确认：目前仅 .xls 走这个分支（需先经 Excel COM 转换
    # 成 .xlsx 才能看到这些部件），不代表「确认没有图片」。
    image_count: int | None = None
    # 含文字的文本框/形状数量：xl/drawings/drawingN.xml 里 <xdr:sp> 元素中，
    # 其 <a:t> 至少有一处非空文本的个数；纯装饰性形状（无文字）不计入。
    # None 语义同 image_count：目前仅 .xls 为 None。
    shape_text_count: int | None = None
    # 带批注的单元格数量（批注文字不翻译，原样保留）。同一格上的多条回复算一条。
    # None 语义同 image_count：目前仅 .xls 为 None。
    comment_count: int | None = None


@dataclass
class ScanSkippedItem:
    """One source that was visible to the scan but cannot be selected."""

    path: Path
    relative_path: str
    reason: str
    format: str = ""


@dataclass
class ExcelScanResult:
    """Typed Excel scan result used by the local API/UI contract.

    ``scan_path`` remains a compatibility wrapper returning only selectable
    items; new code should call :func:`scan_excel_sources` when it must expose
    skipped files, counts and `.xls` compatibility risk to the user.
    """

    root: Path
    items: list[FileItem] = field(default_factory=list)
    skipped: list[ScanSkippedItem] = field(default_factory=list)

    @property
    def summary(self) -> dict[str, int]:
        # image_count / shape_text_count 在部分文件（目前是 .xls）上是 None
        # （「数不出来」），不能当 0 加进汇总，否则会和「确认没有」混为一谈。
        # 因此汇总里额外给出 *_unknown_files，统计有多少文件的计数是未知的。
        known_image_counts = [
            item.image_count for item in self.items if item.image_count is not None
        ]
        known_shape_text_counts = [
            item.shape_text_count
            for item in self.items
            if item.shape_text_count is not None
        ]
        return {
            "scanned_count": len(self.items),
            "selected_count": len(self.items),
            "sheet_count": sum(len(item.sheets) for item in self.items),
            "text_cell_count": sum(item.text_cell_count for item in self.items),
            "xls_count": sum(1 for item in self.items if item.format == "xls"),
            "skipped_count": len(self.skipped),
            "image_count": sum(known_image_counts),
            "image_count_unknown_files": sum(
                1 for item in self.items if item.image_count is None
            ),
            "shape_text_count": sum(known_shape_text_counts),
            "shape_text_count_unknown_files": sum(
                1 for item in self.items if item.shape_text_count is None
            ),
            "comment_count": sum(
                item.comment_count
                for item in self.items
                if item.comment_count is not None
            ),
            "comment_count_unknown_files": sum(
                1 for item in self.items if item.comment_count is None
            ),
        }

    @property
    def risk(self) -> dict[str, object]:
        xls_count = self.summary["xls_count"]
        return {
            "has_xls": bool(xls_count),
            "xls_count": xls_count,
            "requires_explicit_compatibility_confirmation": bool(xls_count),
            "message": (
                "检测到 .xls 文件：优先使用本机 Microsoft Excel 高保真转换；"
                "若选择兼容转换，复杂样式、合并单元格、图片、图表和宏可能无法完整保留。"
                if xls_count
                else ""
            ),
        }


def is_supported_excel_file(path: str | Path) -> bool:
    """判断是否为可处理的单个 Excel 文件（排除 ~ 临时文件）。"""
    path = Path(path)
    return (
        path.is_file()
        and path.suffix.lower() in SUPPORTED_EXCEL_SUFFIXES
        and not path.name.startswith("~")
    )


def _is_generated_output(path: Path) -> bool:
    """跳过程序自己生成的输出目录，避免双语结果在下次扫描时被重复纳入任务。"""
    return any(GENERATED_OUTPUT_DIR_MARKER in part for part in path.parts)


def scan_path(path: str | Path) -> list[FileItem]:
    """Compatibility entry point returning selectable Excel items only."""
    return scan_excel_sources(path).items


def scan_excel_sources(path: str | Path) -> ExcelScanResult:
    """Recursively scan Excel input while retaining skipped-source evidence."""
    source = Path(path).expanduser()
    root = source if source.is_dir() else source.parent
    result = ExcelScanResult(root=root)
    if not source.exists():
        reason = f"路径不存在：{source}"
        logger.warning(reason)
        result.skipped.append(ScanSkippedItem(source, source.name, reason))
        return result

    if source.is_file():
        _scan_one_excel_file(source, source.parent, result)
    elif source.is_dir():
        for candidate in sorted(source.rglob("*")):
            if not candidate.is_file():
                continue
            if candidate.name.startswith("~"):
                continue
            try:
                relative = candidate.relative_to(source)
            except ValueError:
                relative = candidate.name
            if _is_generated_output(Path(relative)):
                continue
            if candidate.suffix.lower() not in SUPPORTED_EXCEL_SUFFIXES:
                continue
            _scan_one_excel_file(candidate, source, result)
    else:
        result.skipped.append(
            ScanSkippedItem(source, source.name, f"路径既不是文件也不是目录：{source}")
        )

    result.items.sort(key=lambda item: str(item.path))
    result.skipped.sort(key=lambda item: str(item.path))
    logger.info(
        f"Excel 扫描完成：{source}，可选 {len(result.items)}，跳过 {len(result.skipped)}"
    )
    return result


def scan_folder(root: str | Path) -> list[FileItem]:
    """
    递归扫描文件夹，返回所有 Excel 文件列表。
    跳过以 ~ 开头的临时文件（Excel 打开时产生）。
    """
    return scan_excel_sources(root).items


def _scan_one_excel_file(
    path: Path,
    root: Path,
    result: ExcelScanResult,
) -> None:
    if not is_supported_excel_file(path):
        result.skipped.append(
            ScanSkippedItem(
                path=path,
                relative_path=_relative_path(path, root),
                reason="不支持的 Excel 文件或 Office 临时文件",
                format=path.suffix.lower().lstrip("."),
            )
        )
        return
    try:
        result.items.append(_build_file_item(path, root=root))
    except Exception as exc:  # scan must not hide a corrupt/unreadable file
        # 这句会原样出现在「查看扫描报告」里。原来直接拼 str(exc)，用户看到的是
        # File is not a zip file 之类的英文库报错，既不知道是哪一类问题也不知道怎么办。
        message = "读取失败：" + humanize_error(
            exc, fallback="这个文件读不出来，可能已损坏或格式与后缀不符。"
        )
        logger.warning(f"扫描文件失败 {path.name}：{exc!r}")
        result.skipped.append(
            ScanSkippedItem(
                path=path,
                relative_path=_relative_path(path, root),
                reason=message,
                format=path.suffix.lower().lstrip("."),
            )
        )


def _count_xlsx_text_cells(wb) -> int:
    """统计 .xlsx 内的非空文本单元格总数（不去重、不按语种过滤）。"""
    total = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if isinstance(value, str) and value.strip():
                    total += 1
    return total


def _count_xls_text_cells(wb) -> int:
    """统计 .xls 内的非空文本单元格总数，无需先转换为 .xlsx。"""
    total = 0
    for index in range(wb.nsheets):
        sheet = wb.sheet_by_index(index)
        try:
            for row in range(sheet.nrows):
                for value in sheet.row_values(row):
                    if isinstance(value, str) and value.strip():
                        total += 1
        finally:
            # on_demand 打开时逐表释放，避免大文件把整本工作簿留在内存里
            wb.unload_sheet(index)
    return total


_DRAWING_PART_RE = re.compile(r"^xl/drawings/drawing\d+\.xml$")
# 传统批注（Excel 里的「注释」）与新版对话式批注。Excel 写新版批注时会同时留一份
# 传统批注做兼容，两边的 ref 指的是同一个格子——按 (部件序号, ref) 去重，否则同一条
# 批注会被数两遍。序号来自文件名：comments3.xml 与 threadedComment3.xml 对同一张表。
_COMMENTS_PART_RE = re.compile(r"^xl/comments(\d+)\.xml$")
_THREADED_COMMENTS_PART_RE = re.compile(
    r"^xl/threadedComments/threadedComment(\d+)\.xml$"
)
_CELL_IMAGES_PART = "xl/cellimages.xml"
_RICH_VALUE_PART = "xl/richData/rdrichvalue.xml"
_RICH_VALUE_STRUCTURE_PART = "xl/richData/rdrichvaluestructure.xml"
_RICH_VALUE_REL_PART = "xl/richData/richValueRel.xml"
# 「置于单元格内」的本地图片在结构表里的类型名；其余类型（_linkedEntity2 等
# 股票/地理位置富数据）不是图片，不能计入。
_LOCAL_IMAGE_STRUCTURE_TYPE = "_localImage"
_XDR_PIC_TAG = f"{{{NS_XDR}}}pic"
_XDR_SP_TAG = f"{{{NS_XDR}}}sp"
_A_T_TAG = f"{{{NS_A}}}t"


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1] if "}" in tag else tag


def _count_drawing_part(data: bytes) -> tuple[int, int]:
    """解析单个 xl/drawings/drawingN.xml：统计图片数与含文字形状数。

    复用 xlsx_patcher.fix_drawing_anchors 已验证过的命名空间常量和解析思路，
    不另起一套。图片：<xdr:pic> 元素（含嵌套在组合形状 <xdr:grpSp> 内的）。
    含文字形状：<xdr:sp> 且其 <a:t> 至少一处非空文本，纯装饰形状不计。
    """
    root = etree.fromstring(data)
    image_count = sum(1 for _ in root.iter(_XDR_PIC_TAG))
    shape_text_count = 0
    for sp in root.iter(_XDR_SP_TAG):
        if any((t.text or "").strip() for t in sp.iter(_A_T_TAG)):
            shape_text_count += 1
    return image_count, shape_text_count


def _count_cell_images_part(data: bytes) -> int:
    """WPS xl/cellimages.xml：每个 <etc:cellImage> 子元素对应一张被 DISPIMG
    公式引用的单元格内图片，按子元素计数（不依赖其内部是否套 <xdr:pic>）。
    """
    root = etree.fromstring(data)
    return sum(1 for child in root if _local_name(child.tag) == "cellImage")


def _local_image_structure_indexes(data: bytes) -> set[int]:
    """xl/richData/rdrichvaluestructure.xml：哪些结构下标代表「本地图片」。

    富数据条目 <rv s="N"> 里的 N 是结构表下标（按 <s> 出现顺序从 0 起）。只有
    t="_localImage" 的结构才是「置于单元格内」的图片，股票、地理位置等富数据
    各有自己的类型名。不做这层过滤的话，一张图都没有的股票表会被报成一堆图片。
    """
    root = etree.fromstring(data)
    structures = [child for child in root if _local_name(child.tag) == "s"]
    return {
        index
        for index, structure in enumerate(structures)
        if structure.get("t") == _LOCAL_IMAGE_STRUCTURE_TYPE
    }


def _count_rich_value_images(data: bytes, image_structures: set[int]) -> int:
    """xl/richData/rdrichvalue.xml：只数结构下标落在 image_structures 里的条目。"""
    root = etree.fromstring(data)
    count = 0
    for child in root:
        if _local_name(child.tag) != "rv":
            continue
        try:
            index = int(child.get("s"))
        except (TypeError, ValueError):
            continue
        if index in image_structures:
            count += 1
    return count


def _count_rich_value_rels(data: bytes) -> int:
    """xl/richData/richValueRel.xml：<rel> 直接指向被富数据引用的图片文件。

    仅在结构表缺失或解析失败时兜底。同一张图片被多个单元格引用时会少算，但绝
    不会把非图片富数据算成图片——宁可少报，不能虚报给用户看的数字。
    """
    root = etree.fromstring(data)
    return sum(1 for child in root if _local_name(child.tag) == "rel")


def _count_comment_refs(data: bytes, tag: str) -> set[str]:
    """收集一个批注部件里所有带批注的单元格地址。

    按地址去重：一个格子上挂三条回复是一条批注，不是三条。
    """
    root = etree.fromstring(data)
    refs: set[str] = set()
    for node in root.iter():
        if _local_name(node.tag) != tag:
            continue
        ref = (node.get("ref") or "").strip()
        if ref:
            refs.add(ref)
    return refs


def _count_xlsx_comments(archive: zipfile.ZipFile, names: list[str], label: str) -> int:
    """统计 .xlsx 内带批注的单元格数量（传统批注 ∪ 新版对话式批注）。"""
    per_sheet: dict[str, set[str]] = {}
    for name in names:
        legacy = _COMMENTS_PART_RE.match(name)
        threaded = _THREADED_COMMENTS_PART_RE.match(name)
        match = legacy or threaded
        if not match:
            continue
        try:
            refs = _count_comment_refs(
                archive.read(name), "comment" if legacy else "threadedComment"
            )
        except Exception as exc:
            logger.debug(f"解析 {name} 失败（{label}）：{exc}")
            continue
        per_sheet.setdefault(match.group(1), set()).update(refs)
    return sum(len(refs) for refs in per_sheet.values())


def _count_xlsx_images_and_shapes(path: Path) -> tuple[int, int, int]:
    """统计 .xlsx 内嵌入图片数、含文字文本框/形状数、带批注单元格数。

    只读取需要的 zip 部件，不解压整包。任何部件缺失或解析失败都不应让扫描
    失败——按 0 处理并继续，损坏文件的判定仍然只看 _build_file_item 外层
    是否抛异常。
    """
    image_count = 0
    shape_text_count = 0
    comment_count = 0
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            comment_count = _count_xlsx_comments(archive, names, path.name)
            for name in names:
                if not _DRAWING_PART_RE.match(name):
                    continue
                try:
                    images, shapes = _count_drawing_part(archive.read(name))
                except Exception as exc:
                    logger.debug(f"解析 {name} 失败（{path.name}）：{exc}")
                    continue
                image_count += images
                shape_text_count += shapes

            if _CELL_IMAGES_PART in names:
                try:
                    image_count += _count_cell_images_part(
                        archive.read(_CELL_IMAGES_PART)
                    )
                except Exception as exc:
                    logger.debug(f"解析 cellimages.xml 失败（{path.name}）：{exc}")

            image_structures: set[int] | None = None
            if _RICH_VALUE_STRUCTURE_PART in names:
                try:
                    image_structures = _local_image_structure_indexes(
                        archive.read(_RICH_VALUE_STRUCTURE_PART)
                    )
                except Exception as exc:
                    logger.debug(f"解析 richData 结构表失败（{path.name}）：{exc}")

            if image_structures is not None and _RICH_VALUE_PART in names:
                try:
                    image_count += _count_rich_value_images(
                        archive.read(_RICH_VALUE_PART), image_structures
                    )
                except Exception as exc:
                    logger.debug(f"解析 richData 部件失败（{path.name}）：{exc}")
            elif _RICH_VALUE_REL_PART in names:
                try:
                    image_count += _count_rich_value_rels(
                        archive.read(_RICH_VALUE_REL_PART)
                    )
                except Exception as exc:
                    logger.debug(f"解析 richValueRel.xml 失败（{path.name}）：{exc}")
    except Exception as exc:
        logger.debug(f"统计嵌入图片/文本框/批注失败（{path.name}）：{exc}")
        return 0, 0, 0
    return image_count, shape_text_count, comment_count


def _relative_path(path: Path, root: Path) -> str:
    try:
        return str(path.relative_to(root))
    except ValueError:
        return path.name


def _build_file_item(path: Path, *, root: Path | None = None) -> FileItem:
    if path.suffix.lower() == ".xls":
        import xlrd
        wb = xlrd.open_workbook(str(path), on_demand=True)
        try:
            sheets = wb.sheet_names()
            text_cell_count = _count_xls_text_cells(wb)
        finally:
            wb.release_resources()
        # .xls（BIFF 老格式）扫描阶段读不到图片/文本框：这些部件是 OOXML
        # 概念，.xls 要先经 Excel COM 转换成 .xlsx 才谈得上。用 None 表示
        # 「未知」，不要猜成 0（那会被误读成「确认没有」）。
        image_count = None
        shape_text_count = None
        comment_count = None
    else:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        try:
            sheets = wb.sheetnames
            text_cell_count = _count_xlsx_text_cells(wb)
        finally:
            wb.close()
        image_count, shape_text_count, comment_count = _count_xlsx_images_and_shapes(path)

    size_kb = path.stat().st_size / 1024

    original_path = path if path.suffix.lower() == ".xls" else None

    return FileItem(
        path=path,
        name=path.stem,
        size_kb=round(size_kb, 1),
        sheets=sheets,
        original_path=original_path,
        relative_path=_relative_path(path, root or path.parent),
        format=path.suffix.lower().lstrip("."),
        text_cell_count=text_cell_count,
        image_count=image_count,
        shape_text_count=shape_text_count,
        comment_count=comment_count,
        risk=(
            {
                "compatibility_required": True,
                "message": (
                    ".xls 需通过 Microsoft Excel 高保真转换，或经用户明确确认后"
                    "使用可能损失样式/合并单元格/图片/图表/宏的兼容转换。"
                ),
            }
            if path.suffix.lower() == ".xls"
            else {}
        ),
    )
