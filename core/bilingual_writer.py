"""
双语回填与文件导出模块。

回填规则：
  - 直接修改目标单元格，格式：原文 + "\\n" + 译文
  - 若译文与原文实质相同（剔除空格大小写），仅保留原文
  - 不插入新行或新列

输出：
  - 输出目录：{源目录}_翻译输出_{timestamp}/
  - 文件名前缀：双语({语言})_{原文件名}.xlsx
  - 可选：保留原始中文分表（sheet 名称加 _原文 后缀）
"""
import re
import shutil
import stat
import tempfile
from datetime import datetime
from pathlib import Path
from copy import copy

from loguru import logger

from config import (
    BILINGUAL_SEPARATOR,
    PRINT_GUARD_LINE_HEIGHT_MULTIPLIER,
    PRINT_GUARD_FONT_STEP,
    PRINT_GUARD_FONT_FLOOR,
)
from core.language_registry import get_target_lang_display
from core.translation_protocol import extract_replace_translation, is_replace_translation
from core.translation_filter import should_translate

_INVALID_FILENAME_FRAGMENT_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]+')


def _ensure_owner_writable(path: Path) -> None:
    """Ensure copied outputs stay writable even when the source workbook is read-only."""
    current_mode = path.stat().st_mode
    if current_mode & stat.S_IWUSR:
        return
    path.chmod(current_mode | stat.S_IWUSR)


def resolve_custom_output_dir(custom_output_dir: str | Path | None) -> Path | None:
    """Normalize a custom output root; return None for empty input."""
    if custom_output_dir is None:
        return None

    normalized = str(custom_output_dir).strip()
    if not normalized:
        return None

    return Path(normalized).expanduser()


def _find_blocking_existing_path(target_path: Path) -> Path | None:
    """Find the first existing ancestor that blocks directory creation."""
    current = target_path
    while not current.exists():
        parent = current.parent
        if parent == current:
            return None
        current = parent
    return None if current.is_dir() else current


def get_custom_output_dir_error(custom_output_dir: str | Path | None) -> str | None:
    """Return a user-friendly validation error for a custom output root."""
    output_root = resolve_custom_output_dir(custom_output_dir)
    if output_root is None:
        return "自定义输出目录不能为空"

    if output_root.exists():
        if output_root.is_dir():
            return None
        return f"输出路径不是目录：{output_root}"

    blocking_path = _find_blocking_existing_path(output_root)
    if blocking_path is not None:
        return f"无法在文件路径下创建目录：{blocking_path}"

    return None


def custom_output_dir_will_be_created(custom_output_dir: str | Path | None) -> bool:
    """Whether the custom output root will be created at runtime."""
    output_root = resolve_custom_output_dir(custom_output_dir)
    if output_root is None or get_custom_output_dir_error(output_root) is not None:
        return False
    return not output_root.exists()


def build_output_dir(source_dir: str | Path, custom_output_dir: str | Path | None = None) -> Path:
    """生成带时间戳的输出目录路径（不创建目录）。
    
    :param source_dir: 源文件夹路径
    :param custom_output_dir: 自定义输出目录（None 或空字符串时使用默认位置）
    :return: 输出目录路径
    
    默认行为：在源文件夹内部创建 {源文件夹名}_翻译输出_{时间戳}
    自定义行为：在指定目录内创建 {源文件夹名}_翻译输出_{时间戳}
    """
    source_dir = Path(source_dir)
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_subdir_name = f"{source_dir.name}_翻译输出_{timestamp}"
    
    custom_output_root = resolve_custom_output_dir(custom_output_dir)
    if custom_output_root is not None:
        return custom_output_root / output_subdir_name

    # 默认：在源文件夹内部创建
    return source_dir / output_subdir_name


def write_bilingual_file(
    source_path: Path,
    output_dir: Path,
    translations: dict[str, str],
    target_lang: str,
    keep_original_sheets: bool,
    formula_display_value_backfill: bool,
    enable_print_guard: bool,
    source_lang: str = "zh",
    lock_row_height: bool = False,
    log_callback=None,
    original_path: Path | None = None,
) -> Path:
    """
    将翻译结果回填至 Excel 文件并保存至输出目录。

    :param source_path:          原始文件路径
    :param output_dir:           输出目录（已创建）
    :param translations:         {原文: 译文} 字典
    :param target_lang:          目标语言代码
    :param keep_original_sheets: 是否保留原始中文分表
    :param formula_display_value_backfill:
                                 是否对公式生成的显示文本按显示值匹配后回填
    :param enable_print_guard:   保留参数（MVP 阶段不生效）
    :param lock_row_height:      是否锁定行高并通过缩小字号适配内容
    :param log_callback:         日志回调 log_callback(msg: str)
    :param original_path:        原 .xls 路径（如果是经过转换的临时文件）
    :return:                     输出文件路径
    """
    lang_display = _sanitize_filename_fragment(
        get_target_lang_display(target_lang, include_optional=True)
    )
    
    # 确定输出文件名，处理源文件可能是临时文件的场景
    basename = original_path.name if original_path else source_path.name
    # 强制最终输出为 .xlsx（即使源文件是 .xls）
    if basename.lower().endswith(".xls"):
        basename = basename[:-4] + ".xlsx"
        
    out_name     = f"双语({lang_display})_{basename}"
    out_path     = output_dir / out_name

    output_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, out_path)
    _ensure_owner_writable(out_path)

    _write_with_openpyxl(
        file_path=out_path,
        translations=translations,
        target_lang=target_lang,
        source_lang=source_lang,
        keep_original_sheets=keep_original_sheets,
        formula_display_value_backfill=formula_display_value_backfill,
        lock_row_height=lock_row_height,
        log_callback=log_callback,
    )

    if log_callback:
        log_callback(f"[OK] 已输出：{out_path.name}")

    return out_path


def _sanitize_filename_fragment(value: str) -> str:
    """Remove Windows-illegal filename characters from user-facing fragments."""
    cleaned = _INVALID_FILENAME_FRAGMENT_RE.sub("_", str(value or "")).strip().rstrip(". ")
    return cleaned or "目标语言"


def _write_with_openpyxl(
    file_path: Path,
    translations: dict[str, str],
    target_lang: str,
    keep_original_sheets: bool,
    formula_display_value_backfill: bool,
    source_lang: str = "zh",
    lock_row_height: bool = False,
    log_callback=None,
) -> None:
    """使用 openpyxl 回填（纯文本模式，不处理图片）。"""
    # KNOWN-ISSUE-VAL-006:
    # Image-related handling is intentionally offline in the current main flow.
    # See docs/KNOWN_ISSUES.md for the retained-but-disabled rationale.
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    wb = load_workbook(str(file_path))
    wb_values = (
        load_workbook(str(file_path), data_only=True)
        if formula_display_value_backfill else None
    )

    try:
        sheet_names = list(wb.sheetnames)

        # ── 第一遍：先统一复制原文副本（保存原始内容，不含回填）────────
        if keep_original_sheets:
            for name in sheet_names:
                new_ws = wb.copy_worksheet(wb[name])
                new_ws.title = f"{name}_原文"

        # ── 第二遍：对原始 sheet 执行双语回填 ────────────────────────────
        for name in sheet_names:
            ws = wb[name]
            ws_values = wb_values[name] if wb_values is not None else None
            adjusted_cell_count = 0

            for row in ws.iter_rows():
                for cell in row:
                    # ── 前置拦截：强力清除 WPS DISPIMG 公式 ──────────────────
                    if cell.value is not None and "DISPIMG" in str(cell.value).upper():
                        cell.value = ""
                        continue

                    source_text = _resolve_cell_source_text(
                        cell,
                        ws_values,
                        formula_display_value_backfill,
                    )
                    if source_text is None:
                        continue
                    if not should_translate(
                        source_text,
                        target_lang=target_lang,
                        source_lang=source_lang,
                    ):
                        continue

                    tgt = translations.get(source_text.strip())
                    if tgt is None:
                        continue
                    if is_replace_translation(tgt):
                        final_text = extract_replace_translation(tgt).strip()
                        if not final_text:
                            continue
                        cell.value = final_text
                    else:
                        if not tgt:
                            continue
                        if source_text.strip().lower() == tgt.strip().lower():
                            continue

                        # ── 双语回填 ──────────────────────────────────────────────
                        cell.value = source_text + BILINGUAL_SEPARATOR + tgt

                    # ── 自动换行 ──────────────────────────────────────────────
                    existing = cell.alignment
                    cell.alignment = Alignment(
                        wrap_text=True,
                        horizontal=existing.horizontal,
                        vertical=existing.vertical,
                        text_rotation=existing.text_rotation,
                        indent=existing.indent,
                        shrink_to_fit=existing.shrink_to_fit,
                    )

                    if lock_row_height:
                        reached_floor = _shrink_font_to_fit_locked_row(cell, ws)
                        adjusted_cell_count += 1
                        if reached_floor and log_callback:
                            log_callback(
                                f"[WARN] {ws.title}!{cell.coordinate} 缩至最小字号 {PRINT_GUARD_FONT_FLOOR:.1f}pt 仍可能无法完全显示"
                            )

            # ── 行高策略 ────────────────────────────────────────────────
            if not lock_row_height:
                # 默认模式 / Excel 精调模式：先用 Python 自动调行高
                _auto_adjust_row_heights(ws)
            elif log_callback and adjusted_cell_count:
                log_callback(f"[INFO] 分表已锁定行高并缩字号：{name}（{adjusted_cell_count} 个单元格）")

            if log_callback:
                log_callback(f"[INFO] 分表已处理：{name}")

        wb.save(str(file_path))
    finally:
        wb.close()
        if wb_values is not None:
            wb_values.close()


def _resolve_cell_source_text(
    cell,
    ws_values,
    formula_display_value_backfill: bool,
) -> str | None:
    value = cell.value
    if not isinstance(value, str):
        return None
    if not formula_display_value_backfill or not _is_formula_cell(cell):
        return value
    if ws_values is None:
        return None

    display_value = ws_values[cell.coordinate].value
    return display_value if isinstance(display_value, str) else None


def _is_formula_cell(cell) -> bool:
    return getattr(cell, "data_type", None) == "f"


def _estimate_chars_per_line(col_width: float, font_size_pt: float) -> int:
    """按列宽+字号估算每行容纳字符数。"""
    # 以 11pt 为基准：字号越小，每行可容纳字符越多
    base_chars_per_width = 1.2
    scale = 11.0 / max(font_size_pt, 0.1)
    return max(1, int(col_width * base_chars_per_width * scale))


def _estimate_required_lines(text: str, chars_per_line: int) -> int:
    total_lines = 0
    for segment in text.split("\n"):
        total_lines += max(1, -(-len(segment) // max(chars_per_line, 1)))
    return max(1, total_lines)


def _estimate_max_visible_lines(row_height: float, font_size_pt: float) -> int:
    line_height = max(1.0, font_size_pt * PRINT_GUARD_LINE_HEIGHT_MULTIPLIER)
    return max(1, int(row_height / line_height))


def _shrink_font_to_fit_locked_row(cell, ws) -> bool:
    """锁定行高模式：逐步缩小字号适配当前行高。返回是否触底。"""
    default_col_width = 8.43
    default_row_height = 15.0

    col_dim = ws.column_dimensions.get(cell.column_letter)
    col_width = col_dim.width if (col_dim and col_dim.width) else default_col_width

    row_dim = ws.row_dimensions.get(cell.row)
    row_height = row_dim.height if (row_dim and row_dim.height) else default_row_height

    original_font = cell.font
    current_size = float(original_font.size or 11.0)
    min_size = float(PRINT_GUARD_FONT_FLOOR)
    step = float(PRINT_GUARD_FONT_STEP)

    while True:
        chars_per_line = _estimate_chars_per_line(col_width, current_size)
        required_lines = _estimate_required_lines(str(cell.value), chars_per_line)
        visible_lines = _estimate_max_visible_lines(row_height, current_size)

        if required_lines <= visible_lines:
            break

        if current_size <= min_size:
            current_size = min_size
            break

        current_size = max(min_size, round(current_size - step, 2))

    if original_font.size != current_size:
        new_font = copy(original_font)
        new_font.size = current_size
        cell.font = new_font

    return current_size <= min_size and _estimate_required_lines(
        str(cell.value), _estimate_chars_per_line(col_width, current_size)
    ) > _estimate_max_visible_lines(row_height, current_size)


def _auto_adjust_row_heights(ws) -> None:
    """自动调整行高以适配双语内容（对所有行统一处理）。

    算法：
      - 遍历每一行，计算所有单元格中最多换行行数
      - 行高 = 最大行数 × 默认字号 × 行距系数
      - 仅当内容超过 1 行时才调整（避免压缩原本宽松的行）

    :param ws: openpyxl Worksheet 对象
    """
    BASE_FONT_SIZE_PT = 11.0   # 默认字号（磅）
    LINE_HEIGHT_RATIO = 1.4    # 行距系数
    DEFAULT_COL_WIDTH = 8.43   # Excel 默认列宽（字符数）
    CHARS_PER_WIDTH   = 1.2    # 每列宽单位对应的字符数（近似值）

    for row in ws.iter_rows():
        row_num = row[0].row
        max_lines = 1

        for cell in row:
            if not cell.value or not isinstance(cell.value, str):
                continue

            # 获取列宽（字符数）
            col_letter = cell.column_letter
            col_dim = ws.column_dimensions.get(col_letter)
            col_width = col_dim.width if (col_dim and col_dim.width) else DEFAULT_COL_WIDTH

            # 估算每行可容纳的字符数
            # 中文字符宽度约为英文的 2 倍，取折中近似
            chars_per_line = max(1, int(col_width * CHARS_PER_WIDTH))

            # 统计换行符 + 估算自动换行行数
            lines = 0
            for segment in cell.value.split("\n"):
                lines += max(1, -(-len(segment) // chars_per_line))  # ceiling division
            max_lines = max(max_lines, lines)

        # 只有内容超过 1 行时才调整行高
        if max_lines > 1:
            ws.row_dimensions[row_num].height = max_lines * BASE_FONT_SIZE_PT * LINE_HEIGHT_RATIO


def autofit_files_batch(
    file_paths: list[Path],
    log_callback=None,
    app=None,
    progress_callback=None,
) -> bool:
    """使用一次 Excel 进程对多个文件批量执行 AutoFit 行高调整。

    相比逐文件启动 Excel，只付出一次进程启动开销（约 5-8s），
    N 个文件的总耗时从 N×8s 降为 5s + N×1-2s。

    :param file_paths:   要处理的 Excel 文件路径列表
    :param log_callback: 日志回调
    :param app:          如果有现成的 xlwings App，可直接传入复用
    :param progress_callback: 进度回调 progress_callback(done, total, current_file)
    :return:             True 表示成功，False 表示 xlwings 不可用（已静默降级）
    """
    if not file_paths:
        return True

    try:
        import xlwings as xw
    except ImportError:
        if log_callback:
            log_callback("[WARN] xlwings 未安装，已跳过 Excel 行高优化（当前使用 Python 估算值）")
        logger.warning("xlwings 未安装，跳过 AutoFit")
        return False

    staged_paths: list[tuple[Path, Path]] = []
    temp_workspace = tempfile.TemporaryDirectory(prefix="xl_translator_autofit_")
    try:
        temp_root = Path(temp_workspace.name)
        for index, file_path in enumerate(file_paths, start=1):
            staged_path = temp_root / f"{index:03d}_{file_path.name}"
            shutil.copy2(file_path, staged_path)
            _ensure_owner_writable(staged_path)
            staged_paths.append((file_path, staged_path))

        def _do_autofit(current_app):
            total = len(staged_paths)
            done = 0
            for original_path, staged_path in staged_paths:
                try:
                    if log_callback:
                        log_callback(
                            f"[INFO] Excel AutoFit 打开临时副本：{staged_path}（原文件：{original_path}）"
                        )
                    wb = current_app.books.open(str(staged_path))
                    for ws in wb.sheets:
                        ws.used_range.rows.autofit()
                    wb.save()
                    wb.close()
                    shutil.copy2(staged_path, original_path)
                    _ensure_owner_writable(original_path)
                    if log_callback:
                        log_callback(f"[INFO] Excel AutoFit 已回写原文件：{original_path}")
                    logger.info(f"AutoFit 完成：{original_path.name}")
                except Exception as e:
                    logger.warning(f"AutoFit 异常（{original_path.name}）：{e}")
                    if log_callback:
                        log_callback(f"[WARN] {original_path.name} AutoFit 失败，已保留 Python 估算值：{e}")
                finally:
                    done += 1
                    if progress_callback:
                        progress_callback(done, total, original_path)

        if app is not None:
            _do_autofit(app)
        else:
            with xw.App(visible=False) as new_app:
                new_app.display_alerts = False
                _do_autofit(new_app)

        if log_callback:
            log_callback(f"[INFO] Excel AutoFit 完成，共处理 {len(file_paths)} 个文件")
        return True
    except Exception as e:
        if log_callback:
            log_callback(f"[WARN] Excel AutoFit 失败，已保留 Python 估算值：{e}")
        logger.warning(f"AutoFit 批量处理异常：{e}")
        return False
    finally:
        temp_workspace.cleanup()
