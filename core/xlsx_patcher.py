"""补丁式 xlsx 写入器。

设计目标：把 .xlsx 当成 zip 包处理，只重写真正被改动的部件，其余部件**原字节照抄**。
openpyxl 的「load → 改 cell → save」是整本重建，保存时会丢弃它不认识的部件
（WPS 的 ``xl/cellimages.xml``、Excel「置于单元格内」的富数据部件、图表、宏……），
这正是翻译后嵌入式图片丢失的根因。

会被重写的部件：
  - ``xl/worksheets/sheetN.xml``  译文回填、复核标色、自动换行、行高
  - ``xl/drawings/drawingN.xml``  行高变化后把 twoCellAnchor / absoluteAnchor 改写成
                                  定尺寸 oneCellAnchor
  - ``xl/styles.xml``             追加 fill / font / xf 条目
  - ``xl/workbook.xml``、``xl/_rels/workbook.xml.rels``、``[Content_Types].xml``
                                  仅在 keep_original_sheets 克隆原文分表时
  - ``xl/calcChain.xml``          只有确实删除过公式时才丢弃，交给 Excel 重建

其余部件（sharedStrings.xml、media、cellimages.xml、richData、metadata、theme、
charts、vbaProject……）一律原字节照抄。

单元格文本回填采用 **inlineStr** 而不是往 sharedStrings.xml 追加条目，原因：
  1. 完全不需要碰 sharedStrings.xml，该部件保持逐字节不变；
  2. 不必维护 count / uniqueCount 与索引，也不必在缺失该部件时新建并注册；
  3. 每个单元格自包含，局部失败不会污染全局字符串表。
"""
from __future__ import annotations

import copy
import os
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

from lxml import etree
from loguru import logger

from config import (
    BILINGUAL_SEPARATOR,
    EXCEL_REVIEW_EXISTING_FILL_POLICY_DEFAULT,
    PRINT_GUARD_FONT_FLOOR,
    PRINT_GUARD_FONT_STEP,
    PRINT_GUARD_LINE_HEIGHT_MULTIPLIER,
    REVIEW_MARK_COLOR_DEFAULTS,
)
from core.mixed_language import MIXED_MARK_UNRESOLVED
from core.translation_filter import should_translate
from core.translation_protocol import extract_replace_translation, is_replace_translation

# ── 命名空间 ──────────────────────────────────────────────────────────────────
NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_CONTENT_TYPES = "http://schemas.openxmlformats.org/package/2006/content-types"
NS_XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
XML_SPACE = "{http://www.w3.org/XML/1998/namespace}space"

REL_TYPE_OFFICE_DOCUMENT = f"{NS_REL}/officeDocument"
REL_TYPE_WORKSHEET = f"{NS_REL}/worksheet"
REL_TYPE_DRAWING = f"{NS_REL}/drawing"
REL_TYPE_VML_DRAWING = f"{NS_REL}/vmlDrawing"
REL_TYPE_COMMENTS = f"{NS_REL}/comments"
REL_TYPE_TABLE = f"{NS_REL}/table"
REL_TYPE_CALC_CHAIN = f"{NS_REL}/calcChain"
# 与分表一一对应、不能被两个分表共享的内部部件，克隆分表时必须一并复制。
REL_TYPES_SHEET_OWNED = (REL_TYPE_DRAWING, REL_TYPE_VML_DRAWING, REL_TYPE_COMMENTS)

CT_WORKSHEET = "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"

# ── 排版估算常量（与旧 openpyxl 写入路径保持同一套数值）────────────────────────
DEFAULT_COL_WIDTH = 8.43        # Excel 默认列宽（字符数）
DEFAULT_ROW_HEIGHT = 15.0       # Excel 默认行高（磅）
BASE_FONT_SIZE_PT = 11.0        # 行高估算用的基准字号
LINE_HEIGHT_RATIO = 1.4         # 行高估算用的行距系数
CHARS_PER_WIDTH = 1.2           # 每单位列宽约可容纳的字符数
MAX_SHEET_TITLE_LEN = 31        # Excel 分表名长度上限

EMU_PER_PIXEL = 9525
EMU_PER_POINT = 12700
MAX_DIGIT_WIDTH = 7             # Calibri 11 的最大数字宽度（像素）

# Excel 行列号上限（0-based），absoluteAnchor 坐标外推越过这个界才算坏数据。
EXCEL_MAX_ROW_INDEX = 1_048_575
EXCEL_MAX_COL_INDEX = 16_383

_EXCEL_REVIEW_RISK_FONT_COLOR = "FF0000"
_EXCEL_REVIEW_MARK_COLORS = dict(REVIEW_MARK_COLOR_DEFAULTS)
_COORD_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def _m(tag: str) -> str:
    return f"{{{NS_MAIN}}}{tag}"


def _r(tag: str) -> str:
    return f"{{{NS_REL}}}{tag}"


def _pr(tag: str) -> str:
    return f"{{{NS_PKG_REL}}}{tag}"


def _ct(tag: str) -> str:
    return f"{{{NS_CONTENT_TYPES}}}{tag}"


def _xdr(tag: str) -> str:
    return f"{{{NS_XDR}}}{tag}"


def _local(el) -> str:
    return etree.QName(el).localname


def _parser() -> etree.XMLParser:
    # huge_tree：真实工作簿的 sheetData 可能远超 lxml 默认的文本节点上限。
    return etree.XMLParser(huge_tree=True)


def _parse(data: bytes):
    return etree.fromstring(data, parser=_parser())


def _serialize(root) -> bytes:
    return etree.tostring(
        root.getroottree(),
        xml_declaration=True,
        encoding="UTF-8",
        standalone=True,
    )


def _column_index(letters: str) -> int:
    """A → 1，AA → 27。"""
    index = 0
    for char in letters.upper():
        index = index * 26 + (ord(char) - 64)
    return index


def _column_letter(index: int) -> str:
    letters = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters or "A"


def _split_coordinate(coordinate: str) -> tuple[int, int] | None:
    match = _COORD_RE.match(str(coordinate or ""))
    if match is None:
        return None
    return int(match.group(2)), _column_index(match.group(1))


# ══════════════════════════════════════════════════════════════════════════════
# zip 包：只重写改动过的部件
# ══════════════════════════════════════════════════════════════════════════════
class _Package:
    """打开 .xlsx 的 zip 包，记录改写/新增/删除，保存时其余条目原字节照抄。"""

    def __init__(self, path: Path):
        self._path = Path(path)
        self._zip = zipfile.ZipFile(self._path, "r")
        self._infos = {info.filename: info for info in self._zip.infolist()}
        self._order = [info.filename for info in self._zip.infolist()]
        self._overrides: dict[str, bytes] = {}
        self._added: list[str] = []
        self._removed: set[str] = set()

    def __enter__(self) -> "_Package":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        self.close()

    def has(self, name: str) -> bool:
        if name in self._removed:
            return False
        return name in self._overrides or name in self._infos

    def read(self, name: str) -> bytes | None:
        if name in self._overrides:
            return self._overrides[name]
        if name in self._removed or name not in self._infos:
            return None
        return self._zip.read(name)

    def write(self, name: str, data: bytes) -> None:
        self._removed.discard(name)
        if name not in self._infos and name not in self._overrides:
            self._added.append(name)
        self._overrides[name] = data

    def remove(self, name: str) -> None:
        self._overrides.pop(name, None)
        if name in self._added:
            self._added.remove(name)
        self._removed.add(name)

    def close(self) -> None:
        self._zip.close()

    def save(self) -> None:
        """把包写到临时文件再整体替换，避免半成品覆盖原文件。"""
        temp_path = self._path.with_name(self._path.name + ".patching")
        with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as target:
            for name in self._order:
                if name in self._removed:
                    continue
                source_info = self._infos[name]
                if name in self._overrides:
                    target.writestr(_clone_zipinfo(source_info), self._overrides[name])
                else:
                    target.writestr(_clone_zipinfo(source_info), self._zip.read(name))
            for name in self._added:
                if name in self._removed:
                    continue
                target.writestr(
                    zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0)),
                    self._overrides[name],
                    zipfile.ZIP_DEFLATED,
                )
        self._zip.close()
        os.replace(temp_path, self._path)


def _clone_zipinfo(source: zipfile.ZipInfo) -> zipfile.ZipInfo:
    info = zipfile.ZipInfo(source.filename, date_time=source.date_time)
    info.compress_type = source.compress_type
    info.external_attr = source.external_attr
    info.internal_attr = source.internal_attr
    info.create_system = source.create_system
    return info


def _resolve_part(base_part: str, target: str) -> str:
    """把关系里的相对 Target 解析成包内绝对部件名（不带前导斜杠）。"""
    target = str(target or "")
    if target.startswith("/"):
        return target[1:]
    base_dir = os.path.dirname(base_part)
    return os.path.normpath(os.path.join(base_dir, target)).replace(os.sep, "/")


def _relative_part(base_part: str, target_part: str) -> str:
    base_dir = os.path.dirname(base_part) or "."
    return os.path.relpath(target_part, base_dir).replace(os.sep, "/")


def _rels_path(part: str) -> str:
    directory, name = os.path.split(part)
    return f"{directory}/_rels/{name}.rels" if directory else f"_rels/{name}.rels"


def _next_free_part(package: _Package, part: str) -> str:
    """按 ``dir/stem<N>.ext`` 规则找一个包内尚未占用的部件名。"""
    directory, name = os.path.split(part)
    stem, ext = os.path.splitext(name)
    prefix = stem.rstrip("0123456789") or stem
    index = 1
    while True:
        candidate = f"{directory}/{prefix}{index}{ext}" if directory else f"{prefix}{index}{ext}"
        if not package.has(candidate):
            return candidate
        index += 1


# ══════════════════════════════════════════════════════════════════════════════
# [Content_Types].xml
# ══════════════════════════════════════════════════════════════════════════════
class _ContentTypes:
    def __init__(self, data: bytes):
        self._root = _parse(data)
        self.dirty = False
        self._overrides = {
            el.get("PartName"): el
            for el in self._root
            if _local(el) == "Override"
        }
        self._defaults = {
            str(el.get("Extension") or "").lower()
            for el in self._root
            if _local(el) == "Default"
        }

    def content_type_of(self, part: str) -> str | None:
        el = self._overrides.get("/" + part)
        return el.get("ContentType") if el is not None else None

    def has_default(self, part: str) -> bool:
        return os.path.splitext(part)[1].lstrip(".").lower() in self._defaults

    def add_override(self, part: str, content_type: str) -> None:
        part_name = "/" + part
        if part_name in self._overrides:
            return
        el = etree.SubElement(self._root, _ct("Override"))
        el.set("PartName", part_name)
        el.set("ContentType", content_type)
        self._overrides[part_name] = el
        self.dirty = True

    def drop_override(self, part: str) -> None:
        el = self._overrides.pop("/" + part, None)
        if el is not None:
            self._root.remove(el)
            self.dirty = True

    def serialize(self) -> bytes:
        return _serialize(self._root)


# ══════════════════════════════════════════════════════════════════════════════
# 关系表
# ══════════════════════════════════════════════════════════════════════════════
def _next_rel_id(rels_root) -> str:
    used = set()
    for el in rels_root:
        match = re.fullmatch(r"rId(\d+)", str(el.get("Id") or ""))
        if match:
            used.add(int(match.group(1)))
    index = 1
    while index in used:
        index += 1
    return f"rId{index}"


def _add_relationship(rels_root, rel_id: str, rel_type: str, target: str) -> None:
    el = etree.SubElement(rels_root, _pr("Relationship"))
    el.set("Id", rel_id)
    el.set("Type", rel_type)
    el.set("Target", target)


# ══════════════════════════════════════════════════════════════════════════════
# styles.xml
# ══════════════════════════════════════════════════════════════════════════════
class _Styles:
    """在 styles.xml 上追加 fill / font / xf，并按内容去重。"""

    def __init__(self, data: bytes):
        self._root = _parse(data)
        self.dirty = False
        self._fonts = self._section("fonts")
        self._fills = self._section("fills")
        self._cell_xfs = self._section("cellXfs")
        self._font_index = self._build_index(self._fonts)
        self._fill_index = self._build_index(self._fills)
        self._xf_index = self._build_index(self._cell_xfs)
        self._resolve_cache: dict[tuple, int] = {}

    # ── 只读查询 ──────────────────────────────────────────────────────────
    def has_fill(self, xf_index: int) -> bool:
        xf = self._xf(xf_index)
        if xf is None:
            return False
        fill = self._child_at(self._fills, int(xf.get("fillId") or 0))
        if fill is None:
            return False
        pattern = fill.find(_m("patternFill"))
        if pattern is None:
            return fill.find(_m("gradientFill")) is not None
        pattern_type = str(pattern.get("patternType") or "").strip().lower()
        return bool(pattern_type and pattern_type != "none")

    def font_size(self, xf_index: int) -> float:
        xf = self._xf(xf_index)
        if xf is None:
            return BASE_FONT_SIZE_PT
        font = self._child_at(self._fonts, int(xf.get("fontId") or 0))
        if font is None:
            return BASE_FONT_SIZE_PT
        size_el = font.find(_m("sz"))
        if size_el is None or not size_el.get("val"):
            return BASE_FONT_SIZE_PT
        try:
            return float(size_el.get("val"))
        except ValueError:
            return BASE_FONT_SIZE_PT

    # ── 写入 ──────────────────────────────────────────────────────────────
    def resolve(
        self,
        base_index: int,
        *,
        wrap_text: bool = False,
        fill_argb: str | None = None,
        font_color_argb: str | None = None,
        font_size: float | None = None,
    ) -> int:
        if not (wrap_text or fill_argb or font_color_argb or font_size is not None):
            return base_index

        cache_key = (base_index, wrap_text, fill_argb, font_color_argb, font_size)
        cached = self._resolve_cache.get(cache_key)
        if cached is not None:
            return cached

        base_xf = self._xf(base_index)
        if base_xf is None:
            base_index = 0
            base_xf = self._xf(0)
        if base_xf is None:  # 极端残缺的 styles.xml，放弃改样式
            return base_index

        xf = _deepcopy(base_xf)

        if fill_argb:
            xf.set("fillId", str(self._ensure_fill(fill_argb)))
            xf.set("applyFill", "1")

        if font_color_argb or font_size is not None:
            font = self._child_at(self._fonts, int(xf.get("fontId") or 0))
            new_font = _deepcopy(font) if font is not None else etree.Element(_m("font"))
            if font_color_argb:
                _set_font_child(new_font, "color", {"rgb": font_color_argb})
            if font_size is not None:
                _set_font_child(new_font, "sz", {"val": _format_number(font_size)})
            xf.set("fontId", str(self._ensure_font(new_font)))
            xf.set("applyFont", "1")

        if wrap_text:
            _set_wrap_alignment(xf)
            xf.set("applyAlignment", "1")

        index = self._ensure_xf(xf)
        self._resolve_cache[cache_key] = index
        return index

    def serialize(self) -> bytes:
        return _serialize(self._root)

    # ── 内部 ──────────────────────────────────────────────────────────────
    def _section(self, tag: str):
        el = self._root.find(_m(tag))
        if el is None:
            el = etree.SubElement(self._root, _m(tag))
        return el

    @staticmethod
    def _build_index(section) -> dict[bytes, int]:
        index: dict[bytes, int] = {}
        for position, child in enumerate(section):
            key = _canonical(child)
            index.setdefault(key, position)
        return index

    @staticmethod
    def _child_at(section, position: int):
        children = list(section)
        if 0 <= position < len(children):
            return children[position]
        return None

    def _xf(self, position: int):
        return self._child_at(self._cell_xfs, position)

    def _append(self, section, index: dict[bytes, int], element) -> int:
        key = _canonical(element)
        existing = index.get(key)
        if existing is not None:
            return existing
        section.append(element)
        position = len(section) - 1
        index[key] = position
        section.set("count", str(len(section)))
        self.dirty = True
        return position

    def _ensure_fill(self, argb: str) -> int:
        fill = etree.Element(_m("fill"))
        pattern = etree.SubElement(fill, _m("patternFill"))
        pattern.set("patternType", "solid")
        etree.SubElement(pattern, _m("fgColor")).set("rgb", argb)
        etree.SubElement(pattern, _m("bgColor")).set("rgb", argb)
        return self._append(self._fills, self._fill_index, fill)

    def _ensure_font(self, font) -> int:
        return self._append(self._fonts, self._font_index, font)

    def _ensure_xf(self, xf) -> int:
        return self._append(self._cell_xfs, self._xf_index, xf)


def _deepcopy(element):
    return copy.deepcopy(element)


def _canonical(element) -> bytes:
    return etree.tostring(element, method="c14n")


def _format_number(value: float) -> str:
    text = f"{float(value):g}"
    return text


def _set_font_child(font, tag: str, attrs: dict[str, str]) -> None:
    """CT_Font 是 xsd:choice，子元素顺序无所谓，就地替换或追加即可。"""
    el = font.find(_m(tag))
    if el is None:
        el = etree.SubElement(font, _m(tag))
    else:
        el.clear()
    for key, value in attrs.items():
        el.set(key, value)


_ALIGNMENT_KEPT_ATTRS = ("horizontal", "vertical", "textRotation", "indent", "shrinkToFit")


def _set_wrap_alignment(xf) -> None:
    """等价于旧路径的 ``Alignment(wrap_text=True, ...)``：只保留 5 个原属性。"""
    existing = xf.find(_m("alignment"))
    kept = {}
    if existing is not None:
        for name in _ALIGNMENT_KEPT_ATTRS:
            value = existing.get(name)
            if value is not None:
                kept[name] = value
        xf.remove(existing)
    alignment = etree.Element(_m("alignment"))
    for name, value in kept.items():
        alignment.set(name, value)
    alignment.set("wrapText", "1")
    xf.insert(0, alignment)  # CT_Xf 里 alignment 必须排在 protection 之前


# ══════════════════════════════════════════════════════════════════════════════
# 复核标记归一化
# ══════════════════════════════════════════════════════════════════════════════
def normalize_review_mark_colors(colors: dict[str, str] | None) -> dict[str, str]:
    raw_colors = dict(colors or {})
    return {
        mark: normalize_excel_rgb(raw_colors.get(mark, ""), fallback=default_color)
        for mark, default_color in _EXCEL_REVIEW_MARK_COLORS.items()
    }


def normalize_review_marks(
    review_marks: dict[str, str] | None,
    review_mark_colors: dict[str, str],
) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for source, mark in (review_marks or {}).items():
        source_key = str(source or "").strip()
        mark_key = str(mark or "").strip()
        if source_key and mark_key in review_mark_colors:
            normalized[source_key] = mark_key
    return normalized


def normalize_existing_fill_policy(policy: str) -> str:
    normalized = str(policy or "").strip()
    if normalized not in {"skip", "overwrite", "red_font"}:
        return EXCEL_REVIEW_EXISTING_FILL_POLICY_DEFAULT
    return normalized


def normalize_excel_rgb(value: str, *, fallback: str) -> str:
    cleaned = re.sub(r"[^0-9A-Fa-f]", "", str(value or "")).upper()
    if len(cleaned) == 8:
        return cleaned[-6:]
    if len(cleaned) == 6:
        return cleaned
    return fallback


def to_excel_argb(value: str) -> str:
    cleaned = normalize_excel_rgb(value, fallback="")
    return f"FF{cleaned}" if cleaned else "FFFF0000"


# ══════════════════════════════════════════════════════════════════════════════
# 排版估算（与旧 openpyxl 路径共用同一套公式）
# ══════════════════════════════════════════════════════════════════════════════
def estimate_chars_per_line(col_width: float, font_size_pt: float) -> int:
    """按列宽+字号估算每行容纳字符数。"""
    scale = 11.0 / max(font_size_pt, 0.1)
    return max(1, int(col_width * CHARS_PER_WIDTH * scale))


def estimate_required_lines(text: str, chars_per_line: int) -> int:
    total_lines = 0
    for segment in str(text).split("\n"):
        total_lines += max(1, -(-len(segment) // max(chars_per_line, 1)))
    return max(1, total_lines)


def estimate_max_visible_lines(row_height: float, font_size_pt: float) -> int:
    line_height = max(1.0, font_size_pt * PRINT_GUARD_LINE_HEIGHT_MULTIPLIER)
    return max(1, int(row_height / line_height))


# ══════════════════════════════════════════════════════════════════════════════
# 分表几何：行高 / 列宽
# ══════════════════════════════════════════════════════════════════════════════
class _SheetGeometry:
    """从 worksheet XML 直接读行高列宽，供行高估算与图片尺寸换算共用。"""

    def __init__(self, root):
        # absoluteAnchor 反查坐标要用到整表范围，但那是少数情况，按需算，见 extent()。
        self._root = root
        self._extent: tuple[int, int] | None = None
        self._extent_ready = False

        fmt = root.find(_m("sheetFormatPr"))
        self.default_row_height = _to_float(
            fmt.get("defaultRowHeight") if fmt is not None else None,
            DEFAULT_ROW_HEIGHT,
        )
        self.default_col_width = _to_float(
            fmt.get("defaultColWidth") if fmt is not None else None,
            DEFAULT_COL_WIDTH,
        )
        self._col_widths: dict[int, float] = {}
        self._hidden_cols: set[int] = set()
        cols = root.find(_m("cols"))
        if cols is not None:
            for col in cols.findall(_m("col")):
                try:
                    first = int(col.get("min") or 1)
                    last = int(col.get("max") or first)
                except ValueError:
                    continue
                width = _to_float(col.get("width"), None)
                hidden = str(col.get("hidden") or "") in {"1", "true"}
                # 跨度过大的 <col>（常见的 min=1 max=16384）只展开到有意义的范围。
                last = min(last, first + 1024)
                for index in range(first, last + 1):
                    if width is not None:
                        self._col_widths[index] = width
                    if hidden:
                        self._hidden_cols.add(index)

        self._row_heights: dict[int, float] = {}
        self._hidden_rows: set[int] = set()
        for row_num, row in _iter_rows(root):
            height = _to_float(row.get("ht"), None)
            if height is not None:
                self._row_heights[row_num] = height
            if str(row.get("hidden") or "") in {"1", "true"}:
                self._hidden_rows.add(row_num)

    def col_width(self, col_index: int) -> float:
        return self._col_widths.get(col_index, self.default_col_width)

    def row_height(self, row_num: int) -> float:
        return self._row_heights.get(row_num, self.default_row_height)

    def last_sized_col(self) -> int:
        """最后一个尺寸被显式定义过的列号（1-based）；一个都没有返回 0。

        绝对坐标反查要靠它划定「必须逐格累加」的范围。不能拿 ``<dimension>``
        代替：``<dimension>`` 只覆盖有内容的单元格，设了列宽并不会把它撑大，
        而「把数据区右边几列拉宽好摆一张图」恰恰是最常见的写法。隐藏列也算
        显式定义——它的 EMU 是 0，同样不能拿默认列宽顶替。
        """
        return max((*self._col_widths, *self._hidden_cols), default=0)

    def last_sized_row(self) -> int:
        """最后一个尺寸被显式定义过的行号（1-based）；一个都没有返回 0。

        理由同 :meth:`last_sized_col`：为了放图片把数据区下面几行拉高，行高
        就落在 ``<dimension>`` 之外了。
        """
        return max((*self._row_heights, *self._hidden_rows), default=0)

    def col_emu(self, col_index: int) -> int:
        if col_index in self._hidden_cols:
            return 0
        width = self.col_width(col_index)
        pixels = int((256 * width + int(128 / MAX_DIGIT_WIDTH)) / 256 * MAX_DIGIT_WIDTH)
        return pixels * EMU_PER_PIXEL

    def row_emu(self, row_num: int) -> int:
        if row_num in self._hidden_rows:
            return 0
        return int(round(self.row_height(row_num) * EMU_PER_POINT))

    def default_col_emu(self) -> int:
        """没有 ``<col>`` 定义的列，Excel 按这个宽度渲染。"""
        width = self.default_col_width
        pixels = int((256 * width + int(128 / MAX_DIGIT_WIDTH)) / 256 * MAX_DIGIT_WIDTH)
        return pixels * EMU_PER_PIXEL

    def default_row_emu(self) -> int:
        """没有 ``<row>`` 定义的行，Excel 按这个高度渲染。"""
        return int(round(self.default_row_height * EMU_PER_POINT))

    def extent(self) -> tuple[int, int] | None:
        """整表的 ``(最大行号, 最大列号)``（都是 1-based）；拿不到返回 ``None``。

        优先读 ``<dimension ref>``（Excel / WPS / openpyxl 都会写），缺失或写坏了
        才退回扫描 sheetData。只有 absoluteAnchor 换算需要它，所以延迟计算并缓存，
        避免给没有绝对锚点的分表白白加一趟全表遍历。
        """
        if self._extent_ready:
            return self._extent
        self._extent_ready = True
        self._extent = _compute_sheet_extent(self._root)
        return self._extent


def _compute_sheet_extent(root) -> tuple[int, int] | None:
    max_row = 0
    max_col = 0

    node = root.find(_m("dimension"))
    ref = str(node.get("ref") or "") if node is not None else ""
    if ref:
        for corner in ref.split(":"):
            coordinate = _split_coordinate(corner.strip())
            if coordinate is None:
                max_row = max_col = 0
                break
            row_num, col_index = coordinate
            max_row = max(max_row, row_num)
            max_col = max(max_col, col_index)
    if max_row > 0 and max_col > 0:
        return max_row, max_col

    for row_num, row in _iter_rows(root):
        for _, col_index, _cell in _iter_cells(row, row_num):
            max_col = max(max_col, col_index)
            max_row = max(max_row, row_num)
    if max_row > 0 and max_col > 0:
        return max_row, max_col
    return None


def _to_float(value, fallback):
    if value is None or value == "":
        return fallback
    try:
        return float(value)
    except (TypeError, ValueError):
        return fallback


def _iter_rows(root):
    """产出 ``(行号, <row> 元素)``；缺 r 属性时按 openpyxl 的规则顺延。"""
    sheet_data = root.find(_m("sheetData"))
    if sheet_data is None:
        return
    counter = 0
    for row in sheet_data:
        if _local(row) != "row":
            continue
        raw = row.get("r")
        if raw:
            try:
                counter = int(float(raw))
            except ValueError:
                counter += 1
        else:
            counter += 1
        yield counter, row


def _iter_cells(row, row_num: int):
    """产出 ``(行号, 列号, <c> 元素)``。"""
    col_counter = 0
    for cell in row:
        if _local(cell) != "c":
            continue
        coordinate = _split_coordinate(cell.get("r") or "")
        if coordinate is not None:
            col_counter = coordinate[1]
        else:
            col_counter += 1
        yield row_num, col_counter, cell


# ══════════════════════════════════════════════════════════════════════════════
# 单元格取值
# ══════════════════════════════════════════════════════════════════════════════
def _text_of_string_item(element) -> str:
    """``<si>`` / ``<is>`` 的纯文本（拼接 t 与 r/t，忽略注音 rPh）。"""
    if element is None:
        return ""
    parts: list[str] = []
    for child in element:
        tag = _local(child)
        if tag == "t":
            parts.append(child.text or "")
        elif tag == "r":
            for run_child in child:
                if _local(run_child) == "t":
                    parts.append(run_child.text or "")
    return "".join(parts)


def _load_shared_strings(package: _Package, part: str) -> list[str]:
    data = package.read(part)
    if data is None:
        return []
    root = _parse(data)
    return [_text_of_string_item(si) for si in root if _local(si) == "si"]


def _cell_formula(cell):
    return cell.find(_m("f"))


def _is_dispimg_formula(formula_el) -> bool:
    if formula_el is None:
        return False
    text = str(formula_el.text or "").lstrip()
    return text.upper().startswith(("DISPIMG(", "_XLFN.DISPIMG("))


def _cell_v_text(cell) -> str | None:
    v = cell.find(_m("v"))
    if v is None:
        return None
    return v.text or ""


def _cell_display_text(cell, shared_strings: list[str]) -> str | None:
    """公式单元格的缓存显示值（等价于 openpyxl ``data_only=True``）。"""
    raw = _cell_v_text(cell)
    if raw is None:
        return None
    cell_type = cell.get("t", "n")
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return None
    if cell_type in {"str", "e"}:
        return raw
    if cell_type == "inlineStr":
        return _text_of_string_item(cell.find(_m("is")))
    return None  # 数字 / 布尔 / 日期在 openpyxl 里不是 str


def _cell_static_text(cell, shared_strings: list[str]) -> str | None:
    """非公式单元格在 openpyxl 里的字符串值；非字符串返回 None。"""
    cell_type = cell.get("t", "n")
    if cell_type == "inlineStr":
        return _text_of_string_item(cell.find(_m("is")))
    raw = _cell_v_text(cell)
    if raw is None:
        return None
    if cell_type == "s":
        try:
            return shared_strings[int(raw)]
        except (ValueError, IndexError):
            return None
    if cell_type in {"str", "e"}:
        return raw
    return None


# ══════════════════════════════════════════════════════════════════════════════
# 写入单元格文本
# ══════════════════════════════════════════════════════════════════════════════
def _set_cell_inline_text(cell, text: str) -> bool:
    """把单元格改写成 inlineStr。返回是否删掉了公式。"""
    removed_formula = False
    for child in list(cell):
        tag = _local(child)
        if tag in {"v", "is", "f"}:
            if tag == "f":
                removed_formula = True
            cell.remove(child)
    cell.set("t", "inlineStr")
    inline = etree.SubElement(cell, _m("is"))
    text_el = etree.SubElement(inline, _m("t"))
    text_el.set(XML_SPACE, "preserve")
    text_el.text = text
    return removed_formula


def _promote_shared_formula(root, formula_el) -> bool:
    """把要被删除的共享公式主单元格的主控权让给组里下一个单元格。

    返回 True 表示已安全处理（无依赖或已成功让渡），False 表示应放弃改写该单元格。
    """
    if str(formula_el.get("t") or "") != "shared":
        return True
    share_id = formula_el.get("si")
    master_ref = formula_el.get("ref")
    if share_id is None or master_ref is None:
        return True  # 本身就是从属单元格，删掉不影响别人

    master_cell = formula_el.getparent()
    master_coord = master_cell.get("r")
    dependents: list[tuple[int, int, object]] = []
    for row_num, row in _iter_rows(root):
        for _, col_index, cell in _iter_cells(row, row_num):
            if cell is master_cell:
                continue
            other = cell.find(_m("f"))
            if other is None or other.get("si") != share_id:
                continue
            dependents.append((row_num, col_index, other))
    if not dependents:
        return True

    if not master_coord or not (formula_el.text or "").strip():
        return False

    try:
        from openpyxl.formula.translate import Translator
    except ImportError:  # pragma: no cover - openpyxl 是硬依赖
        return False

    new_master_row, new_master_col, new_master_f = dependents[0]
    new_master_coord = f"{_column_letter(new_master_col)}{new_master_row}"
    try:
        translated = Translator(
            f"={formula_el.text}", origin=master_coord
        ).translate_formula(new_master_coord)
    except Exception as error:  # noqa: BLE001 - 公式无法平移时放弃改写更安全
        logger.warning(f"共享公式主控权让渡失败（{master_coord}）：{error}")
        return False

    min_row = min(item[0] for item in dependents)
    max_row = max(item[0] for item in dependents)
    min_col = min(item[1] for item in dependents)
    max_col = max(item[1] for item in dependents)
    new_ref = (
        f"{_column_letter(min_col)}{min_row}:{_column_letter(max_col)}{max_row}"
    )

    new_master_f.text = translated.lstrip("=")
    new_master_f.set("t", "shared")
    new_master_f.set("si", share_id)
    new_master_f.set("ref", new_ref)
    return True


# ══════════════════════════════════════════════════════════════════════════════
# drawing：行高变化后固定悬浮图片尺寸
# ══════════════════════════════════════════════════════════════════════════════
def _anchor_position(anchor, tag: str) -> tuple[int, int, int, int] | None:
    node = anchor.find(_xdr(tag))
    if node is None:
        return None
    try:
        col = int(node.findtext(_xdr("col")) or 0)
        col_off = int(node.findtext(_xdr("colOff")) or 0)
        row = int(node.findtext(_xdr("row")) or 0)
        row_off = int(node.findtext(_xdr("rowOff")) or 0)
    except (TypeError, ValueError):
        return None
    return col, col_off, row, row_off


def _anchor_absolute_pos(anchor) -> tuple[int, int] | None:
    """``<xdr:pos x= y=>`` 的绝对 EMU 坐标。"""
    node = anchor.find(_xdr("pos"))
    if node is None:
        return None
    try:
        return int(node.get("x") or 0), int(node.get("y") or 0)
    except (TypeError, ValueError):
        return None


def _locate_emu(
    value: int, span, limit: int, default_size: int, max_index: int
) -> tuple[int, int] | None:
    """把一维绝对 EMU 坐标换算成 ``(0-based 行/列下标, 格内偏移)``。

    ``span(index)`` 取 1-based 下标那一行/列的 EMU 尺寸，``limit`` 是逐格累加的
    上界（1-based，含）。调用方必须把 ``limit`` 取到「``<dimension>`` 边界」和
    「最后一个有显式尺寸定义的行/列」两者的较大值：显式行高/列宽完全可以落在
    ``<dimension>`` 之外——为了摆一张图把数据区下面几行拉高、右边几列拉宽是最
    普通的写法，而 ``<dimension>`` 只覆盖有内容的单元格，不会跟着变大。只累加
    到 ``<dimension>`` 就会把那些真实尺寸当成默认尺寸，图片被算到别的格子里。

    坐标落在 ``limit`` 之外也不能当成没依据——浮在数据区下方/右侧的图片（页脚
    logo、说明图）并不罕见，Excel 自己渲染时这些没有显式定义的行/列一律按
    ``sheetFormatPr`` 的默认尺寸算，我们照做即可：越过 ``limit`` 之后每格尺寸
    都相同（``default_size``），直接整除/取余闭式定位，不再逐格累加——坐标可能
    来自任意远的绝对定位，累加到 Excel 行列号上限（百万级）会是几十万次循环。

    ``default_size`` 为 0 或负数（异常工作簿，量不出默认尺寸）时不能拿它做除
    数，只能放弃外推。定位结果越过 ``max_index``（Excel 本身的行/列号上限，
    0-based）说明坐标本身是坏数据，一并放弃。这两种情况、以及 ``value`` 为负
    时，都返回 ``None``——交还调用方保持原样，好过瞎猜一个位置。
    """
    if value < 0:
        return None
    # <dimension ref> 是外部数据，行号位数不设限（畸形的 "A1:A99999999" 照收），
    # 循环内那条路径会直接把下标交出去，越界检查只拦得住外推那条。先夹住上界，
    # 两条路径就都不可能吐出 Excel 打不开的行列号。
    limit = min(limit, max_index + 1)
    cursor = 0
    for index in range(1, limit + 1):
        size = span(index)
        # 隐藏行/列尺寸为 0，直接跳过，坐标只会落在可见的那一格里。
        if size > 0 and cursor + size > value:
            return index - 1, value - cursor
        cursor += size

    if default_size <= 0:
        return None
    steps, offset = divmod(value - cursor, default_size)
    # limit 之内的下标用的是 index - 1；紧接着 limit 之后第一格的 0-based 下标
    # 正好是 limit 本身，所以外推起点直接从 limit 开始加 steps。
    located_index = limit + steps
    if located_index > max_index:
        return None
    return located_index, offset


def fix_drawing_anchors(
    drawing_root,
    geometry: _SheetGeometry,
    changed_rows: set[int],
    *,
    freeze_all: bool = False,
) -> int:
    """把受行高变化影响的悬浮图片锚点改写成定尺寸 / 定位置的 oneCellAnchor。

    ``geometry`` 必须是**调整行高之前**的几何，这样算出的 ext 才是原渲染尺寸。

    两类锚点各有各的病：

    * ``twoCellAnchor`` 左上右下各钉一格，行一变高就被两个角拽着拉伸变形；
    * ``absoluteAnchor`` 钉的是相对整表原点的绝对坐标，上方行变高后格子整体下移，
      图片却纹丝不动，最后盖到不相干的内容上。

    两者都改写成 ``oneCellAnchor``（只钉左上角 + 写死尺寸），图片跟着所在单元格
    走，既不拉伸也不错位。

    ``freeze_all=True`` 表示调用方之后还要交给 Excel COM 做整表 autofit——那一刀
    会重排 used range 里的**所有**行高，我们无从预知哪些行会变，只能把整张表的悬浮
    图片全部冻结。代价是没人动过的行上的图片从此也不随行高缩放，这是预期语义。

    返回被冻结/改写的锚点数量。
    """
    if not changed_rows and not freeze_all:
        return 0

    frozen = 0
    for anchor in list(drawing_root):
        tag = _local(anchor)
        if tag == "twoCellAnchor":
            frozen += _freeze_two_cell_anchor(
                anchor, geometry, changed_rows, freeze_all=freeze_all
            )
        elif tag == "absoluteAnchor":
            frozen += _freeze_absolute_anchor(
                anchor, geometry, changed_rows, freeze_all=freeze_all
            )
    return frozen


def _freeze_two_cell_anchor(
    anchor,
    geometry: _SheetGeometry,
    changed_rows: set[int],
    *,
    freeze_all: bool,
) -> int:
    start = _anchor_position(anchor, "from")
    end = _anchor_position(anchor, "to")
    if start is None or end is None:
        return 0
    from_col, from_col_off, from_row, from_row_off = start
    to_col, to_col_off, to_row, to_row_off = end
    # drawing 里的行列是 0-based；图片高度只取决于 [from_row, to_row) 这些行。
    if not freeze_all and not any(
        (row + 1) in changed_rows for row in range(from_row, to_row)
    ):
        return 0

    cx = sum(geometry.col_emu(col + 1) for col in range(from_col, to_col))
    cx += to_col_off - from_col_off
    cy = sum(geometry.row_emu(row + 1) for row in range(from_row, to_row))
    cy += to_row_off - from_row_off
    cx = max(int(cx), 1)
    cy = max(int(cy), 1)

    one_cell = etree.Element(_xdr("oneCellAnchor"))
    from_node = anchor.find(_xdr("from"))
    to_node = anchor.find(_xdr("to"))
    one_cell.append(from_node)
    ext = etree.SubElement(one_cell, _xdr("ext"))
    ext.set("cx", str(cx))
    ext.set("cy", str(cy))
    for child in list(anchor):
        if child is to_node:
            continue
        one_cell.append(child)

    _sync_shape_extent(one_cell, cx, cy)

    parent = anchor.getparent()
    parent.replace(anchor, one_cell)
    return 1


def _freeze_absolute_anchor(
    anchor,
    geometry: _SheetGeometry,
    changed_rows: set[int],
    *,
    freeze_all: bool,
) -> int:
    """把 absoluteAnchor 按**调整前**的几何反查成 oneCellAnchor，尺寸原样保留。"""
    position = _anchor_absolute_pos(anchor)
    ext_node = anchor.find(_xdr("ext"))
    if position is None or ext_node is None:
        return 0
    extent = geometry.extent()
    if extent is None:
        return 0  # 几何信息不全，不瞎猜
    max_row, max_col = extent

    # 逐格累加的范围要盖住最后一个显式定义过尺寸的行/列，而不是止步于
    # <dimension>——两者谁大用谁，理由见 _SheetGeometry.last_sized_row()。
    row_limit = max(max_row, geometry.last_sized_row())
    col_limit = max(max_col, geometry.last_sized_col())

    x_emu, y_emu = position
    located_col = _locate_emu(
        x_emu,
        geometry.col_emu,
        col_limit,
        geometry.default_col_emu(),
        EXCEL_MAX_COL_INDEX,
    )
    located_row = _locate_emu(
        y_emu,
        geometry.row_emu,
        row_limit,
        geometry.default_row_emu(),
        EXCEL_MAX_ROW_INDEX,
    )
    if located_col is None or located_row is None:
        return 0  # 坐标是负数，或外推后仍越过 Excel 行列号上限，坏数据，保持原样
    col_index, col_off = located_col
    row_index, row_off = located_row

    # 只有当**上方**有行变高时绝对坐标才会漂移；同一行内变高不影响该行的顶边。
    if not freeze_all and not any(row < row_index + 1 for row in changed_rows):
        return 0

    one_cell = etree.Element(_xdr("oneCellAnchor"))
    from_node = etree.SubElement(one_cell, _xdr("from"))
    for tag, value in (
        ("col", col_index),
        ("colOff", col_off),
        ("row", row_index),
        ("rowOff", row_off),
    ):
        etree.SubElement(from_node, _xdr(tag)).text = str(value)
    pos_node = anchor.find(_xdr("pos"))
    for child in list(anchor):
        if child is pos_node:
            continue
        one_cell.append(child)  # ext 原样搬过来，尺寸不动

    parent = anchor.getparent()
    parent.replace(anchor, one_cell)
    return 1


def _sync_shape_extent(anchor, cx: int, cy: int) -> None:
    """oneCellAnchor 下的形状 xfrm/ext 也要跟着改，避免渲染器读到 0 尺寸。"""
    for ext in anchor.iter(f"{{{NS_A}}}ext"):
        parent = ext.getparent()
        if parent is None or etree.QName(parent).localname != "xfrm":
            continue
        ext.set("cx", str(cx))
        ext.set("cy", str(cy))


# ══════════════════════════════════════════════════════════════════════════════
# 分表处理
# ══════════════════════════════════════════════════════════════════════════════
@dataclass
class _SheetEntry:
    name: str
    part: str


@dataclass
class _SheetSnapshot:
    name: str
    part: str
    sheet_xml: bytes
    rels_xml: bytes | None
    owned_parts: dict[str, bytes] = field(default_factory=dict)
    owned_rels: dict[str, bytes] = field(default_factory=dict)


@dataclass
class _SheetOutcome:
    review_marked: int = 0
    review_skipped: int = 0
    shrunk_cells: int = 0
    mutated_cells: int = 0
    removed_formula: bool = False
    anchor_frozen_count: int = 0


def _process_sheet(
    package: _Package,
    entry: _SheetEntry,
    *,
    shared_strings: list[str],
    styles: _Styles,
    translations: dict[str, str],
    target_lang: str,
    source_lang: str,
    formula_display_value_backfill: bool,
    lock_row_height: bool,
    review_enabled: bool,
    review_mark_map: dict[str, str],
    review_color_map: dict[str, str],
    fill_policy: str,
    review_positions: list[dict[str, str]] | None,
    log_callback,
    allowed_coordinates: set[str] | None = None,
    external_autofit_planned: bool = False,
) -> _SheetOutcome:
    data = package.read(entry.part)
    if data is None:
        return _SheetOutcome()

    root = _parse(data)
    geometry = _SheetGeometry(root)
    outcome = _SheetOutcome()
    dirty = False

    # 行高估算需要整表的最终文本，未改动的单元格也要算进去。
    row_texts: dict[int, list[tuple[int, str]]] = {}

    for row_num, row in _iter_rows(root):
        collected: list[tuple[int, str]] = []
        for _, col_index, cell in _iter_cells(row, row_num):
            formula_el = _cell_formula(cell)

            # WPS 嵌入图片公式：整格不动，也不参与行高估算。
            if _is_dispimg_formula(formula_el):
                continue
            # 富数据（「置于单元格内」的图片等）单元格：整格不动。
            if cell.get("vm") is not None:
                text = (
                    _cell_display_text(cell, shared_strings)
                    if formula_el is not None
                    else _cell_static_text(cell, shared_strings)
                )
                if text:
                    collected.append((col_index, text))
                continue

            if formula_el is not None:
                current_text = "=" + (formula_el.text or "")
            else:
                current_text = _cell_static_text(cell, shared_strings)

            final_text = current_text

            source_text = _resolve_source_text(
                cell,
                formula_el,
                shared_strings,
                formula_display_value_backfill,
            )
            position_allowed = (
                allowed_coordinates is None
                or f"{_column_letter(col_index)}{row_num}" in allowed_coordinates
            )
            mutation = (
                _plan_cell_mutation(
                    source_text,
                    translations=translations,
                    target_lang=target_lang,
                    source_lang=source_lang,
                    review_enabled=review_enabled,
                    review_mark_map=review_mark_map,
                )
                if position_allowed
                else None
            )

            if mutation is not None:
                new_text, mark_kind = mutation
                base_index = _cell_style_index(cell)
                fill_argb: str | None = None
                font_color_argb: str | None = None
                font_size: float | None = None

                if mark_kind:
                    applied, fill_argb, font_color_argb = _plan_review_mark(
                        styles,
                        base_index,
                        mark_kind,
                        review_color_map=review_color_map,
                        fill_policy=fill_policy,
                    )
                    _record_review_position(
                        review_positions,
                        entry.name,
                        f"{_column_letter(col_index)}{row_num}",
                        mark_kind,
                        applied,
                        bool(font_color_argb),
                    )
                    if applied:
                        outcome.review_marked += 1
                    else:
                        outcome.review_skipped += 1

                if new_text is not None:
                    if formula_el is not None and not _promote_shared_formula(root, formula_el):
                        new_text = None

                if new_text is not None:
                    if _set_cell_inline_text(cell, new_text):
                        outcome.removed_formula = True
                    final_text = new_text
                    dirty = True
                    outcome.mutated_cells += 1

                    if lock_row_height:
                        size, reached_floor = _shrink_font_for_locked_row(
                            new_text,
                            col_width=geometry.col_width(col_index),
                            row_height=geometry.row_height(row_num),
                            font_size=styles.font_size(base_index),
                        )
                        if size is not None:
                            font_size = size
                        outcome.shrunk_cells += 1
                        if reached_floor and log_callback:
                            log_callback(
                                f"[WARN] {entry.name}!{_column_letter(col_index)}{row_num} "
                                f"缩至最小字号 {PRINT_GUARD_FONT_FLOOR:.1f}pt 仍可能无法完全显示"
                            )

                    new_index = styles.resolve(
                        base_index,
                        wrap_text=True,
                        fill_argb=fill_argb,
                        font_color_argb=font_color_argb,
                        font_size=font_size,
                    )
                    _set_cell_style_index(cell, new_index)
                elif fill_argb or font_color_argb:
                    new_index = styles.resolve(
                        base_index,
                        fill_argb=fill_argb,
                        font_color_argb=font_color_argb,
                    )
                    _set_cell_style_index(cell, new_index)
                    dirty = True

            if final_text:
                collected.append((col_index, final_text))

        if collected:
            row_texts[row_num] = collected

    changed_rows: set[int] = set()
    if not lock_row_height:
        changed_rows, touched_rows = _auto_adjust_row_heights(root, geometry, row_texts)
        dirty = dirty or touched_rows

    # 锁定行高时行高压根不变，没有任何锚点需要动。
    # 否则：调用方若还要跑 Excel COM 的整表 autofit，我们预知不了它会改哪些行，
    # 只能整表冻结；不跑 autofit 时仍然只冻结自己改过的那些行。
    freeze_all = external_autofit_planned and not lock_row_height
    if changed_rows or freeze_all:
        outcome.anchor_frozen_count = _patch_sheet_drawing(
            package, entry, root, geometry, changed_rows, freeze_all=freeze_all
        )

    if dirty:
        package.write(entry.part, _serialize(root))

    return outcome


def _resolve_source_text(
    cell,
    formula_el,
    shared_strings: list[str],
    formula_display_value_backfill: bool,
) -> str | None:
    if formula_el is None:
        return _cell_static_text(cell, shared_strings)
    if not formula_display_value_backfill:
        return "=" + (formula_el.text or "")
    return _cell_display_text(cell, shared_strings)


def _plan_cell_mutation(
    source_text: str | None,
    *,
    translations: dict[str, str],
    target_lang: str,
    source_lang: str,
    review_enabled: bool,
    review_mark_map: dict[str, str],
) -> tuple[str | None, str | None] | None:
    """复刻旧写入路径的判定顺序。返回 ``(新文本或 None, 标记类型或 None)``。"""
    if source_text is None:
        return None
    if not should_translate(source_text, target_lang=target_lang, source_lang=source_lang):
        return None

    source_key = source_text.strip()
    translated = translations.get(source_key)
    if translated is None:
        return None

    mark_kind = review_mark_map.get(source_key)

    if is_replace_translation(translated):
        final_text = extract_replace_translation(translated).strip()
        if not final_text:
            return (None, mark_kind) if mark_kind else None
        return final_text, mark_kind

    if not translated:
        return (None, mark_kind) if mark_kind else None

    retained_original = source_key.lower() == translated.strip().lower()
    if review_enabled and retained_original and not mark_kind:
        mark_kind = MIXED_MARK_UNRESOLVED
    if retained_original:
        return (None, mark_kind) if mark_kind else None

    return source_text + BILINGUAL_SEPARATOR + translated, mark_kind


def _plan_review_mark(
    styles: _Styles,
    base_index: int,
    mark_kind: str,
    *,
    review_color_map: dict[str, str],
    fill_policy: str,
) -> tuple[bool, str | None, str | None]:
    """返回 ``(是否算标记成功, 填充色 ARGB, 字色 ARGB)``。"""
    fill_color = review_color_map.get(mark_kind)
    if not fill_color:
        return False, None, None

    has_existing_fill = styles.has_fill(base_index)
    if has_existing_fill and fill_policy == "skip":
        return False, None, None
    if has_existing_fill and fill_policy == "red_font":
        return True, None, to_excel_argb(_EXCEL_REVIEW_RISK_FONT_COLOR)
    return True, to_excel_argb(fill_color), None


def _record_review_position(
    collector: list[dict[str, str]] | None,
    worksheet: str,
    coordinate: str,
    category: str,
    applied: bool,
    red_font: bool,
) -> None:
    if collector is None:
        return
    if not applied:
        action = "preserved_existing_fill"
    elif red_font:
        action = "marked_red_font"
    else:
        action = "marked_fill"
    collector.append(
        {
            "worksheet": worksheet,
            "cell": coordinate,
            "category": category,
            "action": action,
        }
    )


def _cell_style_index(cell) -> int:
    try:
        return int(cell.get("s") or 0)
    except ValueError:
        return 0


def _set_cell_style_index(cell, index: int) -> None:
    if index:
        cell.set("s", str(index))
    else:
        cell.attrib.pop("s", None)


def _shrink_font_for_locked_row(
    text: str,
    *,
    col_width: float,
    row_height: float,
    font_size: float,
) -> tuple[float | None, bool]:
    """锁定行高模式：逐步缩字号。返回 ``(新字号或 None, 是否触底且仍装不下)``。"""
    current_size = float(font_size or BASE_FONT_SIZE_PT)
    min_size = float(PRINT_GUARD_FONT_FLOOR)
    step = float(PRINT_GUARD_FONT_STEP)
    original_size = current_size

    while True:
        chars_per_line = estimate_chars_per_line(col_width, current_size)
        required_lines = estimate_required_lines(text, chars_per_line)
        visible_lines = estimate_max_visible_lines(row_height, current_size)
        if required_lines <= visible_lines:
            break
        if current_size <= min_size:
            current_size = min_size
            break
        current_size = max(min_size, round(current_size - step, 2))

    reached_floor = current_size <= min_size and estimate_required_lines(
        text, estimate_chars_per_line(col_width, current_size)
    ) > estimate_max_visible_lines(row_height, current_size)

    new_size = None if current_size == original_size else current_size
    return new_size, reached_floor


def _auto_adjust_row_heights(
    root,
    geometry: _SheetGeometry,
    row_texts: dict[int, list[tuple[int, str]]],
) -> tuple[set[int], bool]:
    """自动调整行高以适配双语内容。

    返回 ``(高度真的变了的行号集合, 是否写过行属性)``。前者用于判断哪些悬浮图片
    需要固定尺寸；后者只是用来决定要不要重写分表 XML。
    """
    changed: set[int] = set()
    touched = False
    for row_num, row in _iter_rows(root):
        max_lines = 1
        for col_index, text in row_texts.get(row_num, ()):
            chars_per_line = max(1, int(geometry.col_width(col_index) * CHARS_PER_WIDTH))
            lines = 0
            for segment in text.split("\n"):
                lines += max(1, -(-len(segment) // chars_per_line))
            max_lines = max(max_lines, lines)

        if max_lines <= 1:
            continue

        new_height = max_lines * BASE_FONT_SIZE_PT * LINE_HEIGHT_RATIO
        old_height = _to_float(row.get("ht"), None)
        row.set("ht", _format_number(new_height))
        row.set("customHeight", "1")
        touched = True
        if old_height is None or abs(old_height - new_height) > 1e-9:
            changed.add(row_num)
    return changed, touched


def _patch_sheet_drawing(
    package: _Package,
    entry: _SheetEntry,
    root,
    geometry: _SheetGeometry,
    changed_rows: set[int],
    *,
    freeze_all: bool = False,
) -> int:
    """返回本分表被冻结/改写的悬浮图片锚点数量。"""
    drawing_el = root.find(_m("drawing"))
    if drawing_el is None:
        return 0
    rel_id = drawing_el.get(_r("id"))
    if not rel_id:
        return 0

    rels_data = package.read(_rels_path(entry.part))
    if rels_data is None:
        return 0
    rels_root = _parse(rels_data)
    target = None
    for rel in rels_root:
        if rel.get("Id") == rel_id:
            target = rel.get("Target")
            break
    if not target:
        return 0

    drawing_part = _resolve_part(entry.part, target)
    drawing_data = package.read(drawing_part)
    if drawing_data is None:
        return 0

    drawing_root = _parse(drawing_data)
    frozen = fix_drawing_anchors(
        drawing_root, geometry, changed_rows, freeze_all=freeze_all
    )
    if frozen:
        package.write(drawing_part, _serialize(drawing_root))
    return frozen


# ══════════════════════════════════════════════════════════════════════════════
# 克隆原文分表
# ══════════════════════════════════════════════════════════════════════════════
def _snapshot_sheet(package: _Package, entry: _SheetEntry) -> _SheetSnapshot:
    """在任何改写之前留存分表及其独占部件的原始字节。"""
    sheet_xml = package.read(entry.part) or b""
    rels_path = _rels_path(entry.part)
    rels_xml = package.read(rels_path)
    snapshot = _SheetSnapshot(
        name=entry.name,
        part=entry.part,
        sheet_xml=sheet_xml,
        rels_xml=rels_xml,
    )
    if rels_xml is None:
        return snapshot

    rels_root = _parse(rels_xml)
    for rel in rels_root:
        if rel.get("Type") not in REL_TYPES_SHEET_OWNED:
            continue
        if str(rel.get("TargetMode") or "") == "External":
            continue
        part = _resolve_part(entry.part, rel.get("Target") or "")
        data = package.read(part)
        if data is None:
            continue
        snapshot.owned_parts[part] = data
        owned_rels = package.read(_rels_path(part))
        if owned_rels is not None:
            snapshot.owned_rels[part] = owned_rels
    return snapshot


def _strip_table_parts(sheet_xml: bytes) -> bytes:
    root = _parse(sheet_xml)
    table_parts = root.find(_m("tableParts"))
    if table_parts is None:
        return sheet_xml
    root.remove(table_parts)
    return _serialize(root)


def _clone_original_sheets(
    package: _Package,
    content_types: _ContentTypes,
    workbook_root,
    workbook_part: str,
    workbook_rels_root,
    snapshots: list[_SheetSnapshot],
    *,
    freeze_all_anchors: bool = False,
) -> int:
    """克隆原文分表；返回克隆件里被冻结的悬浮图片锚点数量。

    ``freeze_all_anchors``：克隆分表的行高虽然是原样照抄的，但之后 Excel COM 的
    autofit 是对**整本工作簿的每张表**跑的，克隆件也会被重排行高、把图片拉变形。
    所以这里也要冻结。克隆件用的是快照自己的几何——它就是「调整前」的几何。
    """
    sheets_el = workbook_root.find(_m("sheets"))
    if sheets_el is None:
        return 0
    frozen_total = 0

    used_names = {
        str(el.get("name") or "")
        for el in sheets_el
        if _local(el) == "sheet"
    }
    used_sheet_ids = set()
    for el in sheets_el:
        try:
            used_sheet_ids.add(int(el.get("sheetId") or 0))
        except ValueError:
            continue
    next_sheet_id = (max(used_sheet_ids) + 1) if used_sheet_ids else 1

    for snapshot in snapshots:
        new_part = _next_free_part(package, "xl/worksheets/sheet1.xml")
        sheet_xml = snapshot.sheet_xml
        has_tables = False

        if snapshot.rels_xml is not None:
            rels_root = _parse(snapshot.rels_xml)
            for rel in list(rels_root):
                rel_type = rel.get("Type")
                if rel_type == REL_TYPE_TABLE:
                    # 表格部件带唯一名称与 id，克隆分表不复制表格，同时删掉 tableParts。
                    rels_root.remove(rel)
                    has_tables = True
                    continue
                if rel_type not in REL_TYPES_SHEET_OWNED:
                    continue
                if str(rel.get("TargetMode") or "") == "External":
                    continue
                source_part = _resolve_part(snapshot.part, rel.get("Target") or "")
                data = snapshot.owned_parts.get(source_part)
                if data is None:
                    continue
                if freeze_all_anchors and rel_type == REL_TYPE_DRAWING:
                    data, frozen = _freeze_snapshot_drawing(snapshot, data)
                    frozen_total += frozen
                cloned_part = _clone_owned_part(
                    package,
                    content_types,
                    source_part,
                    data,
                    snapshot.owned_rels.get(source_part),
                )
                rel.set("Target", _relative_part(new_part, cloned_part))
            package.write(_rels_path(new_part), _serialize(rels_root))

        if has_tables:
            sheet_xml = _strip_table_parts(sheet_xml)
        package.write(new_part, sheet_xml)
        content_types.add_override(new_part, CT_WORKSHEET)

        rel_id = _next_rel_id(workbook_rels_root)
        _add_relationship(
            workbook_rels_root,
            rel_id,
            REL_TYPE_WORKSHEET,
            _relative_part(workbook_part, new_part),
        )

        title = _unique_sheet_title(f"{snapshot.name}_原文", used_names)
        used_names.add(title)
        sheet_el = etree.SubElement(sheets_el, _m("sheet"))
        sheet_el.set("name", title)
        sheet_el.set("sheetId", str(next_sheet_id))
        sheet_el.set(_r("id"), rel_id)
        next_sheet_id += 1

    return frozen_total


def _freeze_snapshot_drawing(
    snapshot: _SheetSnapshot, drawing_data: bytes
) -> tuple[bytes, int]:
    """按快照自己的几何冻结克隆分表的悬浮图片锚点。"""
    try:
        geometry = _SheetGeometry(_parse(snapshot.sheet_xml))
        drawing_root = _parse(drawing_data)
    except etree.XMLSyntaxError as error:  # pragma: no cover - 坏包才会走到
        logger.warning(f"克隆分表 {snapshot.name} 的 drawing 无法解析，原样照抄：{error}")
        return drawing_data, 0

    frozen = fix_drawing_anchors(drawing_root, geometry, set(), freeze_all=True)
    if not frozen:
        return drawing_data, 0
    return _serialize(drawing_root), frozen


def _unique_sheet_title(title: str, used_names: set[str]) -> str:
    """Excel 分表名上限 31 个字符，且必须唯一。"""
    candidate = title[:MAX_SHEET_TITLE_LEN]
    if candidate not in used_names:
        return candidate
    index = 2
    while True:
        suffix = f"_{index}"
        candidate = title[: MAX_SHEET_TITLE_LEN - len(suffix)] + suffix
        if candidate not in used_names:
            return candidate
        index += 1


def _clone_owned_part(
    package: _Package,
    content_types: _ContentTypes,
    source_part: str,
    data: bytes,
    rels_data: bytes | None,
) -> str:
    new_part = _next_free_part(package, source_part)
    package.write(new_part, data)
    if rels_data is not None:
        # drawing 的 rels 用 ../media/... 之类的相对路径，克隆件同目录，可直接照抄。
        package.write(_rels_path(new_part), rels_data)
    content_type = content_types.content_type_of(source_part)
    if content_type:
        content_types.add_override(new_part, content_type)
    elif not content_types.has_default(new_part):
        logger.warning(f"克隆部件缺少内容类型声明：{new_part}")
    return new_part


# ══════════════════════════════════════════════════════════════════════════════
# 入口
# ══════════════════════════════════════════════════════════════════════════════
def _load_sheet_entries(
    package: _Package,
    workbook_root,
    workbook_part: str,
    workbook_rels_root,
) -> list[_SheetEntry]:
    targets = {
        rel.get("Id"): _resolve_part(workbook_part, rel.get("Target") or "")
        for rel in workbook_rels_root
        if rel.get("Type") == REL_TYPE_WORKSHEET
    }
    sheets_el = workbook_root.find(_m("sheets"))
    entries: list[_SheetEntry] = []
    if sheets_el is None:
        return entries
    for el in sheets_el:
        if _local(el) != "sheet":
            continue
        rel_id = el.get(_r("id"))
        part = targets.get(rel_id)
        if not part or not package.has(part):
            continue
        entries.append(_SheetEntry(name=str(el.get("name") or ""), part=part))
    return entries


def _resolve_workbook_child_part(
    workbook_part: str,
    workbook_rels_root,
    rel_type_suffix: str,
    fallback: str,
) -> str:
    for rel in workbook_rels_root:
        if rel.get("Type") == f"{NS_REL}/{rel_type_suffix}":
            return _resolve_part(workbook_part, rel.get("Target") or "")
    return fallback


_MINIMAL_STYLES_XML = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    b'<styleSheet xmlns="' + NS_MAIN.encode() + b'">'
    b'<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
    b'<fills count="2"><fill><patternFill patternType="none"/></fill>'
    b'<fill><patternFill patternType="gray125"/></fill></fills>'
    b'<borders count="1"><border/></borders>'
    b'<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
    b'<cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>'
    b"</styleSheet>"
)


def _register_new_workbook_part(
    package: _Package,
    content_types: _ContentTypes,
    workbook_part: str,
    workbook_rels_root,
    *,
    part: str,
    rel_type: str,
    content_type: str,
    data: bytes,
) -> None:
    package.write(part, data)
    content_types.add_override(part, content_type)
    _add_relationship(
        workbook_rels_root,
        _next_rel_id(workbook_rels_root),
        rel_type,
        _relative_part(workbook_part, part),
    )


def _drop_calc_chain(
    package: _Package,
    content_types: _ContentTypes,
    workbook_part: str,
    workbook_rels_root,
) -> bool:
    """删除过公式后 calcChain 必然失效，直接丢弃让 Excel 重建。"""
    removed = False
    for rel in list(workbook_rels_root):
        if rel.get("Type") != REL_TYPE_CALC_CHAIN:
            continue
        part = _resolve_part(workbook_part, rel.get("Target") or "")
        package.remove(part)
        content_types.drop_override(part)
        workbook_rels_root.remove(rel)
        removed = True
    return removed


def write_bilingual_workbook(
    file_path: Path,
    *,
    translations: dict[str, str],
    target_lang: str,
    source_lang: str = "zh",
    keep_original_sheets: bool = False,
    formula_display_value_backfill: bool = False,
    lock_row_height: bool = False,
    review_marks: dict[str, str] | None = None,
    review_mark_colors: dict[str, str] | None = None,
    mark_review_items: bool = True,
    existing_fill_policy: str = EXCEL_REVIEW_EXISTING_FILL_POLICY_DEFAULT,
    log_callback=None,
    review_positions: list[dict[str, str]] | None = None,
    allowed_positions: dict[str, set[str]] | None = None,
    external_autofit_planned: bool = False,
    stats: dict[str, int] | None = None,
) -> None:
    """就地补丁式回填双语内容（``file_path`` 应当已是输出副本）。

    ``allowed_positions``：``None`` 表示不限制（默认，逐字节等价于旧行为）；
    传入 ``{分表名: {坐标, ...}}`` 时，只有落在该集合内的坐标才会被回填，
    未列出的分表视为坐标集合为空（该分表任何单元格都不会被改写）。
    供 ``core.excel_coverage`` 的按位置补译写入路径使用。

    ``external_autofit_planned``：调用方在本次写入之后还会用 Excel COM 对整表跑
    autofit（``enable_excel_autofit and not lock_row_height``）。为 True 时会把整张表
    的悬浮图片锚点全部冻结——Excel 那一刀重排的是整个 used range 的行高，只冻结我们
    自己改过的行挡不住它。这是**调用方**的决定，本模块不去猜。

    ``stats``：可选的输出统计字典，写入后会写入 ``stats["mutated_cells"]``
    （本次实际回填的单元格数）与 ``stats["anchor_frozen_count"]``
    （被冻结/改写的悬浮图片锚点数），供调用方生成日志与任务结果。
    ``anchor_frozen_count`` 只数译文分表，不含 ``keep_original_sheets`` 克隆出来的
    原文分表——那上面是同一批图片，两边都算会让面向用户的计数翻倍。
    """
    review_enabled = bool(mark_review_items)
    review_color_map = normalize_review_mark_colors(review_mark_colors)
    review_mark_map = (
        normalize_review_marks(review_marks, review_color_map) if review_enabled else {}
    )
    fill_policy = normalize_existing_fill_policy(existing_fill_policy)

    package = _Package(Path(file_path))
    try:
        root_rels_data = package.read("_rels/.rels")
        if root_rels_data is None:
            raise ValueError("不是有效的 xlsx 包：缺少 _rels/.rels")
        root_rels = _parse(root_rels_data)
        workbook_part = ""
        for rel in root_rels:
            if rel.get("Type") == REL_TYPE_OFFICE_DOCUMENT:
                workbook_part = _resolve_part("", rel.get("Target") or "")
                break
        if not workbook_part or not package.has(workbook_part):
            raise ValueError("不是有效的 xlsx 包：找不到 workbook 部件")

        content_types = _ContentTypes(package.read("[Content_Types].xml") or b"")
        workbook_root = _parse(package.read(workbook_part))
        workbook_rels_path = _rels_path(workbook_part)
        workbook_rels_data = package.read(workbook_rels_path)
        if workbook_rels_data is None:
            raise ValueError("不是有效的 xlsx 包：缺少 workbook 关系表")
        workbook_rels_root = _parse(workbook_rels_data)

        styles_part = _resolve_workbook_child_part(
            workbook_part, workbook_rels_root, "styles", "xl/styles.xml"
        )
        styles_data = package.read(styles_part)
        if styles_data is None:
            styles_data = _MINIMAL_STYLES_XML
            _register_new_workbook_part(
                package,
                content_types,
                workbook_part,
                workbook_rels_root,
                part=styles_part,
                rel_type=f"{NS_REL}/styles",
                content_type=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"
                ),
                data=styles_data,
            )
        styles = _Styles(styles_data)

        shared_strings = _load_shared_strings(
            package,
            _resolve_workbook_child_part(
                workbook_part,
                workbook_rels_root,
                "sharedStrings",
                "xl/sharedStrings.xml",
            ),
        )
        entries = _load_sheet_entries(
            package, workbook_root, workbook_part, workbook_rels_root
        )

        snapshots = (
            [_snapshot_sheet(package, entry) for entry in entries]
            if keep_original_sheets
            else []
        )

        removed_formula = False
        workbook_dirty = styles_data is _MINIMAL_STYLES_XML
        total_mutated_cells = 0
        total_anchor_frozen = 0
        for entry in entries:
            sheet_allowed = (
                allowed_positions.get(entry.name, set())
                if allowed_positions is not None
                else None
            )
            outcome = _process_sheet(
                package,
                entry,
                shared_strings=shared_strings,
                styles=styles,
                translations=translations,
                target_lang=target_lang,
                source_lang=source_lang,
                formula_display_value_backfill=formula_display_value_backfill,
                lock_row_height=lock_row_height,
                review_enabled=review_enabled,
                review_mark_map=review_mark_map,
                review_color_map=review_color_map,
                fill_policy=fill_policy,
                review_positions=review_positions,
                log_callback=log_callback,
                allowed_coordinates=sheet_allowed,
                external_autofit_planned=external_autofit_planned,
            )
            removed_formula = removed_formula or outcome.removed_formula
            total_mutated_cells += outcome.mutated_cells
            total_anchor_frozen += outcome.anchor_frozen_count

            if log_callback:
                if outcome.anchor_frozen_count:
                    log_callback(
                        f"[INFO] 分表已固定悬浮图片锚点：{entry.name}"
                        f"（{outcome.anchor_frozen_count} 处）"
                    )
                if lock_row_height and outcome.shrunk_cells:
                    log_callback(
                        f"[INFO] 分表已锁定行高并缩字号：{entry.name}"
                        f"（{outcome.shrunk_cells} 个单元格）"
                    )
                review_summary = ""
                if outcome.review_marked or outcome.review_skipped:
                    review_summary = f"（风险标记 {outcome.review_marked}"
                    if outcome.review_skipped:
                        review_summary += f"，保留原底色未改 {outcome.review_skipped}"
                    review_summary += "）"
                log_callback(f"[INFO] 分表已处理：{entry.name}{review_summary}")

        if keep_original_sheets and snapshots:
            cloned_frozen = _clone_original_sheets(
                package,
                content_types,
                workbook_root,
                workbook_part,
                workbook_rels_root,
                snapshots,
                # Excel 的 autofit 是对整本工作簿每张表跑的，克隆出来的原文分表
                # 也逃不掉，它上面的图片同样要先冻结。
                freeze_all_anchors=external_autofit_planned and not lock_row_height,
            )
            # 故意**不**计入 total_anchor_frozen：克隆件上的图片和译文分表上的是同一张，
            # 用户在原文件里只看得见一张。两边都算会让「N 张悬浮图片已固定尺寸」在开了
            # 「保留原文表」时凭空翻倍，那个数字就谁也对不上了。只记日志。
            if cloned_frozen and log_callback:
                log_callback(
                    f"[INFO] 原文分表克隆件已固定悬浮图片锚点：{cloned_frozen} 处"
                    "（与译文分表是同一批图片，不计入汇总）"
                )
            workbook_dirty = True

        if removed_formula and _drop_calc_chain(
            package, content_types, workbook_part, workbook_rels_root
        ):
            workbook_dirty = True

        if styles.dirty:
            package.write(styles_part, styles.serialize())
        if workbook_dirty:
            package.write(workbook_part, _serialize(workbook_root))
            package.write(workbook_rels_path, _serialize(workbook_rels_root))
        if content_types.dirty:
            package.write("[Content_Types].xml", content_types.serialize())

        if stats is not None:
            stats["mutated_cells"] = total_mutated_cells
            stats["anchor_frozen_count"] = total_anchor_frozen

        package.save()
    finally:
        package.close()
