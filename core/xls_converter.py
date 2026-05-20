"""
XLS 格式转换器模块。
提供将老的 .xls 格式转换为 .xlsx 格式的功能。
提供两种策略：
1. xlwings（优先复用本地 Excel 链路）：保真度更高，但运行前提取决于本机实际环境。
2. xlrd + openpyxl（纯 Python）：降级转换，可能丢失复杂格式（合并单元格样式、图片、宏等）。
"""
import tempfile
from pathlib import Path

from loguru import logger

from core.excel_automation import probe_local_excel_automation


class XlwingsUnavailableError(Exception):
    """当尝试使用 xlwings 但环境不可用时抛出。"""


def is_excel_automation_permission_denied(exc: BaseException | str) -> bool:
    """Return whether an Excel automation error is a macOS privacy denial."""
    text = str(exc or "").lower()
    return (
        "oserror: -1743" in text
        or "the user has declined permission" in text
        or "not authorized to send apple events" in text
        or "自动化权限" in text
    )


def _format_excel_conversion_error(exc: BaseException) -> str:
    message = str(exc)
    if not is_excel_automation_permission_denied(message):
        return f"使用 Excel 转换失败：{message}"

    return (
        "使用 Excel 转换失败：macOS 已拒绝 Translator 控制 Microsoft Excel 的自动化权限。"
        "可在「系统设置 > 隐私与安全性 > 自动化」中允许 Translator 控制 Microsoft Excel，"
        "或将 .xls 手动另存为 .xlsx 后再翻译。"
        f" 原始错误：{message}"
    )


def get_local_excel_availability() -> tuple[bool, str]:
    """检查当前环境是否真的可用本地 Excel 自动化。"""
    return probe_local_excel_automation()


def check_xlwings_available() -> bool:
    """兼容旧接口：返回当前环境是否可用本地 Excel 自动化。"""
    available, _reason = get_local_excel_availability()
    return available


def is_excel_installed() -> bool:
    """兼容旧接口：沿用真实可用性作为判断。"""
    return check_xlwings_available()


def _get_temp_xlsx_path(original_path: Path) -> Path:
    """在系统临时目录生成一个对应的 .xlsx 路径。"""
    temp_dir = Path(tempfile.gettempdir()) / "xl_translator_temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    
    # 避免重名冲突，加入文件大小或修改时间等哈希，这里简单用随机/自增
    import uuid
    safe_name = f"{original_path.stem}_{uuid.uuid4().hex[:6]}.xlsx"
    return temp_dir / safe_name


def convert_with_excel(app, xls_path: Path) -> Path:
    """
    使用由外部管理的 xlwings App 将 .xls 转换为 .xlsx。
    
    :param app: 外部传入的 xlwings.App 实例，实现全局进程复用。
    :param xls_path: 原 .xls 文件路径
    :return: 转换后的临时 .xlsx 文件路径
    """
    out_path = _get_temp_xlsx_path(xls_path)
    logger.info(f"使用 xlwings 将 {xls_path.name} 转换为临时 .xlsx")
    
    try:
        wb = app.books.open(str(xls_path))
        # 统一通过 xlwings 的跨平台 save() 触发 Save As，
        # 由目标扩展名 .xlsx 决定输出格式，避免写死 Windows COM 风格接口。
        wb.save(str(out_path))
        wb.close()
    except Exception as e:
        raise XlwingsUnavailableError(_format_excel_conversion_error(e)) from e
        
    return out_path


def convert_with_fallback(xls_path: Path) -> Path:
    """
    纯 Python 降级方案：使用 xlrd 读取，openpyxl 写出。
    
    :param xls_path: 原 .xls 文件路径
    :return: 转换后的临时 .xlsx 文件路径
    """
    import xlrd
    from openpyxl import Workbook

    out_path = _get_temp_xlsx_path(xls_path)
    logger.info(f"使用降级方案将 {xls_path.name} 转换为临时 .xlsx")

    wb_in = xlrd.open_workbook(str(xls_path), formatting_info=False)
    wb_out = Workbook()
    
    # 删除默认创建的第一个 sheet
    if wb_out.sheetnames:
        del wb_out[wb_out.sheetnames[0]]

    for sheet_idx in range(wb_in.nsheets):
        ws_in = wb_in.sheet_by_index(sheet_idx)
        # 防止重名限制 (最大31字符等 openpyxl 自身会校验，这里直接传递)
        ws_out = wb_out.create_sheet(title=ws_in.name)
        
        for rowx in range(ws_in.nrows):
            # openpyxl 行和列是从 1 开始的
            row_data = ws_in.row_values(rowx)
            for colx, value in enumerate(row_data):
                # 简单值拷贝，忽略单元格上的任何样式、合并和图片
                if value != "":
                    ws_out.cell(row=rowx + 1, column=colx + 1, value=value)

    wb_out.save(str(out_path))
    wb_out.close()
    wb_in.release_resources()
    
    return out_path
