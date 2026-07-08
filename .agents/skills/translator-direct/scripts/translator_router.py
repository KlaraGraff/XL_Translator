#!/usr/bin/env python3
"""Route a path to the Translator Excel, Word, or PDF headless CLI."""

from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
import sys
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any


EXCEL_SUFFIXES = {".xlsx", ".xls"}
WORD_SUFFIXES = {".docx", ".doc"}
PDF_SUFFIXES = {".pdf"}
IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}
GENERATED_OUTPUT_MARKER = "_翻译输出_"
TEXT_CONCURRENCY_DEFAULT = 10
PDF_PAGE_CONCURRENCY_DEFAULT = 3
OPENAI_IMAGE_MODEL_DEFAULT = "gpt-image-2"


@dataclass
class RoutePlan:
    route: str
    source_path: str
    file_count: int
    command: list[str]
    defaults: dict[str, Any] = field(default_factory=dict)


@dataclass
class MissingRequirement:
    route: str
    status: str
    message: str
    options: list[str]


def main(argv: list[str] | None = None) -> int:
    args = _parse_args(argv)
    repo_root = _find_repo_root(args.repo_root)
    if repo_root is None:
        return _emit(
            {
                "ok": False,
                "status": "missing_runtime",
                "message": "Translator repository not found.",
                "missing_requirements": [
                    {
                        "route": "runtime",
                        "status": "missing_repository",
                        "message": "Set TRANSLATOR_REPO_ROOT or run from the Translator repository.",
                        "options": [
                            "Run scripts/bootstrap.py --repo-root /path/to/repo.",
                            "Install the skill from inside a checked-out Translator repository.",
                        ],
                    }
                ],
            },
            args.json,
            exit_code=2,
        )

    source_path = Path(args.source).expanduser()
    if not source_path.is_absolute():
        source_path = source_path.resolve()
    groups = _scan_route_groups(source_path, include_images=args.include_images)
    if not any(groups.values()):
        return _emit(
            {
                "ok": False,
                "status": "no_supported_files",
                "source_path": str(source_path),
                "message": "No supported Excel, Word, or PDF files were found.",
            },
            args.json,
            exit_code=2,
        )

    source_lang = _resolve_source_lang(args, groups)
    current_config = _discover_current_agent_config()
    plans = _build_plans(
        repo_root=repo_root,
        source_path=source_path,
        groups=groups,
        args=args,
        source_lang=source_lang,
        current_config=current_config,
    )
    missing = _preflight(
        repo_root=repo_root,
        groups=groups,
        args=args,
        source_lang=source_lang,
        current_config=current_config,
    )

    dry_payload = {
        "ok": not missing,
        "status": "ready" if not missing else "missing_requirements",
        "source_path": str(source_path),
        "repo_root": str(repo_root),
        "source_lang": source_lang or "auto",
        "defaults": {
            "text_concurrency": args.text_concurrency,
            "pdf_page_concurrency": args.pdf_page_concurrency,
            "provider_preference": "current_agent_config",
        },
        "routes": [asdict(plan) for plan in plans],
        "missing_requirements": [asdict(item) for item in missing],
    }
    if args.dry_run:
        return _emit(dry_payload, args.json, exit_code=0 if not missing else 2)
    if missing and not args.force:
        return _emit(dry_payload, args.json, exit_code=2)

    results = []
    failed = False
    for plan in plans:
        completed = subprocess.run(
            plan.command,
            cwd=repo_root,
            check=False,
            text=True,
            capture_output=True,
        )
        parsed_stdout = _parse_json_stdout(completed.stdout)
        if completed.returncode != 0:
            failed = True
        results.append(
            {
                "route": plan.route,
                "returncode": completed.returncode,
                "stdout_json": parsed_stdout,
                "stdout": "" if parsed_stdout is not None else completed.stdout,
                "stderr": completed.stderr,
            }
        )

    return _emit(
        {
            "ok": not failed,
            "status": "completed" if not failed else "route_failed",
            "source_path": str(source_path),
            "source_lang": source_lang or "auto",
            "results": results,
        },
        args.json,
        exit_code=1 if failed else 0,
    )


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Route Translator work by source path.")
    parser.add_argument("source", help="File or folder path to translate.")
    parser.add_argument("--repo-root", help="Translator repository root.")
    parser.add_argument("--target-lang", help="Target language code or display name.")
    parser.add_argument("--source-lang", default="auto", help="Source language or auto.")
    parser.add_argument("--output-dir", help="Optional custom output root directory.")
    parser.add_argument("--include-images", action="store_true", help="Include standalone images in PDF route.")
    parser.add_argument("--allow-xls-fallback", action="store_true", help="Allow .xls fallback conversion.")
    parser.add_argument("--text-concurrency", type=int, default=TEXT_CONCURRENCY_DEFAULT)
    parser.add_argument("--pdf-page-concurrency", type=int, default=PDF_PAGE_CONCURRENCY_DEFAULT)
    parser.add_argument("--engine-mode", choices=("cloud", "local"), help="Text translation engine mode.")
    parser.add_argument("--cloud-provider", help="Text cloud provider override.")
    parser.add_argument("--cloud-model", help="Text cloud model override.")
    parser.add_argument("--cloud-base-url", help="Text cloud Base URL override.")
    parser.add_argument("--ollama-model", help="Local Ollama model override.")
    parser.add_argument("--image-provider", help="PDF image-generation provider override.")
    parser.add_argument("--image-model", help="PDF image-generation model override.")
    parser.add_argument("--image-base-url", help="PDF image-generation Base URL override.")
    parser.add_argument("--pdf-review", action="store_true", help="Enable PDF review model.")
    parser.add_argument("--review-provider", help="PDF review provider override.")
    parser.add_argument("--review-model", help="PDF review model override.")
    parser.add_argument("--review-base-url", help="PDF review Base URL override.")
    parser.add_argument("--dry-run", action="store_true", help="Only return route and preflight info.")
    parser.add_argument("--force", action="store_true", help="Run even when preflight reports missing config.")
    parser.add_argument("--json", action="store_true", help="Print JSON output.")
    return parser.parse_args(argv)


def _find_repo_root(raw_root: str | None) -> Path | None:
    candidates: list[Path] = []
    if raw_root:
        candidates.append(Path(raw_root).expanduser())
    env_root = str(os.environ.get("TRANSLATOR_REPO_ROOT") or "").strip()
    if env_root:
        candidates.append(Path(env_root).expanduser())
    here = Path(__file__).resolve()
    candidates.extend([Path.cwd(), *Path.cwd().parents, *here.parents])
    for candidate in candidates:
        root = candidate.resolve()
        if (
            (root / "scripts" / "translate_excel_cli.py").exists()
            and (root / "scripts" / "translate_word_cli.py").exists()
            and (root / "scripts" / "translate_pdf_cli.py").exists()
        ):
            return root
    return None


def _scan_route_groups(source_path: Path, *, include_images: bool) -> dict[str, list[Path]]:
    groups = {"excel": [], "word": [], "pdf": []}
    if not source_path.exists():
        return groups
    paths = [source_path] if source_path.is_file() else sorted(p for p in source_path.rglob("*") if p.is_file())
    for path in paths:
        if path.name.startswith(("~", ".")):
            continue
        if GENERATED_OUTPUT_MARKER in str(path):
            continue
        suffix = path.suffix.lower()
        if suffix in EXCEL_SUFFIXES:
            groups["excel"].append(path)
        elif suffix in WORD_SUFFIXES:
            groups["word"].append(path)
        elif suffix in PDF_SUFFIXES or (include_images and suffix in IMAGE_SUFFIXES):
            groups["pdf"].append(path)
    return groups


def _resolve_source_lang(args: argparse.Namespace, groups: dict[str, list[Path]]) -> str:
    requested = str(args.source_lang or "").strip()
    if requested and requested.casefold() != "auto":
        return requested
    samples: list[str] = []
    for path in [*groups.get("excel", [])[:3], *groups.get("word", [])[:3]]:
        text = _sample_text(path)
        if text:
            samples.append(text)
    return _detect_lang(" ".join(samples))


def _sample_text(path: Path) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(str(path), read_only=True, data_only=True)
            values: list[str] = []
            try:
                for sheet in workbook.worksheets[:2]:
                    for row in sheet.iter_rows(max_row=30, values_only=True):
                        for value in row:
                            if value is not None:
                                values.append(str(value))
                            if len(values) >= 80:
                                return " ".join(values)
            finally:
                workbook.close()
            return " ".join(values)
        if suffix == ".docx":
            return _sample_docx_text(path)
    except Exception:
        return ""
    return ""


def _sample_docx_text(path: Path) -> str:
    with zipfile.ZipFile(path) as package:
        with package.open("word/document.xml") as document:
            root = ET.fromstring(document.read())
    texts = []
    for node in root.iter():
        if node.tag.endswith("}t") and node.text:
            texts.append(node.text)
        if len(texts) >= 120:
            break
    return " ".join(texts)


def _detect_lang(text: str) -> str:
    if not text:
        return ""
    counts = {
        "zh": len(re.findall(r"[\u4e00-\u9fff]", text)),
        "ar": len(re.findall(r"[\u0600-\u06ff]", text)),
        "ru": len(re.findall(r"[\u0400-\u04ff]", text)),
        "en": len(re.findall(r"[A-Za-z]", text)),
    }
    code, count = max(counts.items(), key=lambda item: item[1])
    return code if count >= 8 else ""


def _discover_current_agent_config() -> dict[str, dict[str, str]]:
    custom_key = _first_env("TRANSLATOR_API_KEY", "OPENAI_COMPATIBLE_API_KEY", "CUSTOM_OPENAI_API_KEY")
    custom_base_url = _first_env("TRANSLATOR_BASE_URL", "OPENAI_COMPATIBLE_BASE_URL", "CUSTOM_OPENAI_BASE_URL")
    openai_key = _first_env("OPENAI_API_KEY")
    openai_base_url = _first_env("OPENAI_BASE_URL")
    text_model = _first_env("TRANSLATOR_TEXT_MODEL", "OPENAI_COMPATIBLE_MODEL", "OPENAI_MODEL")
    image_model = _first_env("TRANSLATOR_IMAGE_MODEL", "OPENAI_IMAGE_MODEL")
    review_model = _first_env("TRANSLATOR_PDF_REVIEW_MODEL", "OPENAI_VISION_MODEL", "OPENAI_MODEL")

    text: dict[str, str] = {}
    image: dict[str, str] = {}
    review: dict[str, str] = {}
    if custom_key or custom_base_url:
        text = {"provider": "custom_openai", "model": text_model, "base_url": custom_base_url}
        image = {
            "provider": "custom_openai",
            "model": image_model or OPENAI_IMAGE_MODEL_DEFAULT,
            "base_url": custom_base_url,
        }
        review = {"provider": "custom_openai", "model": review_model, "base_url": custom_base_url}
    elif openai_key or openai_base_url:
        text = {"provider": "openai", "model": text_model, "base_url": openai_base_url}
        image = {
            "provider": "openai",
            "model": image_model or OPENAI_IMAGE_MODEL_DEFAULT,
            "base_url": openai_base_url,
        }
        review = {"provider": "openai", "model": review_model, "base_url": openai_base_url}
    elif _first_env("ANTHROPIC_API_KEY", "CLAUDE_API_KEY"):
        text = {"provider": "claude", "model": _first_env("ANTHROPIC_MODEL", "CLAUDE_MODEL"), "base_url": ""}
    elif _first_env("DASHSCOPE_API_KEY"):
        text = {"provider": "dashscope", "model": _first_env("DASHSCOPE_MODEL"), "base_url": ""}
    elif _first_env("ZHIPUAI_API_KEY", "ZHIPU_API_KEY"):
        text = {"provider": "zhipu", "model": _first_env("ZHIPU_MODEL"), "base_url": ""}
    elif _first_env("SILICONFLOW_API_KEY"):
        model = _first_env("SILICONFLOW_MODEL")
        text = {"provider": "siliconflow", "model": model, "base_url": _first_env("SILICONFLOW_BASE_URL")}
        image = {
            "provider": "siliconflow",
            "model": _first_env("SILICONFLOW_IMAGE_MODEL") or model,
            "base_url": _first_env("SILICONFLOW_BASE_URL"),
        }
        review = {
            "provider": "siliconflow",
            "model": _first_env("SILICONFLOW_VISION_MODEL") or model,
            "base_url": _first_env("SILICONFLOW_BASE_URL"),
        }
    return {"text": text, "image": image, "review": review}


def _first_env(*names: str) -> str:
    for name in names:
        value = str(os.environ.get(name) or "").strip()
        if value:
            return value
    return ""


def _build_plans(
    *,
    repo_root: Path,
    source_path: Path,
    groups: dict[str, list[Path]],
    args: argparse.Namespace,
    source_lang: str,
    current_config: dict[str, dict[str, str]],
) -> list[RoutePlan]:
    python_path = _project_python(repo_root)
    plans: list[RoutePlan] = []
    if groups["excel"]:
        command = _text_command(
            python_path,
            repo_root / "scripts" / "translate_excel_cli.py",
            source_path,
            args,
            source_lang,
            current_config.get("text") or {},
        )
        if args.allow_xls_fallback:
            command.append("--allow-xls-fallback")
        plans.append(RoutePlan("excel", str(source_path), len(groups["excel"]), command))
    if groups["word"]:
        command = _text_command(
            python_path,
            repo_root / "scripts" / "translate_word_cli.py",
            source_path,
            args,
            source_lang,
            current_config.get("text") or {},
        )
        plans.append(RoutePlan("word", str(source_path), len(groups["word"]), command))
    if groups["pdf"]:
        command = [
            str(python_path),
            str(repo_root / "scripts" / "translate_pdf_cli.py"),
            str(source_path),
            "--quiet",
            "--json",
            "--pdf-page-concurrency",
            str(args.pdf_page_concurrency),
            "--no-pdf-review" if not args.pdf_review else "--pdf-review",
        ]
        if args.target_lang:
            command.extend(["--target-lang", args.target_lang])
        if args.output_dir:
            command.extend(["--output-dir", args.output_dir])
        if args.include_images:
            command.append("--include-images")
        image_config = _merged_route_config(args, "image", current_config.get("image") or {})
        _extend_model_args(command, "image", image_config)
        if args.pdf_review:
            review_config = _merged_route_config(args, "review", current_config.get("review") or {})
            _extend_model_args(command, "review", review_config)
        plans.append(
            RoutePlan(
                "pdf",
                str(source_path),
                len(groups["pdf"]),
                command,
                defaults={"pdf_page_concurrency": args.pdf_page_concurrency},
            )
        )
    return plans


def _text_command(
    python_path: Path,
    script_path: Path,
    source_path: Path,
    args: argparse.Namespace,
    source_lang: str,
    current_text_config: dict[str, str],
) -> list[str]:
    command = [
        str(python_path),
        str(script_path),
        str(source_path),
        "--quiet",
        "--json",
        "--concurrency",
        str(args.text_concurrency),
    ]
    if args.target_lang:
        command.extend(["--target-lang", args.target_lang])
    if source_lang:
        command.extend(["--source-lang", source_lang])
    if args.output_dir:
        command.extend(["--output-dir", args.output_dir])
    if args.engine_mode:
        command.extend(["--engine-mode", args.engine_mode])
    if args.ollama_model:
        command.extend(["--ollama-model", args.ollama_model])
    text_config = _merged_route_config(args, "cloud", current_text_config)
    if text_config.get("provider"):
        command.extend(["--cloud-provider", text_config["provider"]])
    if text_config.get("model"):
        command.extend(["--cloud-model", text_config["model"]])
    if text_config.get("base_url"):
        command.extend(["--cloud-base-url", text_config["base_url"]])
    return command


def _merged_route_config(
    args: argparse.Namespace,
    prefix: str,
    current_config: dict[str, str],
) -> dict[str, str]:
    if prefix == "cloud":
        provider = args.cloud_provider
        model = args.cloud_model
        base_url = args.cloud_base_url
    elif prefix == "image":
        provider = args.image_provider
        model = args.image_model
        base_url = args.image_base_url
    else:
        provider = args.review_provider
        model = args.review_model
        base_url = args.review_base_url
    return {
        "provider": str(provider or current_config.get("provider") or "").strip(),
        "model": str(model or current_config.get("model") or "").strip(),
        "base_url": str(base_url or current_config.get("base_url") or "").strip(),
    }


def _extend_model_args(command: list[str], prefix: str, config: dict[str, str]) -> None:
    if config.get("provider"):
        command.extend([f"--{prefix}-provider", config["provider"]])
    if config.get("model"):
        command.extend([f"--{prefix}-model", config["model"]])
    if config.get("base_url"):
        command.extend([f"--{prefix}-base-url", config["base_url"]])


def _preflight(
    *,
    repo_root: Path,
    groups: dict[str, list[Path]],
    args: argparse.Namespace,
    source_lang: str,
    current_config: dict[str, dict[str, str]],
) -> list[MissingRequirement]:
    sys.path.insert(0, str(repo_root))
    from core.api_config_check import check_translation_api_config
    from core.headless_pdf_translate import build_pdf_runtime_settings
    from core.model_roles import (
        ROLE_IMAGE,
        ROLE_PDF_REVIEW,
        SOURCE_INDEPENDENT,
        provider_supports_capability,
        resolve_effective_model_config,
    )
    from settings import load_settings, set_cloud_provider_config

    missing: list[MissingRequirement] = []
    if groups["excel"] or groups["word"]:
        text_settings = load_settings().model_copy(deep=True)
        if source_lang:
            text_settings.source_lang = source_lang
        text_settings.engine.concurrency = args.text_concurrency
        if args.engine_mode:
            text_settings.engine.mode = args.engine_mode
        if args.ollama_model:
            text_settings.engine.ollama_model = args.ollama_model
            text_settings.engine.local_model = args.ollama_model
        text_config = _merged_route_config(args, "cloud", current_config.get("text") or {})
        if text_config.get("provider"):
            text_settings.engine.cloud_provider = text_config["provider"]
            if not text_config.get("base_url"):
                text_settings.engine.cloud_base_url = ""
        if text_config.get("model"):
            text_settings.engine.cloud_model = text_config["model"]
        if text_config.get("base_url"):
            text_settings.engine.cloud_base_url = text_config["base_url"]
        if text_config.get("provider") or text_config.get("model") or text_config.get("base_url"):
            set_cloud_provider_config(
                text_settings.engine,
                text_settings.engine.cloud_provider,
                cloud_model=text_settings.engine.cloud_model,
                cloud_base_url=text_settings.engine.cloud_base_url,
            )
        check = check_translation_api_config(text_settings)
        if not check.ok:
            missing.append(_missing_text_requirement(check.status, check.message))

    if groups["pdf"]:
        pdf_settings = build_pdf_runtime_settings(
            base_settings=None,
            target_lang=args.target_lang,
            page_concurrency=args.pdf_page_concurrency,
            review_enabled=args.pdf_review,
        )
        image_config = _merged_route_config(args, "image", current_config.get("image") or {})
        if image_config.get("provider"):
            pdf_settings.image_model_role.source_role = SOURCE_INDEPENDENT
            pdf_settings.image_model_role.cloud_provider = image_config["provider"]
            if not image_config.get("base_url"):
                pdf_settings.image_model_role.cloud_base_url = ""
        if image_config.get("model"):
            pdf_settings.image_model_role.source_role = SOURCE_INDEPENDENT
            pdf_settings.image_model_role.cloud_model = image_config["model"]
        if image_config.get("base_url"):
            pdf_settings.image_model_role.source_role = SOURCE_INDEPENDENT
            pdf_settings.image_model_role.cloud_base_url = image_config["base_url"]
        if image_config.get("provider") or image_config.get("model") or image_config.get("base_url"):
            set_cloud_provider_config(
                pdf_settings.image_model_role,
                pdf_settings.image_model_role.cloud_provider,
                cloud_model=pdf_settings.image_model_role.cloud_model,
                cloud_base_url=pdf_settings.image_model_role.cloud_base_url,
            )
        image_model = resolve_effective_model_config(pdf_settings, ROLE_IMAGE)
        if not provider_supports_capability(image_model.provider, "image"):
            missing.append(_missing_pdf_requirement("unsupported_image_provider", image_model.provider))
        elif not image_model.model:
            missing.append(_missing_pdf_requirement("missing_image_model", image_model.provider))
        elif not image_model.api_key:
            missing.append(_missing_pdf_requirement("missing_image_api_key", image_model.provider))

        if args.pdf_review:
            review_config = _merged_route_config(args, "review", current_config.get("review") or {})
            if review_config.get("provider"):
                pdf_settings.pdf_review_model_role.source_role = SOURCE_INDEPENDENT
                pdf_settings.pdf_review_model_role.cloud_provider = review_config["provider"]
                if not review_config.get("base_url"):
                    pdf_settings.pdf_review_model_role.cloud_base_url = ""
            if review_config.get("model"):
                pdf_settings.pdf_review_model_role.source_role = SOURCE_INDEPENDENT
                pdf_settings.pdf_review_model_role.cloud_model = review_config["model"]
            if review_config.get("base_url"):
                pdf_settings.pdf_review_model_role.source_role = SOURCE_INDEPENDENT
                pdf_settings.pdf_review_model_role.cloud_base_url = review_config["base_url"]
            if review_config.get("provider") or review_config.get("model") or review_config.get("base_url"):
                set_cloud_provider_config(
                    pdf_settings.pdf_review_model_role,
                    pdf_settings.pdf_review_model_role.cloud_provider,
                    cloud_model=pdf_settings.pdf_review_model_role.cloud_model,
                    cloud_base_url=pdf_settings.pdf_review_model_role.cloud_base_url,
                )
            review_model = resolve_effective_model_config(pdf_settings, ROLE_PDF_REVIEW)
            if not provider_supports_capability(review_model.provider, "vision_text"):
                missing.append(_missing_review_requirement("unsupported_review_provider", review_model.provider))
            elif not review_model.model:
                missing.append(_missing_review_requirement("missing_review_model", review_model.provider))
            elif not review_model.api_key:
                missing.append(_missing_review_requirement("missing_review_api_key", review_model.provider))
    return missing


def _missing_text_requirement(status: str, message: str) -> MissingRequirement:
    return MissingRequirement(
        route="text",
        status=status,
        message=message,
        options=[
            "Use the current agent API config via OPENAI_API_KEY/OPENAI_MODEL or provider-specific env vars.",
            "Provide a temporary API key/model for this run.",
            "Save provider credentials in Translator settings/keys.json.",
            "Switch to a configured local model with --engine-mode local --ollama-model <model>.",
        ],
    )


def _missing_pdf_requirement(status: str, provider: str) -> MissingRequirement:
    return MissingRequirement(
        route="pdf",
        status=status,
        message=f"PDF image translation needs an image-generation model; current provider: {provider or 'unset'}.",
        options=[
            "Use the current agent OpenAI config; default image model is gpt-image-2 when OPENAI_API_KEY is available.",
            "Provide --image-provider, --image-model, and optionally --image-base-url for this run.",
            "Save an image-generation model role in Translator settings.",
            "Run only Excel/Word routes if the available key only supports text generation.",
        ],
    )


def _missing_review_requirement(status: str, provider: str) -> MissingRequirement:
    return MissingRequirement(
        route="pdf_review",
        status=status,
        message=f"PDF review needs a vision-text model; current provider: {provider or 'unset'}.",
        options=[
            "Disable PDF review for this run.",
            "Provide --review-provider, --review-model, and optionally --review-base-url.",
            "Save a PDF review model role in Translator settings.",
        ],
    )


def _project_python(repo_root: Path) -> Path:
    unix_python = repo_root / ".venv" / "bin" / "python3"
    if unix_python.exists():
        return unix_python
    win_python = repo_root / ".venv" / "Scripts" / "python.exe"
    if win_python.exists():
        return win_python
    return Path(sys.executable)


def _parse_json_stdout(stdout: str) -> Any:
    text = str(stdout or "").strip()
    if not text:
        return None
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        return None


def _emit(payload: dict[str, Any], as_json: bool, *, exit_code: int) -> int:
    if as_json:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
    elif payload.get("ok"):
        print(f"{payload.get('status', 'ok')}: {payload.get('source_path', '')}")
    else:
        print(json.dumps(payload, ensure_ascii=False, indent=2))
    return exit_code


if __name__ == "__main__":
    raise SystemExit(main())
